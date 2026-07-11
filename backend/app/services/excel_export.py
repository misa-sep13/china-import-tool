import io
import re
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TAOTARO_SPEC_LABEL_RE = re.compile(r"(颜色|顏色|规格|規格|尺码|尺寸|款式)\s*[：:]\s*")
GRIP_TRAINER_SPEC_RE = re.compile(r"握笔器六代【[^】]*彩盒装】")


def normalize_taotaro_spec(spec) -> str:
    """タオタロウのExcel取込で読める仕様文字に整える。"""
    text = "" if spec is None else str(spec).strip()
    if not text:
        return ""

    text = text.replace("；", "、").replace(";", "、")
    text = TAOTARO_SPEC_LABEL_RE.sub("", text)

    parts = [part.strip(" 、,，") for part in re.split(r"[、,，]+", text)]
    parts = [part for part in parts if part and part not in {"无规格", "無規格"}]
    text = "、".join(parts) if parts else ""

    # 1688の選択肢名と商品マスタ側の表記ゆれを吸収する。
    text = text.replace("握笔器第六代", "握笔器六代")
    text = text.replace("握笔器六代代", "握笔器六代")

    if GRIP_TRAINER_SPEC_RE.fullmatch(text):
        return f"颜色：{text}；规格：无规格；"

    return text.strip(" 、,，")


def build_taotaro_excel(items: List[Dict]) -> bytes:
    """タオタロウインポート用Excelを生成してbytesで返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 1行目：タイトル（A1を結合して「導入例」）
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "導入例"
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.row_dimensions[1].height = 20

    # 2行目：ヘッダー
    # A: 発注先URL（「↓※発注先URLをここに入れる」付き）
    headers = [
        "発注先URL　↓※発注先URLをここに入れる",  # A
        "仕様",       # B
        "数量",       # C
        "単価",       # D
        "ASIN",       # E
        "FNSKU",      # F
        "お客様専用メモ",  # G
        "備考",       # H
    ]
    header_fill = PatternFill("solid", fgColor="E2EFDA")
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 25

    # 3行目以降：データ
    for i, item in enumerate(items):
        row_num = 3 + i
        spec = item.get("spec") or "　".join(filter(None, [item.get("color", ""), item.get("size", "")]))
        spec = normalize_taotaro_spec(spec)
        values = [
            item.get("buy_url", ""),       # A: 発注先URL
            spec,                           # B: 仕様
            item.get("qty", 0),             # C: 数量
            item.get("price", 0),           # D: 単価
            item.get("asin", ""),           # E: ASIN
            item.get("fnsku", ""),          # F: FNSKU
            item.get("customer_memo", ""),  # G: お客様専用メモ
            item.get("note", ""),           # H: 備考
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1 and val:
                cell.font = Font(color="0563C1", underline="single")

    # 列幅調整
    col_widths = {1: 45, 2: 22, 3: 8, 4: 8, 5: 14, 6: 14, 7: 25, 8: 20}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_rakuten_taotaro_excel(items: List[Dict]) -> bytes:
    """楽天版 タオタロウインポート用Excel（Amazonと同じ8列形式・ASIN/FNSKUは空欄）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 1行目：ヘッダー（タオタロウの取り込みは1行目をヘッダーとして読み飛ばし、
    # 2行目からデータとして読む。タイトル行を入れるとヘッダー行が商品として
    # 誤読される（公式テンプレートの導入内容シートも1行目ヘッダー・2行目データ））
    headers = [
        "発注先URL　↓※発注先URLをここに入れる",  # A
        "仕様",           # B
        "数量",           # C
        "単価",           # D
        "ASIN",           # E（楽天では空欄）
        "FNSKU",          # F（楽天では空欄）
        "お客様専用メモ",  # G
        "備考",           # H
    ]
    header_fill = PatternFill("solid", fgColor="E2EFDA")
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 25

    # 2行目以降：データ（ASIN/FNSKUは空欄、customer_memoがあればASIN欄に出力）
    for i, item in enumerate(items):
        row_num = 2 + i
        supplier_spec = normalize_taotaro_spec(item.get("supplier_spec", "") or item.get("spec", ""))
        values = [
            item.get("buy_url", ""),        # A: 発注先URL
            supplier_spec,                   # B: 仕様（中国語優先）
            item.get("qty", 0),             # C: 数量
            item.get("price", 0),           # D: 単価
            "",                             # E: ASIN（空欄）
            "",                             # F: FNSKU（空欄）
            item.get("customer_memo", ""),  # G: お客様専用メモ
            item.get("notes", ""),          # H: 備考
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1 and val:
                cell.font = Font(color="0563C1", underline="single")

    # 列幅調整（Amazonと同じ）
    col_widths = {1: 45, 2: 22, 3: 8, 4: 8, 5: 14, 6: 14, 7: 25, 8: 20}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
