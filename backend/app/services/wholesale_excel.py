"""発注書のExcelを作る。

取引先が見慣れた様式をそのまま使う。書式を手で組み直すと罫線や
列幅が微妙にずれるので、実際に送っていたファイルから中身だけ
抜いたものをテンプレートにして、値を差し込む。

金額の計算はExcelの数式に任せる（元ファイルと同じ）。ただし
メール本文や画面に出す合計はPython側でも同じ式で出しておく。
openpyxlは数式を計算しないので、保存しただけでは値が読めないため。
"""
import io
import os
from datetime import date, datetime, timedelta, timezone

import openpyxl

# サーバーはUTCで動く。そのまま date.today() を使うと日本の朝9時前が
# 前日になり、発注書の日付がずれる
JST = timezone(timedelta(hours=9))


def today_jst():
    return datetime.now(JST).date()

TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "templates", "発注書テンプレート.xlsx")

ROW_START = 15          # 明細の開始行
ROW_END = 42            # 明細の終了行（テンプレートの罫線がここまで）
TAX_RATE = 0.1


def calc_totals(items):
    """小計・消費税・合計。元のExcelと同じ出し方をする。

    合計は ROUNDDOWN(小計+消費税, 0)。円未満は切り捨て。
    """
    subtotal = sum((i.get("unit_price") or 0) * (i.get("qty") or 0) for i in items)
    tax = subtotal * TAX_RATE
    total = int(subtotal + tax)          # ROUNDDOWN(x, 0) と同じ
    return subtotal, tax, total


def build(supplier, order, items):
    """発注書のExcelを作ってバイト列で返す。

    supplier: {name, honorific}
    order   : {order_date, order_no, subject, delivery_date,
               deliver_zip, deliver_address, deliver_note, payment_terms}
    items   : [{item_code, jan_code, name, unit_price, qty, note}]
    """
    if len(items) > ROW_END - ROW_START + 1:
        raise ValueError(
            f"1枚に書ける明細は{ROW_END - ROW_START + 1}件までです"
            f"（今回{len(items)}件）。分けて発注してください")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["発注書"]

    # 宛先
    ws["A2"] = f"{supplier['name']}　{supplier.get('honorific') or '御中'}"

    # 発注日。テンプレートは =TODAY() だが、あとで開き直したときに
    # 日付が変わってしまうので、実際の発注日を値で入れる
    d = order.get("order_date")
    ws["F3"] = date.fromisoformat(d) if isinstance(d, str) and d else today_jst()

    if order.get("order_no"):
        ws["F2"] = order["order_no"]
    if order.get("subject"):
        ws["B6"] = order["subject"]
    if order.get("delivery_date"):
        ws["B7"] = order["delivery_date"]

    # 納品場所。郵便番号と住所は続けて1つのセルに入っていた
    zip_, addr = order.get("deliver_zip") or "", order.get("deliver_address") or ""
    if zip_ or addr:
        ws["B8"] = f"{zip_}　{addr}".strip()
    if order.get("deliver_note"):
        ws["C9"] = order["deliver_note"]
    if order.get("payment_terms"):
        ws["B9"] = order["payment_terms"]

    # 明細
    for n, it in enumerate(items):
        r = ROW_START + n
        ws.cell(r, 1).value = it.get("item_code") or None
        ws.cell(r, 2).value = it.get("jan_code") or None
        ws.cell(r, 3).value = it.get("name") or None
        ws.cell(r, 4).value = it.get("unit_price") or None
        ws.cell(r, 5).value = it.get("qty") or None
        ws.cell(r, 7).value = it.get("note") or None
        # F列（金額）はテンプレートの数式のまま

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def file_name(supplier_name, order_date):
    """送るときのファイル名。

    今までは「美園工芸社発注書2026.8.24.xlsx」の形だったので、
    取引先が受け取ったときに違和感がないよう合わせる。
    """
    d = order_date or today_jst().isoformat()
    y, m, dd = d.split("-")
    return f"美園工芸社発注書{y}.{int(m)}.{int(dd)}.xlsx"
