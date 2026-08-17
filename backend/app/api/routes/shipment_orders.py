from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from app.core.database import get_db
from app.models.shipment_order import ShipmentOrder, ShipmentOrderItem
from app.models.product import Product
from app.models.rakuten_product import RakutenProduct
from app.models.rakuten_order import RakutenOrderHistory
from app.models.inventory_reflection_log import InventoryReflectionLog
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import asyncio
import io, uuid

router = APIRouter(prefix="/shipment-orders", tags=["shipment_orders"])


def _url_match_key(url: str) -> str:
    value = (url or "").strip().lower()
    if not value:
        return ""
    import re
    m = re.search(r"offer/(\d+)\.html", value)
    if m:
        return f"1688:{m.group(1)}"
    m = re.search(r"[?&]id=(\d+)", value)
    if m:
        return f"id:{m.group(1)}"
    return value.split("?")[0].rstrip("/")


def _norm_text(value: str) -> str:
    return (value or "").strip().replace(" ", "").replace("　", "")


class ShipmentOrderItemPatch(BaseModel):
    product_id: int


@router.get("/debug-match-score")
def debug_match_score(buy_url: str, color: str = "", size: str = "", unit_price_cny: float = 0, qty: int = 0, db: Session = Depends(get_db)):
    """デバッグ用: 指定した仕入URLに紐づく全候補のスコア内訳を返す（一時的な調査用）"""
    rakuten_products = db.query(RakutenProduct).filter(
        RakutenProduct.buy_url.isnot(None), RakutenProduct.is_active == True
    ).all()
    pending_rows = (
        db.query(RakutenOrderHistory.sku, sqlfunc.sum(RakutenOrderHistory.qty))
        .filter(RakutenOrderHistory.is_deleted == False, RakutenOrderHistory.is_delivered == False)
        .group_by(RakutenOrderHistory.sku)
        .all()
    )
    pending_by_sku = {sku: qty2 or 0 for sku, qty2 in pending_rows}
    url_key_counts = {}
    for p in rakuten_products:
        key = _url_match_key(p.buy_url or "")
        if key:
            url_key_counts[key] = url_key_counts.get(key, 0) + 1

    item = {"buy_url": buy_url, "color": color, "size": size, "unit_price_cny": unit_price_cny, "qty": qty}
    target_key = _url_match_key(buy_url)
    results = []
    for p in rakuten_products:
        pk = _url_match_key(p.buy_url or "")
        if pk != target_key:
            continue
        detail = {}
        score = 0
        item_key = _url_match_key(item.get("buy_url", ""))
        product_key = pk
        detail["url_match"] = bool(item_key and product_key and item_key == product_key)
        if detail["url_match"]:
            score += 45
            detail["url_unique_bonus"] = url_key_counts.get(product_key, 0) == 1
            if detail["url_unique_bonus"]:
                score += 10
        color_n = _norm_text(item.get("color", ""))
        size_n = _norm_text(item.get("size", ""))
        spec_n = _norm_text(p.supplier_spec or "")
        combo1 = _norm_text(f"{item.get('color', '')}、{item.get('size', '')}")
        combo2 = _norm_text(f"{item.get('color', '')} {item.get('size', '')}")
        detail["spec"] = p.supplier_spec
        detail["spec_norm"] = spec_n
        detail["color_norm"] = color_n
        detail["combo1"] = combo1
        detail["combo2"] = combo2
        if spec_n and color_n and spec_n == color_n:
            score += 35
            detail["spec_branch"] = "exact"
        elif spec_n and color_n and size_n and spec_n in {combo1, combo2}:
            score += 35
            detail["spec_branch"] = "combo"
        elif spec_n and color_n and (spec_n in color_n or color_n in spec_n):
            score += 18
            detail["spec_branch"] = "partial"
        else:
            detail["spec_branch"] = None
        try:
            ip = float(item.get("unit_price_cny") or 0)
            pp = float(p.price or 0)
            detail["price_match"] = ip > 0 and abs(ip - pp) < 0.011
            if detail["price_match"]:
                score += 15
        except Exception:
            detail["price_match"] = "error"
        results.append({"sku": p.sku, "id": p.id, "score": score, "detail": detail})
    results.sort(key=lambda x: -x["score"])
    return {"item": item, "url_key": target_key, "candidates": results}


@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel読み込みエラー: {str(e)}")

    ws = wb.active

    # ヘッダー情報の取得（行/列は固定レイアウトに依存）
    def find_cell(keyword, max_row=10):
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
            for i, cell in enumerate(row):
                if cell and keyword in str(cell):
                    return row, i
        return None, -1

    shipped_date = ""
    tracking_no = ""
    order_no = ""
    box_count = 0
    total_weight_kg = 0.0

    # 固定位置から値を取得（send-order-listのレイアウト）
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        row_vals = [str(c) if c is not None else "" for c in row]
        joined = "".join(row_vals)
        if "出荷日" in joined:
            # 出荷日は同じ行の次のセルにある
            for i, c in enumerate(row):
                if c and "出荷日" in str(c) and i + 1 < len(row):
                    shipped_date = str(row[i + 1] or "")[:10]
        if "配送依赖No" in joined or "配送依頼No" in joined:
            for i, c in enumerate(row):
                if c and ("配送依赖No" in str(c) or "配送依頼No" in str(c)) and i + 1 < len(row):
                    order_no = str(row[i + 1] or "")
        if "追跡番号" in joined:
            for i, c in enumerate(row):
                if c and "追跡番号" in str(c) and i + 1 < len(row):
                    tracking_no = str(row[i + 1] or "")
        if "実際重量" in joined:
            for i, c in enumerate(row):
                if c and "実際重量" in str(c) and i + 1 < len(row):
                    try:
                        total_weight_kg = float(row[i + 1] or 0)
                    except Exception:
                        pass
        if "箱数" in joined:
            for i, c in enumerate(row):
                if c and "箱数" in str(c) and i + 1 < len(row):
                    try:
                        box_count = int(row[i + 1] or 0)
                    except Exception:
                        pass

    # 商品行のヘッダーを探す（発注時間・商品名・色・サイズ・URL・単価・数量）
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_strs = [str(c) if c else "" for c in row]
        if "商品名" in row_strs and "数量" in row_strs:
            header_row_idx = i
            for ci, c in enumerate(row_strs):
                col_map[c] = ci
            break

    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="商品データ行が見つかりません")

    col_name = col_map.get("商品名", -1)
    col_color = col_map.get("色", -1)
    col_size = col_map.get("サイズ", -1)
    col_url = col_map.get("商品URL", -1)
    col_price = col_map.get("単価", -1)
    col_qty = col_map.get("数量", -1)

    items_raw = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(c is None for c in row):
            continue
        name_cn = str(row[col_name] or "") if col_name >= 0 else ""
        if not name_cn:
            continue
        buy_url = str(row[col_url] or "") if col_url >= 0 else ""
        items_raw.append({
            "name_cn": name_cn,
            "color": str(row[col_color] or "") if col_color >= 0 else "",
            "size": str(row[col_size] or "") if col_size >= 0 else "",
            "buy_url": buy_url,
            "unit_price_cny": float(row[col_price] or 0) if col_price >= 0 else 0,
            "qty": int(row[col_qty] or 0) if col_qty >= 0 else 0,
        })

    return {
        "shipped_date": shipped_date,
        "tracking_no": tracking_no,
        "order_no": order_no,
        "box_count": box_count,
        "total_weight_kg": total_weight_kg,
        "items": items_raw,
    }


@router.post("/match")
def match_products(items: List[dict], db: Session = Depends(get_db)):
    """配送依頼明細を楽天商品マスタと照合して照合結果を返す"""
    rakuten_products = db.query(RakutenProduct).filter(
        RakutenProduct.buy_url.isnot(None), RakutenProduct.is_active == True
    ).all()
    pending_rows = (
        db.query(RakutenOrderHistory.sku, sqlfunc.sum(RakutenOrderHistory.qty))
        .filter(RakutenOrderHistory.is_deleted == False, RakutenOrderHistory.is_delivered == False)
        .group_by(RakutenOrderHistory.sku)
        .all()
    )
    pending_by_sku = {sku: qty or 0 for sku, qty in pending_rows}

    url_key_counts = {}
    for p in rakuten_products:
        key = _url_match_key(p.buy_url or "")
        if key:
            url_key_counts[key] = url_key_counts.get(key, 0) + 1

    def score_product(item: dict, product: RakutenProduct) -> int:
        score = 0
        item_key = _url_match_key(item.get("buy_url", ""))
        product_key = _url_match_key(product.buy_url or "")
        if item_key and product_key and item_key == product_key:
            score += 45
            if url_key_counts.get(product_key, 0) == 1:
                score += 10

        color = _norm_text(item.get("color", ""))
        size = _norm_text(item.get("size", ""))
        spec = _norm_text(product.supplier_spec or "")
        # 1688仕入元によっては色/仕様の情報が「色」欄ではなく「サイズ」欄に入ることがある
        # （例: 色欄が空でサイズ欄に"蓝色12粒"のように色名込みで入っている）。
        # 色欄だけで判定すると常に不一致になり単価だけが決め手になってしまうため、
        # サイズ欄単独でもspecと突き合わせる。
        if spec and color and spec == color:
            score += 35
        elif spec and size and spec == size:
            score += 35
        elif spec and color and size and spec in {
            _norm_text(f"{item.get('color', '')}、{item.get('size', '')}"),
            _norm_text(f"{item.get('color', '')} {item.get('size', '')}"),
        }:
            score += 35
        elif spec and color and (spec in color or color in spec):
            score += 18
        elif spec and size and (spec in size or size in spec):
            score += 18

        try:
            item_price = float(item.get("unit_price_cny") or 0)
            product_price = float(product.price or 0)
            if item_price > 0 and abs(item_price - product_price) < 0.011:
                score += 15
        except Exception:
            pass
        try:
            set_size = product.set_size or 1
            received_qty = int(item.get("qty") or 0) // set_size if set_size > 1 else int(item.get("qty") or 0)
            pending_qty = (pending_by_sku.get(product.sku, 0) or 0) + (product.inbound or 0) + (product.standard_stock or 0)
            if received_qty > 0 and pending_qty == received_qty:
                score += 25
        except Exception:
            pass
        return score

    matched = []
    unmatched = []
    for item in items:
        candidates = sorted(
            ((score_product(item, p), p) for p in rakuten_products),
            key=lambda x: x[0],
            reverse=True,
        )
        best_score, product = candidates[0] if candidates else (0, None)
        second_score = candidates[1][0] if len(candidates) > 1 else 0
        if product and best_score >= 50 and best_score > second_score:
            matched.append({**item, "product_id": product.id, "sku": product.sku, "name_jp": product.name})
        else:
            unmatched.append(item)

    return {"matched": matched, "unmatched": unmatched}


class ShipmentOrderSaveIn(BaseModel):
    shipped_date: str
    tracking_no: str
    order_no: str
    box_count: int = 0
    total_weight_kg: float = 0
    note: str = ""
    matched: List[dict]
    unmatched: List[dict]


@router.post("/save")
def save_shipment_order(data: ShipmentOrderSaveIn, db: Session = Depends(get_db)):
    order = ShipmentOrder(
        tracking_no=data.tracking_no,
        order_no=data.order_no,
        shipped_date=data.shipped_date,
        box_count=data.box_count,
        total_weight_kg=data.total_weight_kg,
        note=data.note,
        status="pending",
    )
    db.add(order)
    db.flush()

    for item in data.matched:
        db.add(ShipmentOrderItem(
            shipment_order_id=order.id,
            product_id=item.get("product_id"),
            name_cn=item.get("name_cn", ""),
            color=item.get("color", ""),
            size=item.get("size", ""),
            buy_url=item.get("buy_url", ""),
            unit_price_cny=item.get("unit_price_cny", 0),
            qty=item.get("qty", 0),
            is_matched=True,
        ))

    for item in data.unmatched:
        db.add(ShipmentOrderItem(
            shipment_order_id=order.id,
            product_id=item.get("product_id"),  # 手動照合後はセット済み
            name_cn=item.get("name_cn", ""),
            color=item.get("color", ""),
            size=item.get("size", ""),
            buy_url=item.get("buy_url", ""),
            unit_price_cny=item.get("unit_price_cny", 0),
            qty=item.get("qty", 0),
            is_matched=bool(item.get("product_id")),
            # アップロード画面で「対象外」にした行はここで確定させる。
            # 反映しないと入荷後に「未反映の行が残っている配送依頼」へ毎回再登場してしまう。
            is_excluded=bool(item.get("excluded")),
        ))

    db.commit()
    return {"shipment_order_id": order.id}


@router.get("/")
def list_shipment_orders(db: Session = Depends(get_db)):
    orders = db.query(ShipmentOrder).order_by(ShipmentOrder.created_at.desc()).all()
    result = []
    for o in orders:
        items = db.query(ShipmentOrderItem).filter(ShipmentOrderItem.shipment_order_id == o.id).all()
        unmatched_count = sum(1 for i in items if not i.is_matched and not i.is_excluded)
        # 入荷済みなのに在庫へ入っていない行（紐づけ間違い・未照合の取りこぼし）。
        # 対象外にした行（梱包材など）は集計から除外する。
        unreflected_count = sum(1 for i in items if not i.is_reflected and not i.is_excluded)
        pending_reimport = sum(1 for i in items if not i.is_reflected and not i.is_excluded and i.product_id)
        excluded_count = sum(1 for i in items if i.is_excluded)
        result.append({
            "id": o.id,
            "tracking_no": o.tracking_no,
            "order_no": o.order_no,
            "shipped_date": o.shipped_date,
            "box_count": o.box_count,
            "total_weight_kg": o.total_weight_kg,
            "status": o.status,
            "received_at": o.received_at,
            "note": o.note,
            "item_count": len(items),
            "unmatched_count": unmatched_count,
            "unreflected_count": unreflected_count,
            "pending_reimport": pending_reimport,
            "excluded_count": excluded_count,
        })
    return result


@router.get("/{order_id}/items")
def get_shipment_order_items(order_id: int, db: Session = Depends(get_db)):
    items = db.query(ShipmentOrderItem).filter(ShipmentOrderItem.shipment_order_id == order_id).all()
    result = []
    for item in items:
        product = db.query(RakutenProduct).filter(RakutenProduct.id == item.product_id).first() if item.product_id else None
        result.append({
            "id": item.id,
            "product_id": item.product_id,
            "sku": product.sku if product else None,
            "name_jp": product.name if product else None,
            "name_cn": item.name_cn,
            "color": item.color,
            "size": item.size,
            "buy_url": item.buy_url,
            "unit_price_cny": item.unit_price_cny,
            "qty": item.qty,
            "is_matched": item.is_matched,
            "is_reflected": bool(item.is_reflected),
            "is_excluded": bool(item.is_excluded),
        })
    return result


@router.patch("/{order_id}/items/{item_id}/exclude")
def exclude_shipment_item(order_id: int, item_id: int, data: dict, db: Session = Depends(get_db)):
    """在庫に入れる必要がない行（梱包材など）を未反映カウントから除外する／解除する。"""
    item = db.query(ShipmentOrderItem).filter(
        ShipmentOrderItem.id == item_id,
        ShipmentOrderItem.shipment_order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="アイテムが見つかりません")
    item.is_excluded = bool(data.get("excluded", True))
    db.commit()
    return {"ok": True, "is_excluded": item.is_excluded}


@router.patch("/{order_id}/items/{item_id}/match")
def match_item(order_id: int, item_id: int, data: ShipmentOrderItemPatch, db: Session = Depends(get_db)):
    """未照合アイテムに手動でSKUを紐付ける"""
    item = db.query(ShipmentOrderItem).filter(
        ShipmentOrderItem.id == item_id,
        ShipmentOrderItem.shipment_order_id == order_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="アイテムが見つかりません")
    product = db.query(RakutenProduct).filter(RakutenProduct.id == data.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="商品が見つかりません")

    # すでに在庫へ反映済みの行を別商品へ付け替える場合、
    # 元の商品に入れた分を戻してから未反映に戻す。
    # （戻さないと間違ったSKUに在庫が残ったままになる）
    reverted = None
    if item.is_reflected and item.product_id and item.product_id != data.product_id:
        old = db.query(RakutenProduct).filter(RakutenProduct.id == item.product_id).first()
        if old:
            set_size = old.set_size or 1
            qty_units = item.qty // set_size if set_size > 1 else item.qty
            before_stock = old.stock or 0
            old.stock = max(0, before_stock - qty_units)
            db.add(InventoryReflectionLog(
                event_id=str(uuid.uuid4()),
                source="shipment_order",
                source_label="紐づけ修正（取消）",
                source_id=order_id,
                source_ref=f"配送依頼#{order_id}",
                sku=old.sku,
                name=old.name,
                supplier=old.supplier,
                received_qty=-qty_units,
                stock_before=before_stock,
                stock_after=old.stock,
                inbound_before=old.inbound or 0,
                inbound_after=old.inbound or 0,
                standard_stock_before=old.standard_stock or 0,
                standard_stock_after=old.standard_stock or 0,
                rms_push_items=0,
                note="紐づけ先の変更にともない在庫加算を取り消し",
            ))
            reverted = {"sku": old.sku, "qty": qty_units,
                        "stock_before": before_stock, "stock_after": old.stock}
        item.is_reflected = False

    item.product_id = data.product_id
    item.is_matched = True
    db.commit()
    return {"ok": True, "reverted": reverted}


@router.post("/{order_id}/receive")
async def receive_shipment(order_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """入荷済みにして在庫を加算する"""
    order = db.query(ShipmentOrder).filter(ShipmentOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="配送依頼が見つかりません")
    if order.status == "received":
        raise HTTPException(status_code=400, detail="すでに入荷済みです")

    items = db.query(ShipmentOrderItem).filter(ShipmentOrderItem.shipment_order_id == order_id).all()
    return await _apply_receive(db, order, items, mark_received=True)


@router.post("/{order_id}/receive-remaining")
async def receive_remaining(order_id: int, db: Session = Depends(get_db)):
    """入荷済みの配送依頼のうち、まだ在庫に反映されていない行だけを取り込む。

    紐づけ間違いや未照合で取りこぼした行を、紐づけを直したあとに復旧するための操作。
    すでに反映済みの行(is_reflected=True)は対象外なので二重加算にはならない。
    """
    order = db.query(ShipmentOrder).filter(ShipmentOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="配送依頼が見つかりません")

    items = db.query(ShipmentOrderItem).filter(
        ShipmentOrderItem.shipment_order_id == order_id,
        ShipmentOrderItem.is_reflected == False,
        ShipmentOrderItem.is_excluded == False,
        ShipmentOrderItem.product_id != None,
    ).all()
    if not items:
        raise HTTPException(status_code=400, detail="未反映の行はありません（すべて反映済み、または未照合のままです）")

    return await _apply_receive(db, order, items, mark_received=False)


async def _apply_receive(db: Session, order: ShipmentOrder, items, mark_received: bool):
    """配送依頼の行を在庫へ反映する共通処理。receive / receive-remaining から呼ぶ。"""
    updated = 0
    skipped = 0
    skipped_rows = []   # 取り込めなかった行（画面に理由付きで出す）
    duplicate_skipped = 0
    order_consumed = 0  # 発注済みリストから消化した件数
    consumed_skus = set()  # 消化が発生したSKU（発注済2→1の繰り上げ判定用）
    updated_skus = set()   # 在庫を加算したSKU（セット再計算・RMS反映用）
    reflection_rows = []
    # product_id -> その商品で既に反映した(色, サイズ)の組み合わせ
    processed_variants: dict[int, set[tuple[str, str]]] = {}

    def _skip(item, reason: str, sku: str = ""):
        skipped_rows.append({
            "name_cn": item.name_cn or "",
            "color": item.color or "",
            "size": item.size or "",
            "qty": item.qty or 0,
            "buy_url": item.buy_url or "",
            "sku": sku,
            "reason": reason,
        })

    for item in items:
        if not item.product_id:
            skipped += 1
            _skip(item, "商品マスタと未照合")
            continue
        product = db.query(RakutenProduct).filter(RakutenProduct.id == item.product_id).first()
        if not product:
            skipped += 1
            _skip(item, "照合先の商品が見つからない（削除済み）")
            continue

        # 配送依頼の数量は仕入れ単位。販売在庫はset_sizeで割った単位で管理する。
        set_size = product.set_size or 1
        received_qty = item.qty // set_size if set_size > 1 else item.qty
        if received_qty <= 0:
            skipped += 1
            _skip(item, f"セット入数{set_size}に満たないため0個換算", product.sku or "")
            continue

        before = {
            "stock": product.stock or 0,
            "inbound": product.inbound or 0,
            "standard_stock": product.standard_stock or 0,
        }

        # 同一商品・同一色/サイズの複数行は合算する（航空便・船便を別々に発注して
        # 結局同じ便で届いた場合など、正当に同じ商品が複数行に分かれるケースがある）。
        # 色/サイズが違うのに同じ商品IDに解決された場合は紐づけ間違いの疑いが強いため、
        # 黙って捨てずに画面へ出して確認を促す（在庫が入らず気付けなくなるのを防ぐ）。
        variant_key = ((item.color or "").strip(), (item.size or "").strip())
        seen_variants = processed_variants.setdefault(item.product_id, set())
        if seen_variants and variant_key not in seen_variants:
            duplicate_skipped += 1
            _skip(item, f"同じ商品（{product.sku}）に色/サイズが異なる複数行が紐づいているため未反映。紐づけ先を確認してください",
                  product.sku or "")
            continue
        seen_variants.add(variant_key)

        # 在庫加算
        product.stock = (product.stock or 0) + received_qty
        item.is_reflected = True   # 再取り込みで二重加算しないための印
        updated += 1
        updated_skus.add(product.sku)

        # 発注済みリストをSKU・古い順に消化
        remaining = received_qty
        orders = db.query(RakutenOrderHistory).filter(
            RakutenOrderHistory.sku == product.sku,
            RakutenOrderHistory.is_deleted == False,
        ).order_by(RakutenOrderHistory.created_at.asc()).all()

        for o in orders:
            if remaining <= 0:
                break
            if remaining >= o.qty:
                remaining -= o.qty
                o.is_deleted = True
                order_consumed += 1
            else:
                o.qty -= remaining
                remaining = 0
                order_consumed += 1

        # 旧方式で商品マスタに残っている発注済1/2も、移行漏れ対策として消化する。
        if remaining > 0:
            consume_legacy = min(remaining, product.inbound or 0)
            product.inbound = (product.inbound or 0) - consume_legacy
            remaining -= consume_legacy
        if remaining > 0:
            consume_legacy2 = min(remaining, product.standard_stock or 0)
            product.standard_stock = (product.standard_stock or 0) - consume_legacy2
            remaining -= consume_legacy2
        consumed_skus.add(product.sku)

        reflection_rows.append({
            "sku": product.sku,
            "name": product.name,
            "supplier": product.supplier,
            "received_qty": received_qty,
            "stock_before": before["stock"],
            "stock_after": product.stock or 0,
            "inbound_before": before["inbound"],
            "inbound_after": product.inbound or 0,
            "standard_stock_before": before["standard_stock"],
            "standard_stock_after": product.standard_stock or 0,
        })

    # 発注済1が空になったSKUは、残っている発注済2を発注済1へ繰り上げる。
    # SessionLocalはautoflush=Falseなので、消化でセットしたis_deleted=Trueを
    # 先にflushしないと「消化済みの発注済1がまだ残っている」と誤判定して繰り上げを飛ばす。
    db.flush()
    promoted = 0
    for sku in consumed_skus:
        remaining_orders = db.query(RakutenOrderHistory).filter(
            RakutenOrderHistory.sku == sku,
            RakutenOrderHistory.is_deleted == False,
            RakutenOrderHistory.is_delivered == False,
        ).all()
        remaining_orders = [o for o in remaining_orders if not o.is_deleted]
        if any((o.stage or 1) == 1 for o in remaining_orders):
            continue
        for o in remaining_orders:
            if o.stage == 2:
                o.stage = 1
                promoted += 1

    # セット在庫を再計算し、RMSへ反映するitemsを組み立てる。
    # pushしないと毎分のRMS在庫取得(pull)で楽天側の古い在庫に巻き戻ってしまうため必須。
    rms_items = []
    if updated_skus:
        from app.api.routes.rakuten import _recalc_dependent_set_stock, _build_rms_stock_items
        all_products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
        sku_stock = {p.sku: (p.stock or 0) for p in all_products}
        _recalc_dependent_set_stock(all_products, sku_stock, updated_skus)
        rms_items = _build_rms_stock_items(all_products, sku_stock, updated_skus)

    if mark_received:
        order.status = "received"
        order.received_at = datetime.now(timezone.utc)
    event_id = str(uuid.uuid4())
    source_ref = order.tracking_no or order.order_no or f"配送依頼#{order.id}"
    label = "配送依頼" if mark_received else "配送依頼（未反映分の再取込）"
    for row in reflection_rows:
        db.add(InventoryReflectionLog(
            event_id=event_id,
            source="shipment_order",
            source_label=label,
            source_id=order.id,
            source_ref=source_ref,
            note=f"未照合スキップ: {skipped}件 / 発注済消化: {order_consumed}件",
            rms_push_items=0,
            **row,
        ))
    db.commit()

    push_items = 0
    push_result = {"ok": 0, "fail": 0, "errors": [], "details": []}
    if rms_items:
        from app.api.routes.rakuten import _get_or_create_settings
        settings = _get_or_create_settings(db)
        if settings.rms_service_secret and settings.rms_license_key:
            from app.services.rakuten_rms import push_inventory_and_record
            push_items = len(rms_items)
            for attempt in range(3):
                push_result = await push_inventory_and_record(
                    settings.rms_service_secret,
                    settings.rms_license_key,
                    rms_items,
                    source="shipment_order",
                    source_label="配送依頼",
                    event_id=event_id,
                )
                if push_result.get("fail", 0) == 0:
                    break
                await asyncio.sleep(2 * (attempt + 1))
            note_suffix = (
                f" / RMS push ok:{push_result.get('ok', 0)}"
                f" fail:{push_result.get('fail', 0)}"
            )
            db.query(InventoryReflectionLog).filter(
                InventoryReflectionLog.event_id == event_id,
            ).update({
                "rms_push_items": push_items,
                "note": f"未照合スキップ: {skipped}件 / 発注済消化: {order_consumed}件{note_suffix}",
            })
            db.commit()

    failed_details = [d for d in (push_result.get("details") or []) if not d.get("ok")]
    err_by_sku = {e.get("sku"): e.get("detail") for e in (push_result.get("errors") or [])}
    for d in failed_details:
        d["detail"] = err_by_sku.get(d.get("sku"))

    return {"updated": updated, "skipped": skipped, "duplicate_skipped": duplicate_skipped,
            "skipped_rows": skipped_rows,
            "order_consumed": order_consumed,
            "stage_promoted": promoted, "rms_push_items": push_items,
            "rms_push_ok": push_result.get("ok", 0),
            "rms_push_fail": push_result.get("fail", 0),
            "rms_push_errors": push_result.get("errors", []),
            "rms_push_failed": failed_details,
            "rms_push_details": push_result.get("details", [])}
