from datetime import datetime, timezone
import base64
import io
import json
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.rakuten_product import RakutenProduct
from app.models.welfare import WelfareInventoryItem, WelfareInventoryMovement, WelfareWorkInstruction


router = APIRouter(prefix="/welfare", tags=["welfare"])


def _norm_url(url: str | None) -> str:
    if not url:
        return ""
    return str(url).split("?")[0].rstrip("/").lower()


def _unit_per_set(product: RakutenProduct | None) -> int:
    if not product:
        return 1
    if product.set_size and product.set_size > 1:
        return int(product.set_size)
    try:
        comps = json.loads(product.set_components or "[]")
    except Exception:
        comps = []
    if len(comps) == 1:
        qty = comps[0].get("qty") or 1
        try:
            return max(1, int(qty))
        except Exception:
            return 1
    return 1


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _image_data_url(image) -> str:
    try:
        raw = image._data()
    except Exception:
        return ""
    if not raw:
        return ""
    fmt = (getattr(image, "format", "") or "").lower()
    mime = "image/jpeg" if fmt in ("jpeg", "jpg") else f"image/{fmt or 'png'}"
    return f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}"


def _sheet_images_by_row(ws) -> dict[int, str]:
    images = {}
    for image in getattr(ws, "_images", []):
        try:
            row_no = image.anchor._from.row + 1
        except Exception:
            continue
        images.setdefault(row_no, _image_data_url(image))
    return {row_no: data for row_no, data in images.items() if data}


def _parse_excel(content: bytes):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel読み込みエラー: {str(e)}")

    parsed = []
    for ws in wb.worksheets:
        image_by_row = _sheet_images_by_row(ws)
        header_row_idx = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if "商品名" in vals and "数量" in vals:
                header_row_idx = i
                col_map = {v: idx for idx, v in enumerate(vals) if v}
                break
        if header_row_idx is None:
            continue

        def col(name):
            return col_map.get(name, -1)

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
            if all(c is None for c in row):
                continue
            def cell(name):
                idx = col(name)
                if idx < 0 or idx >= len(row):
                    return None
                return row[idx]

            name_cn = _cell_text(cell("商品名"))
            if not name_cn:
                continue
            buy_url = _cell_text(cell("商品URL"))
            if "?" in buy_url:
                buy_url = buy_url.split("?")[0]
            try:
                units = int(float(cell("数量") or 0))
            except Exception:
                units = 0
            if units <= 0:
                continue
            remaining_raw = cell("残")
            try:
                remaining_units = int(float(remaining_raw)) if remaining_raw not in (None, "") else None
            except Exception:
                remaining_units = None
            parsed.append({
                "sheet": ws.title,
                "order_date": _cell_text(cell("発注時間"))[:10],
                "order_no": _cell_text(cell("オーダー番号")),
                "name_cn": name_cn,
                "supplier_spec": _cell_text(cell("色")),
                "size": _cell_text(cell("サイズ")),
                "buy_url": buy_url,
                "image_data_url": image_by_row.get(row_idx, ""),
                "unit_price": _cell_text(cell("単価")),
                "units": units,
                "instruction": _cell_text(cell("指示")),
                "note": "",
                "remaining_units": remaining_units,
            })
    return parsed


def _product_indexes(db: Session):
    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    by_url_spec = {}
    by_url = {}
    by_url_all = {}
    url_counts = {}
    for p in products:
        url = _norm_url(p.buy_url)
        if not url:
            continue
        spec = (p.supplier_spec or "").strip()
        if spec:
            by_url_spec[(url, spec)] = p
        url_counts[url] = url_counts.get(url, 0) + 1
        by_url[url] = p
        by_url_all.setdefault(url, []).append(p)
    unique_url = {url: p for url, p in by_url.items() if url_counts.get(url) == 1}
    return by_url_spec, unique_url, by_url_all


def _match_product(row: dict, by_url_spec: dict, unique_url: dict, by_url_all: dict | None = None):
    url = _norm_url(row.get("buy_url"))
    spec = (row.get("supplier_spec") or "").strip()
    if url and spec and (url, spec) in by_url_spec:
        return by_url_spec[(url, spec)], "url+spec"
    if url and url in unique_url:
        return unique_url[url], "url"
    if by_url_all and url and url in by_url_all:
        color = (row.get("color") or "").strip()
        size = (row.get("size") or "").strip()
        candidates = by_url_all[url]
        for combo in [f"{spec}、{size}", f"{color}、{size}", spec, color]:
            combo = combo.strip("、 ")
            if not combo:
                continue
            for p in candidates:
                p_spec = (p.supplier_spec or "").strip()
                if p_spec and (p_spec.startswith(combo) or combo in p_spec):
                    return p, "url+fuzzy"
    return None, None


def _import_key(product: RakutenProduct | None, row: dict):
    return (
        product.sku if product else None,
        _norm_url(row.get("buy_url")),
        row.get("order_no") or "",
        row.get("supplier_spec") or "",
        int(row.get("units") or 0),
        row.get("sheet") or "",
    )


def _out(item: WelfareInventoryItem):
    return {
        "id": item.id,
        "product_id": item.product_id,
        "sku": item.sku,
        "name_jp": item.name_jp,
        "name_cn": item.name_cn,
        "supplier_spec": item.supplier_spec,
        "buy_url": item.buy_url,
        "image_data_url": item.image_data_url,
        "unit_per_set": item.unit_per_set or 1,
        "total_received_units": item.total_received_units or 0,
        "total_received_qty": item.total_received_qty or 0,
        "withdrawn_qty": item.withdrawn_qty or 0,
        "remaining_qty": item.remaining_qty or 0,
        "instruction": item.instruction or "",
        "note": item.note or "",
        "last_received_at": item.last_received_at.isoformat() if item.last_received_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


@router.get("/inventory")
def list_inventory(q: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(WelfareInventoryItem)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (WelfareInventoryItem.sku.ilike(like)) |
            (WelfareInventoryItem.name_jp.ilike(like)) |
            (WelfareInventoryItem.name_cn.ilike(like)) |
            (WelfareInventoryItem.supplier_spec.ilike(like))
        )
    rows = query.order_by(WelfareInventoryItem.remaining_qty.desc(), WelfareInventoryItem.sku.asc()).all()
    pids = [r.product_id for r in rows if r.product_id]
    products = {p.id: p for p in db.query(RakutenProduct).filter(RakutenProduct.id.in_(pids)).all()} if pids else {}
    result = []
    for r in rows:
        d = _out(r)
        p = products.get(r.product_id)
        d["product_unit_per_set"] = _unit_per_set(p) if p else (r.unit_per_set or 1)
        # SKU・名称は登録時点のコピーではなく、常に商品マスタの最新値を見せる
        # （マスタ側でSKUをリネームしたときに就労支援側が古いままにならないように）
        if p:
            d["sku"] = p.sku
            if p.name:
                d["name_jp"] = p.name
        result.append(d)
    return result


def _work_out(row: WelfareWorkInstruction):
    return {
        "id": row.id,
        "product_id": row.product_id,
        "sku": row.sku,
        "order_date": row.order_date,
        "source_file": row.source_file,
        "source_sheet": row.source_sheet,
        "source_order_no": row.source_order_no,
        "name_jp": row.name_jp,
        "source_product_name": row.source_product_name,
        "color": row.color or row.supplier_spec,
        "size": row.size,
        "supplier_spec": row.supplier_spec,
        "buy_url": row.buy_url,
        "image_data_url": row.image_data_url,
        "unit_price": row.unit_price,
        "units": row.units or 0,
        "unit_per_set": row.unit_per_set or 1,
        "qty": row.qty or 0,
        "instruction": row.instruction or "",
        "remaining_units": row.remaining_units if row.remaining_units is not None else (row.remaining_qty or 0) * (row.unit_per_set or 1),
        "remaining_qty": row.remaining_qty or 0,
        "note": row.note or "",
        "is_reflected": bool(row.is_reflected),
        "reflected_at": row.reflected_at.isoformat() if row.reflected_at else None,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


@router.get("/work-instructions")
def list_work_instructions(q: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(WelfareWorkInstruction)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (WelfareWorkInstruction.sku.ilike(like)) |
            (WelfareWorkInstruction.name_jp.ilike(like)) |
            (WelfareWorkInstruction.source_product_name.ilike(like)) |
            (WelfareWorkInstruction.color.ilike(like)) |
            (WelfareWorkInstruction.size.ilike(like)) |
            (WelfareWorkInstruction.supplier_spec.ilike(like)) |
            (WelfareWorkInstruction.source_order_no.ilike(like))
        )
    rows = query.order_by(
        WelfareWorkInstruction.order_date.desc(),
        WelfareWorkInstruction.id.desc(),
    ).limit(2000).all()
    pids = [r.product_id for r in rows if r.product_id]
    products = {p.id: p for p in db.query(RakutenProduct).filter(RakutenProduct.id.in_(pids)).all()} if pids else {}
    result = []
    for r in rows:
        d = _work_out(r)
        p = products.get(r.product_id)
        # SKU・名称は商品マスタの最新値を見せる（マスタ側でSKUをリネームしても
        # 就労支援側の表示が古いSKUのままにならないように）
        if p:
            d["sku"] = p.sku
            if p.name:
                d["name_jp"] = p.name
        result.append(d)
    return result


@router.post("/preview-excel")
async def preview_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    rows = _parse_excel(await file.read())
    by_url_spec, unique_url, by_url_all = _product_indexes(db)
    result = []
    matched = 0
    for row in rows:
        product, match_type = _match_product(row, by_url_spec, unique_url, by_url_all)
        unit = _unit_per_set(product)
        qty = row["units"] // unit
        if product:
            matched += 1
        result.append({
            **row,
            "matched": bool(product),
            "match_type": match_type,
            "product_id": product.id if product else None,
            "sku": product.sku if product else None,
            "name_jp": product.name if product else None,
            "unit_per_set": unit,
            "qty": qty,
            "remainder_units": row["units"] % unit,
        })
    return {"rows": result, "matched": matched, "unmatched": len(result) - matched}


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    rows = _parse_excel(await file.read())
    return _import_rows(rows, db, source_file=file.filename, clear_existing=False)


class WelfareGoogleImportIn(BaseModel):
    url: str
    clear_existing: bool = False


def _clear_welfare_data(db: Session):
    db.query(WelfareWorkInstruction).delete()
    db.query(WelfareInventoryMovement).delete()
    db.query(WelfareInventoryItem).delete()


def _import_rows(rows: list[dict], db: Session, *, source_file: str, clear_existing: bool = False):
    if clear_existing:
        _clear_welfare_data(db)
        db.flush()

    by_url_spec, unique_url, by_url_all = _product_indexes(db)
    now = datetime.now(timezone.utc)
    imported = 0
    unmatched = 0
    work_imported = 0
    existing_movement_keys = set()
    for m in db.query(WelfareInventoryMovement).filter(WelfareInventoryMovement.movement_type == "import").all():
        existing_movement_keys.add((
            m.sku,
            _norm_url(m.buy_url),
            m.source_order_no or "",
            m.supplier_spec or "",
            int(m.units or 0),
            m.source_sheet or "",
        ))
    existing_work_keys = set()
    for w in db.query(WelfareWorkInstruction).all():
        existing_work_keys.add((
            w.sku,
            _norm_url(w.buy_url),
            w.source_order_no or "",
            w.supplier_spec or "",
            int(w.units or 0),
            w.source_sheet or "",
        ))

    # 同じ商品は前回と同じ指示（保管／戻し等）である場合が多いので、
    # 過去（一番新しいもの）の指示をデフォルト値として引き継ぐ。
    # Excel側に指示が入っている行はそちらを優先する。
    last_instruction_by_product: dict[int, str] = {}
    for pid, instruction in (
        db.query(WelfareWorkInstruction.product_id, WelfareWorkInstruction.instruction)
        .filter(
            WelfareWorkInstruction.product_id.isnot(None),
            WelfareWorkInstruction.instruction.isnot(None),
            WelfareWorkInstruction.instruction != "",
        )
        .order_by(WelfareWorkInstruction.id.asc())
        .all()
    ):
        last_instruction_by_product[pid] = instruction

    unmatched_items = []
    imported_items = []
    skipped_items = []
    for row in rows:
        product, _match_type = _match_product(row, by_url_spec, unique_url, by_url_all)
        if not product:
            unmatched += 1
            unmatched_items.append({
                "name_cn": row.get("name_cn", ""),
                "supplier_spec": row.get("supplier_spec", ""),
                "units": row.get("units", 0),
                "buy_url": row.get("buy_url", ""),
                "sheet": row.get("sheet", ""),
            })
        unit = _unit_per_set(product)
        qty = row["units"] // unit
        remaining_units = row.get("remaining_units")
        remaining_units_value = row["units"] if remaining_units is None else remaining_units
        remaining_qty = remaining_units_value // unit
        if product and qty <= 0:
            continue
        key = _import_key(product, row)
        already_imported = key in existing_movement_keys
        already_work = key in existing_work_keys

        fallback_name = None
        if not product:
            raw_url = row.get("buy_url") or ""
            if raw_url:
                prev = db.query(WelfareWorkInstruction.name_jp).filter(
                    WelfareWorkInstruction.buy_url == raw_url,
                    WelfareWorkInstruction.name_jp.isnot(None),
                    WelfareWorkInstruction.name_jp != "",
                ).first()
                if prev:
                    fallback_name = prev[0]

        if not already_work:
            new_work = WelfareWorkInstruction(
                product_id=product.id if product else None,
                sku=product.sku if product else None,
                order_date=row.get("order_date") or None,
                source_file=source_file,
                source_sheet=row.get("sheet"),
                source_order_no=row.get("order_no"),
                name_jp=product.name if product else fallback_name,
                source_product_name=row.get("name_cn"),
                color=row.get("supplier_spec"),
                size=row.get("size"),
                supplier_spec=row.get("supplier_spec"),
                buy_url=row.get("buy_url"),
                image_data_url=row.get("image_data_url"),
                unit_price=row.get("unit_price"),
                units=row["units"],
                unit_per_set=unit,
                qty=qty,
                instruction=(
                    row.get("instruction")
                    or (last_instruction_by_product.get(product.id, "") if product else "")
                ),
                remaining_units=remaining_units_value,
                remaining_qty=remaining_qty,
                note=None,
                is_reflected=False,
            )
            db.add(new_work)
            work_imported += 1
            if product and new_work.instruction:
                last_instruction_by_product[product.id] = new_work.instruction
        if not product:
            continue
        # 取込では就労支援在庫へ加算しない。
        # 「こっちに持ち帰る分(戻し)」と「施設で保管する分」は荷受け処理で仕分けるため、
        # 指示と残を確定したあとに POST /welfare/work-instructions/reflect で残の数量だけ在庫化する。
        if already_imported:
            skipped_items.append({
                "sku": product.sku,
                "name_jp": product.name,
                "units": row["units"],
                "qty": qty,
                "status": "既取込済",
            })
            continue
        existing_movement_keys.add(key)
    db.commit()
    return {
        "imported": imported,
        "work_imported": work_imported,
        "unmatched": unmatched,
        "imported_items": imported_items,
        "skipped_items": skipped_items,
        "unmatched_items": unmatched_items,
    }


class WelfareReflectIn(BaseModel):
    ids: Optional[list[int]] = None   # 未指定なら未反映の全行


def _is_full_return(instruction: Optional[str]) -> bool:
    """指示が「全量こちらに戻す」かどうか。

    戻しの行でも「残」には返す個数を入れたままにする運用（施設側が何個返すか分かるように）なので、
    在庫へ入れるかどうかは残の値ではなく指示で判定する。
    - 「戻し」           → 全量戻し（在庫化しない）
    - 「作業後戻し」等   → 施設で作業してから戻すので在庫化する
    - 「30戻し」「戻し10」→ 一部だけ戻す。残が施設に置く数なので在庫化する
    """
    import re
    s = (instruction or "").strip()
    if "戻し" not in s:
        return False
    if "作業" in s:
        return False
    if re.search(r"\d", s):
        return False
    return True


@router.post("/work-instructions/reflect")
def reflect_work_instructions(data: WelfareReflectIn, db: Session = Depends(get_db)):
    """荷受け行の「残」の数量だけを就労支援在庫へ反映する。

    取込時点では在庫化せず、指示（保管／戻し）と残を確定したあとにこれを実行する。
    残が0の行（＝全部こちらに持ち帰る＝戻し）は在庫化されない。
    反映済みの行は対象外なので、何度押しても二重計上しない。
    """
    q = db.query(WelfareWorkInstruction).filter(
        WelfareWorkInstruction.is_reflected == False,
        WelfareWorkInstruction.product_id != None,
    )
    if data.ids:
        q = q.filter(WelfareWorkInstruction.id.in_(data.ids))
    rows = q.all()
    if not rows:
        return {"reflected": 0, "skipped_zero": 0, "skipped_return": 0, "items": [],
                "message": "反映できる行がありません（すべて反映済み、または商品未照合です）"}

    now = datetime.now(timezone.utc)
    reflected = 0
    skipped_zero = 0
    skipped_return = 0
    results = []
    for w in rows:
        product = db.query(RakutenProduct).filter(RakutenProduct.id == w.product_id).first()
        if not product:
            continue
        unit = w.unit_per_set or 1
        keep_units = w.remaining_units if w.remaining_units is not None else (w.remaining_qty or 0) * unit
        keep_qty = (w.remaining_qty if w.remaining_qty is not None else keep_units // unit) or 0

        # 「戻し」の行は残に返す個数が入っていても在庫化しない
        if _is_full_return(w.instruction):
            w.is_reflected = True
            w.reflected_at = now
            skipped_return += 1
            results.append({"sku": w.sku, "name_jp": w.name_jp, "keep_qty": 0,
                            "instruction": w.instruction or "", "status": "在庫化せず（戻し）"})
            continue

        # 残0＝施設に置かない。在庫は作らず反映済みにして次回対象から外す。
        if keep_qty <= 0:
            w.is_reflected = True
            w.reflected_at = now
            skipped_zero += 1
            results.append({"sku": w.sku, "name_jp": w.name_jp, "keep_qty": 0,
                            "instruction": w.instruction or "", "status": "在庫化せず（残0）"})
            continue

        item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.product_id == product.id).first()
        is_new = item is None
        before_qty = (item.remaining_qty or 0) if item else 0
        if not item:
            item = WelfareInventoryItem(
                product_id=product.id,
                sku=product.sku,
                name_jp=product.name,
                name_cn=w.source_product_name,
                supplier_spec=w.supplier_spec,
                buy_url=w.buy_url,
                image_data_url=w.image_data_url,
                unit_per_set=unit,
                total_received_units=0,
                total_received_qty=0,
                withdrawn_qty=0,
                remaining_qty=0,
            )
            db.add(item)
            db.flush()

        item.name_cn = w.source_product_name or item.name_cn
        item.supplier_spec = w.supplier_spec or item.supplier_spec
        item.buy_url = w.buy_url or item.buy_url
        item.image_data_url = w.image_data_url or item.image_data_url
        item.unit_per_set = unit
        item.total_received_units = (item.total_received_units or 0) + keep_units
        item.total_received_qty = (item.total_received_qty or 0) + keep_qty
        item.remaining_qty = (item.remaining_qty or 0) + keep_qty
        item.last_received_at = now
        if w.instruction:
            item.instruction = w.instruction

        db.add(WelfareInventoryMovement(
            item_id=item.id,
            product_id=item.product_id,
            sku=item.sku,
            movement_type="import",
            source_file=w.source_file,
            source_sheet=w.source_sheet,
            source_order_no=w.source_order_no,
            name_cn=w.source_product_name,
            supplier_spec=w.supplier_spec,
            buy_url=w.buy_url,
            units=keep_units,
            qty=keep_qty,
            note=f"荷受け反映（{w.instruction or '指示なし'}）",
        ))
        w.is_reflected = True
        w.reflected_at = now
        reflected += 1
        results.append({
            "sku": w.sku, "name_jp": w.name_jp, "keep_qty": keep_qty,
            "instruction": w.instruction or "",
            "before_qty": before_qty, "after_qty": item.remaining_qty,
            "status": "新規" if is_new else "追加",
        })
    db.commit()
    return {"reflected": reflected, "skipped_zero": skipped_zero,
            "skipped_return": skipped_return, "items": results}


@router.post("/import-google-sheet")
async def import_google_sheet(data: WelfareGoogleImportIn, db: Session = Depends(get_db)):
    import re
    import httpx

    m = re.search(r"/spreadsheets/d/([^/]+)", data.url or "")
    if not m:
        raise HTTPException(status_code=400, detail="GoogleスプレッドシートURLを指定してください")
    sheet_id = m.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            res = await client.get(export_url)
        res.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Googleスプレッドシート取得エラー: {str(e)}")
    rows = _parse_excel(res.content)
    return _import_rows(rows, db, source_file="google-sheet", clear_existing=data.clear_existing)


class WelfareClearIn(BaseModel):
    confirm: str


@router.post("/clear")
def clear_welfare(data: WelfareClearIn, db: Session = Depends(get_db)):
    if data.confirm != "CLEAR":
        raise HTTPException(status_code=400, detail="confirm に CLEAR を指定してください")
    _clear_welfare_data(db)
    db.commit()
    return {"ok": True}


class WelfareMemoIn(BaseModel):
    name_jp: Optional[str] = None
    instruction: Optional[str] = None
    note: Optional[str] = None
    unit_per_set: Optional[int] = None


@router.patch("/inventory/{item_id}")
def update_inventory_item(item_id: int, data: WelfareMemoIn, db: Session = Depends(get_db)):
    item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="就労支援在庫が見つかりません")
    if data.name_jp is not None:
        item.name_jp = data.name_jp
    if data.instruction is not None:
        item.instruction = data.instruction
    if data.note is not None:
        item.note = data.note
    if data.unit_per_set is not None and data.unit_per_set >= 1:
        item.unit_per_set = data.unit_per_set
    db.commit()
    db.refresh(item)
    return _out(item)


class WelfareWithdrawIn(BaseModel):
    qty: int
    note: Optional[str] = None


class WelfareAdjustIn(BaseModel):
    remaining_qty: int
    note: Optional[str] = None


class WelfareWorkInstructionIn(BaseModel):
    product_id: Optional[int] = None
    sku: Optional[str] = None
    name_jp: Optional[str] = None
    source_product_name: Optional[str] = None
    instruction: Optional[str] = None
    remaining_units: Optional[int] = None
    remaining_qty: Optional[int] = None
    note: Optional[str] = None


@router.delete("/inventory/{item_id}")
def delete_inventory_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="就労支援在庫が見つかりません")
    db.query(WelfareInventoryMovement).filter(WelfareInventoryMovement.item_id == item_id).delete()
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.post("/inventory/{item_id}/withdraw")
def withdraw_inventory(item_id: int, data: WelfareWithdrawIn, db: Session = Depends(get_db)):
    item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="就労支援在庫が見つかりません")
    if data.qty <= 0:
        raise HTTPException(status_code=400, detail="減算数は1以上で入力してください")
    if data.qty > (item.remaining_qty or 0):
        raise HTTPException(status_code=400, detail="残量を超えて減算できません")
    item.remaining_qty = (item.remaining_qty or 0) - data.qty
    item.withdrawn_qty = (item.withdrawn_qty or 0) + data.qty
    db.add(WelfareInventoryMovement(
        item_id=item.id,
        product_id=item.product_id,
        sku=item.sku,
        movement_type="withdraw",
        name_cn=item.name_cn,
        supplier_spec=item.supplier_spec,
        buy_url=item.buy_url,
        units=data.qty * (item.unit_per_set or 1),
        qty=-data.qty,
        note=data.note,
    ))
    db.commit()
    db.refresh(item)
    return _out(item)


@router.post("/inventory/{item_id}/adjust")
def adjust_inventory(item_id: int, data: WelfareAdjustIn, db: Session = Depends(get_db)):
    item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="就労支援在庫が見つかりません")
    if data.remaining_qty < 0:
        raise HTTPException(status_code=400, detail="残量は0以上で入力してください")

    before = item.remaining_qty or 0
    after = data.remaining_qty
    diff = after - before
    if diff == 0:
        return _out(item)

    item.remaining_qty = after
    db.add(WelfareInventoryMovement(
        item_id=item.id,
        product_id=item.product_id,
        sku=item.sku,
        movement_type="adjust",
        name_cn=item.name_cn,
        supplier_spec=item.supplier_spec,
        buy_url=item.buy_url,
        units=diff * (item.unit_per_set or 1),
        qty=diff,
        note=data.note or f"残量修正: {before} -> {after}",
    ))
    db.commit()
    db.refresh(item)
    return _out(item)


@router.patch("/work-instructions/{instruction_id}")
def update_work_instruction(instruction_id: int, data: WelfareWorkInstructionIn, db: Session = Depends(get_db)):
    row = db.query(WelfareWorkInstruction).filter(WelfareWorkInstruction.id == instruction_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="作業指示が見つかりません")
    if data.product_id is not None:
        row.product_id = data.product_id
    if data.sku is not None:
        row.sku = data.sku
    if data.name_jp is not None:
        row.name_jp = data.name_jp
    if data.source_product_name is not None:
        row.source_product_name = data.source_product_name
    if data.instruction is not None:
        row.instruction = data.instruction
    if data.remaining_units is not None:
        if data.remaining_units < 0:
            raise HTTPException(status_code=400, detail="残は0以上で入力してください")
        row.remaining_units = data.remaining_units
        row.remaining_qty = data.remaining_units // (row.unit_per_set or 1)
    if data.remaining_qty is not None:
        if data.remaining_qty < 0:
            raise HTTPException(status_code=400, detail="残は0以上で入力してください")
        row.remaining_qty = data.remaining_qty
        row.remaining_units = data.remaining_qty * (row.unit_per_set or 1)
    if data.note is not None:
        row.note = data.note
    db.commit()
    db.refresh(row)
    return _work_out(row)


@router.post("/work-instructions/backfill-products")
def backfill_work_instruction_products(db: Session = Depends(get_db)):
    """product_id未設定の荷受けレコードをbuy_urlで再照合し、SKU・日本語名を埋める"""
    by_url_spec, unique_url, by_url_all = _product_indexes(db)
    rows = db.query(WelfareWorkInstruction).filter(WelfareWorkInstruction.product_id.is_(None)).all()
    updated = 0
    for row in rows:
        product, _ = _match_product({
            "buy_url": row.buy_url,
            "supplier_spec": row.supplier_spec or row.color or "",
            "color": row.color or "",
            "size": row.size or "",
        }, by_url_spec, unique_url, by_url_all)
        if product:
            row.product_id = product.id
            row.sku = product.sku
            row.name_jp = product.name
            row.unit_per_set = _unit_per_set(product)
            if row.units and row.unit_per_set:
                row.qty = row.units // row.unit_per_set
                row.remaining_qty = (row.remaining_units or row.units) // row.unit_per_set
            updated += 1
    db.commit()
    return {"ok": True, "checked": len(rows), "updated": updated}


class WelfareWorkBatchUpdateSheet(BaseModel):
    old_sheet: str
    new_sheet: str


@router.post("/work-instructions/batch-update-sheet")
def batch_update_work_sheet(data: WelfareWorkBatchUpdateSheet, db: Session = Depends(get_db)):
    rows = db.query(WelfareWorkInstruction).filter(WelfareWorkInstruction.source_sheet == data.old_sheet).all()
    for r in rows:
        r.source_sheet = data.new_sheet
    db.commit()
    return {"ok": True, "updated": len(rows)}


@router.delete("/work-instructions/{instruction_id}")
def delete_work_instruction(instruction_id: int, db: Session = Depends(get_db)):
    row = db.query(WelfareWorkInstruction).filter(WelfareWorkInstruction.id == instruction_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="作業指示が見つかりません")
    db.delete(row)
    db.commit()
    return {"ok": True}


class WelfareUndoMovementsIn(BaseModel):
    movement_ids: list[int]


@router.post("/undo-movements")
def undo_movements(data: WelfareUndoMovementsIn, db: Session = Depends(get_db)):
    movements = db.query(WelfareInventoryMovement).filter(
        WelfareInventoryMovement.id.in_(data.movement_ids),
        WelfareInventoryMovement.movement_type == "import",
    ).all()
    undone = []
    for m in movements:
        if m.item_id:
            item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.id == m.item_id).first()
            if item:
                item.total_received_units = max(0, (item.total_received_units or 0) - (m.units or 0))
                item.total_received_qty = max(0, (item.total_received_qty or 0) - (m.qty or 0))
                item.remaining_qty = max(0, (item.remaining_qty or 0) - (m.qty or 0))
        undone.append(m.sku)
        db.delete(m)
    db.commit()
    return {"undone": len(undone), "skus": undone}


class WelfareBulkCreateIn(BaseModel):
    product_ids: list[int] = []
    items: list[dict] = []


@router.post("/inventory/bulk-create")
def bulk_create_inventory(data: WelfareBulkCreateIn, db: Session = Depends(get_db)):
    existing_pids = set(
        r[0] for r in db.query(WelfareInventoryItem.product_id).all() if r[0]
    )
    existing_skus = set(
        r[0] for r in db.query(WelfareInventoryItem.sku).all() if r[0]
    )
    created = []
    for pid in data.product_ids:
        if pid in existing_pids:
            continue
        product = db.query(RakutenProduct).filter(RakutenProduct.id == pid).first()
        if not product:
            continue
        item = WelfareInventoryItem(
            product_id=product.id,
            sku=product.sku,
            name_jp=product.name,
            buy_url=product.buy_url,
            supplier_spec=product.supplier_spec,
            unit_per_set=_unit_per_set(product),
            total_received_units=0,
            total_received_qty=0,
            withdrawn_qty=0,
            remaining_qty=0,
        )
        db.add(item)
        existing_pids.add(pid)
        existing_skus.add(product.sku)
        created.append(product.sku)
    for entry in data.items:
        sku = entry.get("sku", "")
        if not sku or sku in existing_skus:
            continue
        product = db.query(RakutenProduct).filter(RakutenProduct.sku == sku).first()
        item = WelfareInventoryItem(
            product_id=product.id if product else None,
            sku=sku,
            name_jp=entry.get("name_jp") or (product.name if product else ""),
            buy_url=entry.get("buy_url") or (product.buy_url if product else None),
            supplier_spec=entry.get("supplier_spec") or (product.supplier_spec if product else None),
            unit_per_set=entry.get("unit_per_set") or (_unit_per_set(product) if product else 1),
            total_received_units=0,
            total_received_qty=0,
            withdrawn_qty=0,
            remaining_qty=entry.get("remaining_qty", 0),
        )
        db.add(item)
        existing_skus.add(sku)
        if product:
            existing_pids.add(product.id)
        created.append(sku)
    db.commit()
    return {"created": len(created), "skus": created}


@router.post("/inventory/backfill")
def backfill_from_product(db: Session = Depends(get_db)):
    items = db.query(WelfareInventoryItem).filter(
        WelfareInventoryItem.product_id.isnot(None)
    ).all()
    updated = 0
    for item in items:
        product = db.query(RakutenProduct).filter(RakutenProduct.id == item.product_id).first()
        if not product:
            continue
        changed = False
        if product.buy_url and not item.buy_url:
            item.buy_url = product.buy_url
            changed = True
        if product.supplier_spec and not item.supplier_spec:
            item.supplier_spec = product.supplier_spec
            changed = True
        correct_unit = _unit_per_set(product)
        if item.unit_per_set != correct_unit:
            item.unit_per_set = correct_unit
            changed = True
        if changed:
            updated += 1
    db.commit()
    return {"updated": updated}


class WelfareMergeProductsIn(BaseModel):
    keep_sku: str    # 統合先（残す方）
    remove_sku: str  # 統合元（消す方）


@router.post("/inventory/merge-products")
def merge_duplicate_products(data: WelfareMergeProductsIn, db: Session = Depends(get_db)):
    """スペル違いなどで別商品として登録されてしまった重複マスタを統合する。

    remove_sku側の就労支援在庫（残数）をkeep_sku側へ合算し、履歴（荷受け・入出庫）も
    keep_sku側の商品IDに付け替えたうえで、remove_sku側のマスタは非表示化（is_active=False）する。
    """
    keep = db.query(RakutenProduct).filter(RakutenProduct.sku == data.keep_sku, RakutenProduct.is_active == True).first()
    remove = db.query(RakutenProduct).filter(RakutenProduct.sku == data.remove_sku, RakutenProduct.is_active == True).first()
    if not keep:
        raise HTTPException(404, f"統合先SKUが見つかりません: {data.keep_sku}")
    if not remove:
        raise HTTPException(404, f"統合元SKUが見つかりません: {data.remove_sku}")
    if keep.id == remove.id:
        raise HTTPException(400, "同じ商品は統合できません")

    keep_item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.product_id == keep.id).first()
    remove_item = db.query(WelfareInventoryItem).filter(WelfareInventoryItem.product_id == remove.id).first()

    moved_qty = 0
    if remove_item:
        moved_qty = remove_item.remaining_qty or 0
        if not keep_item:
            keep_item = WelfareInventoryItem(
                product_id=keep.id, sku=keep.sku, name_jp=keep.name,
                buy_url=keep.buy_url, supplier_spec=keep.supplier_spec,
                unit_per_set=_unit_per_set(keep),
                total_received_units=0, total_received_qty=0, withdrawn_qty=0, remaining_qty=0,
            )
            db.add(keep_item)
            db.flush()
        keep_item.total_received_units = (keep_item.total_received_units or 0) + (remove_item.total_received_units or 0)
        keep_item.total_received_qty = (keep_item.total_received_qty or 0) + (remove_item.total_received_qty or 0)
        keep_item.withdrawn_qty = (keep_item.withdrawn_qty or 0) + (remove_item.withdrawn_qty or 0)
        keep_item.remaining_qty = (keep_item.remaining_qty or 0) + (remove_item.remaining_qty or 0)
        db.delete(remove_item)

    # 履歴はkeep側の商品IDへ付け替えて残す（消さない）
    db.query(WelfareInventoryMovement).filter(WelfareInventoryMovement.product_id == remove.id).update(
        {"product_id": keep.id, "sku": keep.sku}
    )
    db.query(WelfareWorkInstruction).filter(WelfareWorkInstruction.product_id == remove.id).update(
        {"product_id": keep.id, "sku": keep.sku}
    )

    remove.is_active = False
    db.commit()

    return {
        "ok": True,
        "keep_sku": keep.sku,
        "removed_sku": remove.sku,
        "moved_remaining_qty": moved_qty,
        "keep_item_remaining_qty": keep_item.remaining_qty if keep_item else 0,
    }


@router.get("/movements")
def list_movements(item_id: Optional[int] = None, limit: int = 200,
                   movement_type: Optional[str] = None, db: Session = Depends(get_db)):
    limit = max(1, min(limit, 5000))
    query = db.query(WelfareInventoryMovement)
    if item_id:
        query = query.filter(WelfareInventoryMovement.item_id == item_id)
    if movement_type:
        query = query.filter(WelfareInventoryMovement.movement_type == movement_type)
    rows = query.order_by(WelfareInventoryMovement.id.desc()).limit(limit).all()
    return [{
        "id": r.id,
        "item_id": r.item_id,
        "sku": r.sku,
        "movement_type": r.movement_type,
        "source_file": r.source_file,
        "source_order_no": r.source_order_no,
        "name_cn": r.name_cn,
        "supplier_spec": r.supplier_spec,
        "buy_url": r.buy_url,
        "units": r.units,
        "qty": r.qty,
        "note": r.note,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in rows]
