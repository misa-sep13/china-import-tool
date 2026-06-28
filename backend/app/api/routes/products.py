from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, timezone
from app.core.database import get_db
from app.models.product import Product

router = APIRouter(prefix="/products", tags=["products"])


class ProductCreate(BaseModel):
    sku: str
    fnsku: Optional[str] = ""
    asin: Optional[str] = ""
    name: Optional[str] = ""
    amazon_url: Optional[str] = ""
    buy_url: Optional[str] = ""
    supplier: Optional[str] = "タオタロウ"
    photo_url: Optional[str] = ""
    color: Optional[str] = ""
    size: Optional[str] = ""
    price: Optional[float] = 0
    repack: Optional[str] = ""
    spec: Optional[str] = ""
    customer_memo: Optional[str] = ""
    note: Optional[str] = ""
    set_size: Optional[int] = 1
    extra_stock: Optional[int] = 0
    amazon_fee_rate: Optional[float] = 0.1

def _restore_deleted_product(existing: Product, data: ProductCreate, db: Session) -> Product:
    for k, v in data.model_dump().items():
        setattr(existing, k, v)
    existing.is_active = True
    if existing.no is None:
        existing.no = db.query(Product).count() + 1
    db.commit()
    db.refresh(existing)
    return existing

def _retire_deleted_sku(product: Product):
    product.sku = f"__deleted__{product.id}__{product.sku or 'sku'}"

class ProductUpdate(BaseModel):
    sku: Optional[str] = None
    fnsku: Optional[str] = None
    asin: Optional[str] = None
    name: Optional[str] = None
    amazon_url: Optional[str] = None
    buy_url: Optional[str] = None
    supplier: Optional[str] = None
    photo_url: Optional[str] = None
    color: Optional[str] = None
    size: Optional[str] = None
    price: Optional[float] = None
    repack: Optional[str] = None
    spec: Optional[str] = None
    customer_memo: Optional[str] = None
    note: Optional[str] = None
    set_size: Optional[int] = None
    extra_stock: Optional[int] = None
    amazon_fee_rate: Optional[float] = None
    selling_price: Optional[float] = None
    fba_fee: Optional[float] = None

class ProductOut(ProductCreate):
    id: int
    no: Optional[int] = None
    order_qty: int = 0
    is_active: bool = True
    selling_price: Optional[float] = None
    fba_fee: Optional[float] = None
    fees_updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True

@router.get("/", response_model=List[ProductOut])
def list_products(db: Session = Depends(get_db)):
    return db.query(Product).filter(Product.is_active == True).order_by(Product.no).all()

@router.post("/fba-import", response_model=dict)
def import_from_fba(db: Session = Depends(get_db)):
    try:
        from app.services.amazon_api import fetch_inventory
        inventory = fetch_inventory()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SP-APIエラー: {str(e)}")

    added = 0
    skipped = 0
    fixed = 0
    for fnsku, item in inventory.items():
        asin = item.get("asin", "")
        api_sku = item.get("sku", "")

        # SKUで既存商品を検索してFNSKUを修正
        if api_sku:
            by_sku = db.query(Product).filter(Product.sku == api_sku).first()
            if by_sku:
                if by_sku.fnsku != fnsku:
                    by_sku.fnsku = fnsku
                    fixed += 1
                else:
                    skipped += 1
                continue

        # FNSKUで既存商品があればスキップ
        if db.query(Product).filter(Product.fnsku == fnsku).first():
            skipped += 1
            continue

        # 新規追加（SKUはSP-APIのSKUを使用）
        sku = api_sku or fnsku
        if db.query(Product).filter(Product.sku == sku).first():
            skipped += 1
            continue
        max_no = db.query(Product).count()
        p = Product(sku=sku, fnsku=fnsku, asin=asin, name="", no=max_no + 1)
        db.add(p)
        added += 1

    db.commit()
    return {"added": added, "skipped": skipped, "fixed": fixed}

@router.post("/", response_model=ProductOut)
def create_product(data: ProductCreate, db: Session = Depends(get_db)):
    existing = db.query(Product).filter(Product.sku == data.sku).first()
    if existing:
        if existing.is_active:
            raise HTTPException(status_code=400, detail="SKUが既に存在します")
        return _restore_deleted_product(existing, data, db)
    max_no = db.query(Product).count()
    p = Product(**data.model_dump(), no=max_no + 1)
    db.add(p)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="SKUが既に存在します")
    db.refresh(p)
    return p

@router.post("/restore", response_model=ProductOut)
def restore_product(data: ProductCreate, db: Session = Depends(get_db)):
    existing = db.query(Product).filter(Product.sku == data.sku).first()
    if not existing:
        return create_product(data, db)
    if existing.is_active:
        raise HTTPException(status_code=400, detail="SKUが既に存在します")
    return _restore_deleted_product(existing, data, db)

@router.put("/{product_id}", response_model=ProductOut)
def update_product(product_id: int, data: ProductUpdate, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="商品が見つかりません")
    updates = data.model_dump(exclude_none=True)
    if updates.get("sku") and updates["sku"] != p.sku:
        existing = db.query(Product).filter(Product.sku == updates["sku"], Product.id != product_id).first()
        if existing:
            if existing.is_active:
                raise HTTPException(status_code=400, detail="SKUが既に存在します")
            _retire_deleted_sku(existing)
    for k, v in updates.items():
        setattr(p, k, v)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="SKUが既に存在します")
    db.refresh(p)
    return p

@router.post("/refresh-fees")
def refresh_fees(db: Session = Depends(get_db)):
    """全商品の販売価格・FBA手数料をSP-APIから取得してDBに保存"""
    from app.core.config import settings as app_settings
    if not app_settings.SP_API_REFRESH_TOKEN:
        raise HTTPException(status_code=400, detail="SP-API未設定")

    products = db.query(Product).filter(Product.is_active == True, Product.sku != None).all()
    sku_list = [p.sku for p in products if p.sku]
    if not sku_list:
        return {"updated": 0}

    from app.services.amazon_api import fetch_prices_and_fees
    asin_map = {p.sku: p.asin for p in products if p.sku and p.asin}
    fallback_prices = {
        p.sku: p.selling_price
        for p in products
        if p.sku and p.selling_price and p.selling_price > 0
    }
    fees_map = fetch_prices_and_fees(
        sku_list,
        asin_map=asin_map,
        fallback_prices=fallback_prices,
    )

    now = datetime.now(timezone.utc)
    updated = 0
    price_updated = 0
    fee_updated = 0
    price_missing = 0
    fee_missing = 0
    price_sources = {}
    fee_sources = {}
    for p in products:
        info = fees_map.get(p.sku)
        if not info:
            continue
        if info["selling_price"] is not None:
            p.selling_price = info["selling_price"]
            price_updated += 1
        else:
            price_missing += 1
        if info["fba_fee"] is not None:
            p.fba_fee = info["fba_fee"]
            fee_updated += 1
        else:
            fee_missing += 1
        source = info.get("price_source") or "missing"
        price_sources[source] = price_sources.get(source, 0) + 1
        fee_source = info.get("fee_source") or "missing"
        fee_sources[fee_source] = fee_sources.get(fee_source, 0) + 1
        p.fees_updated_at = now
        updated += 1

    db.commit()
    return {
        "updated": updated,
        "price_updated": price_updated,
        "fee_updated": fee_updated,
        "price_missing": price_missing,
        "fee_missing": fee_missing,
        "price_sources": price_sources,
        "fee_sources": fee_sources,
    }


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="商品が見つかりません")
    p.is_active = False
    db.commit()
    return {"ok": True}


@router.get("/export/t4s-cost")
def export_t4s_cost(db: Session = Depends(get_db)):
    """Tool4Seller コスト一括入力用Excelを生成してダウンロード"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io
    from app.models.settings import OrderSettings

    settings_row = db.query(OrderSettings).first()
    exchange_rate = getattr(settings_row, 'exchange_rate', 21.0) or 21.0

    products = db.query(Product).filter(Product.is_active == True).order_by(Product.no).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 行1: タイトル
    ws.cell(row=1, column=1, value="Cost_Template")

    # 行2: ヘッダー
    headers = ["セラーアカウント", "店舗名", "セラーID", "マーケットプレイスID",
               "商品名", "子ASIN", "SKU", "仕入れ単価", "物流単価", "通貨"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)

    # 固定値（ThreeSky+ Japan店舗）
    SELLER_ACCOUNT = "ThreeSky+"
    SHOP_NAME = "Japan"
    SELLER_ID = "A29K12KTHSASJ0"
    MARKETPLACE_ID = "A1VC38T7YXB528"

    # 行3以降: 商品データ
    for row_idx, p in enumerate(products, 3):
        cost_jpy = round(p.price or 0)
        ws.cell(row=row_idx, column=1, value=SELLER_ACCOUNT)
        ws.cell(row=row_idx, column=2, value=SHOP_NAME)
        ws.cell(row=row_idx, column=3, value=SELLER_ID)
        ws.cell(row=row_idx, column=4, value=MARKETPLACE_ID)
        ws.cell(row=row_idx, column=5, value=p.name or "")
        ws.cell(row=row_idx, column=6, value=p.asin or "")
        ws.cell(row=row_idx, column=7, value=p.sku or "")
        ws.cell(row=row_idx, column=8, value=cost_jpy)
        ws.cell(row=row_idx, column=9, value=0)
        ws.cell(row=row_idx, column=10, value="JPY")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=t4s_cost.xlsx"}
    )
