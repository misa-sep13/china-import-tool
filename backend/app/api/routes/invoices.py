from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.invoice import Invoice, InvoiceItem
from app.models.product import Product
from pydantic import BaseModel
from typing import List, Optional
from app.services import invoice_calc
import io, re

router = APIRouter(prefix="/invoices", tags=["invoices"])


class InvoiceItemIn(BaseModel):
    sku: str = ""            # 空欄=対象外としてスキップする
    asin: str = ""
    name_cn: str = ""
    name_jp: str = ""
    qty: int
    unit_price_cny: float
    buy_url: str = ""
    permit_col: Optional[int] = None  # 手動で指定した申告欄番号
    goods_id: str = ""   # M列（商品ID）：箱单シートと突き合わせて送料を重量配賦する


class PermitColumnIn(BaseModel):
    col_no: int
    item_name: str = ""
    hs_code: str = ""
    cif_jpy: int = 0
    tariff_rate: float = 0.0
    tariff_rate_str: str = ""
    duty_jpy: int = 0
    # 欄ごとの内国消費税の実額。税率から再計算せずこの額を按分する
    consumption_tax_jpy: int = 0
    local_tax_jpy: int = 0
    bpr_coeff: float = 0.0


class InvoiceIn(BaseModel):
    invoice_no: str
    invoice_date: str
    exchange_rate: float
    domestic_freight: float = 0
    international_freight: float = 0
    total_weight: float = 0
    total_volume: float = 0
    note: str = ""
    # 輸入許可書から取得
    customs_duty: int = 0
    consumption_tax: int = 0
    local_consumption_tax: int = 0
    total_tax: int = 0
    import_tax_jpy: float = 0   # 輸入税合計（関税+消費税+地方消費税）。原価へ按分する
    bl_number: str = ""
    declaration_no: str = ""
    items: List[InvoiceItemIn]
    permit_columns: List[PermitColumnIn] = []  # 空=従来の一律按分
    force_save: bool = False  # 検算NGでも承知の上で保存する
    # 箱シートから読んだ箱の計費重量と中身。あれば送料を重量で配る
    box_data: dict | None = None
    # 通関料（円）。船便は一律2000円、航空便は無し。許可書には載らない費用
    customs_fee_jpy: float = 0


@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel読み込みエラー: {str(e)}")

    ws = wb.active

    # インボイス番号を探す（VIP...の行）
    invoice_no = ""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for cell in row:
            if cell and str(cell).startswith("VIP"):
                invoice_no = str(cell)
                break

    # 商品行を解析（ヘッダー行を探す）
    header_row = None
    col_asin = None  # ASIN列インデックス
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] == "10) Name of Commodity":
            header_row = i + 1  # 次の行がデータ開始
            # ヘッダー行からASIN列を探す
            for ci, cell in enumerate(row):
                if cell and "ASIN" in str(cell).upper():
                    col_asin = ci
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="商品データが見つかりません")

    items = []
    added_value = 0
    domestic_freight = 0
    international_freight = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        label = str(row[0]) if row[0] else ""
        # Added Value(増値費用)は検品・ラベル貼付などの加工費。便全体で1行にまとまっており
        # 商品ごとの内訳は無いので、国内送料と同じく金額比で按分する。
        # これを拾わないと原価に一切反映されず、丸ごと漏れる。
        if label and ("Added Value" in label or "增值" in label or "増値" in label):
            added_value = row[8] or 0
            continue
        if label.startswith("Domestic") or "国内运费" in label or "国内送料" in label:
            domestic_freight = row[8] or 0
            continue
        if label.startswith("International") or "国际运费" in label or "国際送料" in label:
            international_freight = row[8] or 0
            continue
        if label.startswith("MADE IN"):
            break
        if row[0] is None and row[6] and row[7]:
            asin_val = str(row[col_asin] or "") if col_asin is not None and col_asin < len(row) else ""
            items.append({
                "sku": str(int(row[12])) if row[12] else "",
                "asin": asin_val,
                "name_cn": str(row[1] or ""),
                "name_jp": str(row[2] or ""),
                "qty": int(row[6]) if row[6] else 0,
                "unit_price_cny": float(row[7]) if row[7] else 0,
                "total_price_cny": float(row[8]) if row[8] else 0,
                "buy_url": str(row[11] or ""),
                # 箱单シートと突き合わせるキー。送料を重量で配るのに使う
                "goods_id": str(int(row[12])) if row[12] else "",
            })

    # 箱規シートから重量・容積を取得
    total_weight = 0
    total_volume = 0
    if "箱规" in wb.sheetnames:
        ws2 = wb["箱规"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                total_weight += float(row[7] or 0)
                total_volume += float(row[6] or 0)

    # 箱シート（箱规・箱单）があれば送料を重量で配れる。無ければ従来の金額比
    box_data = invoice_calc.parse_box_sheets(wb)

    return {
        "invoice_no": invoice_no,
        # 楽天版と同じく Added Value は国内送料に合算して按分対象にする
        "domestic_freight": round(domestic_freight + added_value, 2),
        "added_value": round(added_value, 2),  # 内訳確認用
        "international_freight": round(international_freight, 2),
        "total_weight": round(total_weight, 2),
        "total_volume": round(total_volume, 4),
        "items": items,
        "box_data": box_data,
        "has_box_data": box_data.get("available", False),
        # 船便のインボイスには海運用の記入要点シートが付く。確実ではないので
        # 画面の初期値としてだけ使い、ユーザーが確認・変更できるようにする
        "shipping_method": invoice_calc.guess_shipping_method(wb),
        "customs_fee_sea_jpy": invoice_calc.CUSTOMS_FEE_SEA_JPY,
    }


@router.post("/parse-import-permit")
async def parse_import_permit(file: UploadFile = File(...)):
    """輸入許可書PDFを解析して為替レート・税額等を返す"""
    content = await file.read()
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"PDF読み込みエラー: {str(e)}")

    def find_value(pattern, text, cast=str, default=None):
        m = re.search(pattern, text)
        if m:
            try:
                return cast(m.group(1).replace(",", "").strip())
            except Exception:
                pass
        return default

    exchange_rate = find_value(r'通貨レート\s+CNY\s*-\s*([\d,\.]+)', text, float, 0.0)
    customs_duty = find_value(r'関税\s*\\([\d,]+)', text, lambda x: int(x.replace(",", "")), 0)
    consumption_tax = find_value(r'消費税\s*\\([\d,]+)', text, lambda x: int(x.replace(",", "")), 0)
    local_consumption_tax = find_value(r'地方消費税\s*\\([\d,]+)', text, lambda x: int(x.replace(",", "")), 0)
    total_tax = find_value(r'納税額合計\s*\\([\d,]+)', text, lambda x: int(x.replace(",", "")), 0)
    invoice_cny = find_value(r'仕入書価格\s+[A-Z]\s+-\s+CIF\s+-\s+CNY\s+-\s+([\d,\.]+)', text, float, 0.0)
    bl_number = find_value(r'Ｂ／Ｌ番号\(1\)(\S+)', text, str, "")
    declaration_no = find_value(r'申告番号\s+([\d\s]+)', text, lambda x: x.replace(" ", ""), "")

    # 申告欄ごとの関税率（楽天と同じ共通処理）。欄が取れれば税率別計算ができる
    permit_columns = invoice_calc.parse_permit_columns(text)
    import_tax_jpy = total_tax or (customs_duty + consumption_tax + local_consumption_tax)

    return {
        "bl_number": bl_number,
        "declaration_no": declaration_no,
        "exchange_rate": exchange_rate,
        "customs_duty": customs_duty,
        "consumption_tax": consumption_tax,
        "local_consumption_tax": local_consumption_tax,
        "total_tax": total_tax,
        "import_tax_jpy": import_tax_jpy,
        "invoice_cny": invoice_cny,
        "permit_columns": permit_columns,
    }


@router.post("/validate-pair")
async def validate_pair(
    invoice_file: UploadFile = File(...),
    permit_file: UploadFile = File(...),
):
    """インボイスXLSと輸入許可書PDFのCNY合計が一致するか検証する"""
    # インボイスのCNY合計を計算
    inv_content = await invoice_file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(inv_content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"インボイス読み込みエラー: {str(e)}")

    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] == "10) Name of Commodity":
            header_row = i + 1
            break
    if not header_row:
        raise HTTPException(status_code=400, detail="インボイスの商品データが見つかりません")

    total_cny = 0.0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] and str(row[0]).startswith("MADE IN"):
            break
        if row[0] is None and row[6] and row[7]:
            qty = float(row[6]) if row[6] else 0
            price = float(row[7]) if row[7] else 0
            total_cny += qty * price

    # 輸入許可書のCNYを取得
    permit_content = await permit_file.read()
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(permit_content)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"輸入許可書読み込みエラー: {str(e)}")

    m = re.search(r'仕入書価格\s+[A-Z]\s+-\s+CIF\s+-\s+CNY\s+-\s+([\d,\.]+)', text)
    permit_cny = float(m.group(1).replace(",", "")) if m else 0.0

    total_cny = round(total_cny, 2)
    diff = abs(total_cny - permit_cny)
    ok = diff <= 1.0  # 1元以内の誤差は許容

    return {
        "ok": ok,
        "invoice_cny": total_cny,
        "permit_cny": permit_cny,
        "diff": round(diff, 2),
        "message": "照合OK" if ok else f"金額不一致（インボイス: {total_cny}元、輸入許可書: {permit_cny}元、差額: {round(diff,2)}元）",
    }


def _find_invoice_product(db: Session, item: InvoiceItemIn):
    """明細に対応する商品マスタを探す。SKU → ASIN の順で照合する。"""
    sku = (item.sku or "").strip()
    if sku:
        p = db.query(Product).filter(Product.sku == sku).first()
        if p:
            return p
    if item.asin:
        return db.query(Product).filter(Product.asin == item.asin).first()
    return None


def _build_cost_rows(data: InvoiceIn, db: Session):
    """明細ごとの原価を計算する。calculate と save で同じ結果になるよう共通化する。

    1便に楽天商品・Amazon商品・発送資材が混載されるため、両方のマスタを照合して
    便の実額（送料・税）を全明細へ配り切る。片方のマスタだけを見ていた頃は、
    相手側の明細が按分の分母には残るのに原価には反映されず、その分が消えていた。

    - permit_columns があれば申告欄ごとの税率、無ければ金額比の一律按分
    - 原価 = (小計 + 按分送料) × 為替 + 按分税額 を set_size で割った1個あたり
    - 資材(is_material)は商品原価を更新せず、資材費として別途記録する
    - どのマスタにも無い明細はカバー率として警告する
    """
    total_cny = sum(i.qty * i.unit_price_cny for i in data.items)
    total_freight = data.domestic_freight + data.international_freight
    import_tax_jpy = data.import_tax_jpy or 0
    use_tariff = bool(data.permit_columns)

    classified = invoice_calc.classify_invoice_lines(db, data.items)
    kind_by_index = {c["index"]: c["kind"] for c in classified}
    product_by_index = {c["index"]: c["product"] for c in classified}

    item_totals = [i.qty * i.unit_price_cny for i in data.items]
    coverage = invoice_calc.calc_coverage(classified, item_totals)

    # 送料は箱の実測重量で配る。金額比だと安くて嵩張るものが送料をほとんど
    # 負担しない（実測で17倍の差が出た）。箱データが無い便は金額比に落ちる
    freight_res = invoice_calc.calc_freight_by_weight(
        [i.goods_id for i in data.items], item_totals, data.box_data, total_freight,
    )
    freight_by_index = freight_res["alloc"]

    # 通関料は書類1件あたりの手続き費用なので、重量ではなく金額比で配る
    customs_fee = data.customs_fee_jpy or 0
    customs_fee_by_index = invoice_calc.calc_customs_fee_alloc(item_totals, customs_fee)

    # 税額の按分対象は全明細。資材や相手側商品を外すと、その分の税が宙に浮く
    valid_items = [(idx, total) for idx, total in enumerate(item_totals)]

    tariff_info = {}
    if use_tariff and valid_items:
        tariff_info = invoice_calc.calc_tariff_tax(
            valid_items,
            exchange_rate=data.exchange_rate,
            domestic_freight=data.domestic_freight,
            international_freight=data.international_freight,
            permit_cols_by_index={idx: data.items[idx].permit_col for idx, _ in valid_items},
            columns=data.permit_columns,
        )

    rows = []
    material_rows = []
    unknown = 0
    for idx, item in enumerate(data.items):
        item_total = item_totals[idx]
        freight_alloc = freight_by_index.get(idx, 0.0)

        ti = tariff_info.get(idx) if use_tariff else None
        if ti:
            tax_alloc_jpy = ti["total_tax_jpy"]
        else:
            tax_alloc_jpy = (item_total / total_cny * import_tax_jpy) if total_cny > 0 else 0

        kind = kind_by_index.get(idx, invoice_calc.KIND_UNKNOWN)
        product = product_by_index.get(idx)
        fee_alloc = customs_fee_by_index.get(idx, 0.0)
        total_cost_jpy = (
            (item_total + freight_alloc) * data.exchange_rate + tax_alloc_jpy + fee_alloc
        )

        if kind == invoice_calc.KIND_MATERIAL:
            # 梱包資材は売上原価だが、仕入時点ではどの商品に何枚使うか決まらないので
            # 商品ごとの原価には配らず、資材費として月次で集計する
            material_rows.append({
                "index": idx,
                "item": item,
                "product": product,
                "total_price_cny": round(item_total, 2),
                "freight_alloc_cny": round(freight_alloc, 2),
                "tax_alloc_jpy": round(tax_alloc_jpy, 1),
                "customs_fee_alloc_jpy": round(fee_alloc, 1),
                "total_cost_jpy": round(total_cost_jpy, 1),
            })
            continue

        if kind == invoice_calc.KIND_UNKNOWN:
            unknown += 1
            continue

        set_size = (product.set_size or 1) if product else 1
        sell_units = item.qty / set_size if item.qty > 0 else 0
        cost_jpy = (total_cost_jpy / sell_units) if sell_units > 0 else 0

        rows.append({
            "index": idx,
            "item": item,
            "product": product,
            "kind": kind,
            "total_price_cny": round(item_total, 2),
            "freight_alloc_cny": round(freight_alloc, 2),
            "tax_alloc_jpy": round(tax_alloc_jpy, 1),
            "customs_fee_alloc_jpy": round(fee_alloc, 1),
            "cost_per_unit_jpy": round(cost_jpy, 1),
            "set_size": set_size,
            "col_no": ti["col_no"] if ti else None,
            "tariff_rate": ti["tariff_rate"] if ti else None,
            "tariff_rate_str": ti["tariff_rate_str"] if ti else None,
            "duty_jpy": ti["duty_jpy"] if ti else None,
        })

    verification = invoice_calc.verify_allocation(
        rows, material_rows, coverage,
        total_freight_cny=total_freight,
        import_tax_jpy=import_tax_jpy,
        permit_columns=data.permit_columns,
        customs_fee_jpy=customs_fee,
    )

    return {
        "rows": rows,
        "material_rows": material_rows,
        "skipped": unknown,
        "coverage": coverage,
        "verification": verification,
        "freight_method": freight_res,
        "total_cny": total_cny,
        "total_freight_cny": total_freight,
        "import_tax_jpy": import_tax_jpy,
        "customs_fee_jpy": customs_fee,
        "use_tariff": use_tariff,
    }


@router.post("/calculate")
def calculate_cost(data: InvoiceIn, db: Session = Depends(get_db)):
    calc = _build_cost_rows(data, db)

    items = []
    for r in calc["rows"]:
        item = r["item"]
        items.append({
            **item.model_dump(),
            "name_jp": item.name_jp or (r["product"].name if r["product"] else ""),
            "matched_sku": r["product"].sku if r["product"] else "",
            "set_size": r["set_size"],
            "total_price_cny": r["total_price_cny"],
            "freight_alloc_cny": r["freight_alloc_cny"],
            "tax_alloc_jpy": r["tax_alloc_jpy"],
            "customs_fee_alloc_jpy": r["customs_fee_alloc_jpy"],
            "cost_per_unit_jpy": r["cost_per_unit_jpy"],
            "col_no": r["col_no"],
            "tariff_rate": r["tariff_rate"],
            "tariff_rate_str": r["tariff_rate_str"],
            "duty_jpy": r["duty_jpy"],
        })

    materials = [
        {
            "sku": r["item"].sku,
            "name": (r["product"].name if r["product"] else "") or r["item"].name_jp,
            "qty": r["item"].qty,
            "total_price_cny": r["total_price_cny"],
            "freight_alloc_cny": r["freight_alloc_cny"],
            "tax_alloc_jpy": r["tax_alloc_jpy"],
            "customs_fee_alloc_jpy": r["customs_fee_alloc_jpy"],
            "total_cost_jpy": r["total_cost_jpy"],
        }
        for r in calc["material_rows"]
    ]

    total_cny = calc["total_cny"]
    total_freight = calc["total_freight_cny"]
    return {
        "items": items,
        "materials": materials,
        "material_total_jpy": round(sum(m["total_cost_jpy"] for m in materials), 0),
        "coverage": calc["coverage"],
        "verification": calc["verification"],
        "freight_method": {
            "reason": calc["freight_method"]["reason"],
            "fallback": calc["freight_method"]["fallback"],
        },
        "skipped": calc["skipped"],
        "use_tariff": calc["use_tariff"],
        "total_qty": sum(r["item"].qty for r in calc["rows"]),
        "total_cny": round(total_cny, 2),
        "total_freight_cny": round(total_freight, 2),
        "import_tax_jpy": round(calc["import_tax_jpy"], 0),
        "customs_fee_jpy": calc["customs_fee_jpy"],
        "grand_total_jpy": round(
            (total_cny + total_freight) * data.exchange_rate
            + calc["import_tax_jpy"] + calc["customs_fee_jpy"], 0
        ),
    }


@router.post("/save")
def save_invoice(data: InvoiceIn, db: Session = Depends(get_db)):
    calc = _build_cost_rows(data, db)

    # 検算NGのまま保存すると、誤った原価が最新版として値付け・発注判断に使われる。
    # 保存を止めて、何が合わないかを画面に返す（force=true で承知の上の強行は可能）
    v = calc["verification"]
    if not v["ok"] and not data.force_save:
        ng = [c for c in v["checks"] if not c["ok"] and c["level"] == "error"]
        raise HTTPException(400, {
            "message": "検算に失敗したため保存しませんでした",
            "failed": ng,
            "verification": v,
        })

    invoice = Invoice(
        invoice_no=data.invoice_no,
        invoice_date=data.invoice_date,
        exchange_rate=data.exchange_rate,
        domestic_freight=data.domestic_freight,
        international_freight=data.international_freight,
        total_weight=data.total_weight,
        total_volume=data.total_volume,
        note=data.note,
        customs_duty=data.customs_duty,
        consumption_tax=data.consumption_tax,
        local_consumption_tax=data.local_consumption_tax,
        total_tax=data.total_tax,
        import_tax_jpy=calc["import_tax_jpy"],
        customs_fee_jpy=calc["customs_fee_jpy"],
        bl_number=data.bl_number,
        declaration_no=data.declaration_no,
    )
    db.add(invoice)
    db.flush()

    updated_products = 0
    updated_rakuten = 0
    for r in calc["rows"]:
        item = r["item"]
        product = r["product"]
        is_amazon = r["kind"] == invoice_calc.KIND_AMAZON

        db.add(InvoiceItem(
            invoice_id=invoice.id,
            sku=item.sku,
            # product_id は Amazon products への外部キーなので、楽天商品ならNoneにする
            product_id=product.id if (product and is_amazon) else None,
            name_cn=item.name_cn,
            name_jp=item.name_jp or (product.name if product else ""),
            qty=item.qty,
            unit_price_cny=item.unit_price_cny,
            total_price_cny=r["total_price_cny"],
            freight_alloc_cny=r["freight_alloc_cny"],
            cost_per_unit_jpy=r["cost_per_unit_jpy"],
            buy_url=item.buy_url,
            tax_alloc_jpy=r["tax_alloc_jpy"],
            customs_fee_alloc_jpy=r["customs_fee_alloc_jpy"],
            duty_jpy=r["duty_jpy"] or 0,
            col_no=r["col_no"],
            tariff_rate=r["tariff_rate"] or 0,
        ))

        # 円建て原価は cost_jpy に入れる。price は元単価のまま残す
        # （price を上書きすると発注管理の「単価(元)」が円に化けるため）
        # 同じ便に混載された楽天商品も、楽天マスタ側の原価を更新する
        if product:
            product.cost_jpy = r["cost_per_unit_jpy"]
            if item.unit_price_cny:
                product.price = item.unit_price_cny
            if is_amazon:
                updated_products += 1
            else:
                updated_rakuten += 1

    # 発送資材は商品原価に載せず、資材費として月次集計用に記録する
    from app.models.material_cost import MaterialCost
    for r in calc["material_rows"]:
        item = r["item"]
        product = r["product"]
        db.add(MaterialCost(
            invoice_no=data.invoice_no,
            invoice_date=data.invoice_date,
            source="amazon",
            sku=item.sku,
            name=(product.name if product else "") or item.name_jp,
            qty=item.qty,
            unit_price_cny=item.unit_price_cny,
            total_price_cny=r["total_price_cny"],
            freight_alloc_cny=r["freight_alloc_cny"],
            tax_alloc_jpy=r["tax_alloc_jpy"],
            customs_fee_alloc_jpy=r["customs_fee_alloc_jpy"],
            total_cost_jpy=r["total_cost_jpy"],
            exchange_rate=data.exchange_rate,
        ))

    db.commit()
    return {
        "invoice_id": invoice.id,
        "updated_products": updated_products,
        "updated_rakuten": updated_rakuten,
        "material_count": len(calc["material_rows"]),
        "coverage": calc["coverage"],
        "skipped": calc["skipped"],
    }


@router.get("/")
def list_invoices(db: Session = Depends(get_db)):
    invoices = db.query(Invoice).order_by(Invoice.created_at.desc()).all()
    return invoices


@router.get("/{invoice_id}/items")
def get_invoice_items(invoice_id: int, db: Session = Depends(get_db)):
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).all()
    return items
