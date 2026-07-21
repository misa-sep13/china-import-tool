from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import date, datetime, timezone, timedelta
import csv, io, json
import uuid, time, threading
import re
from pydantic import BaseModel
import asyncio
from app.core.database import get_db, SessionLocal
from app.services.rakuten_rms import calc_set_avail, build_component_share_counts

# 日本時間(JST)。RenderはUTCで動くため、表示用の時刻(sales_updated_at, last_sync)は
# JSTの壁時計時刻(タイムゾーンなし)で保存し、画面や報告で日本時間として正しく見えるようにする。
JST = timezone(timedelta(hours=9))
def _now_jst():
    return datetime.now(JST).replace(tzinfo=None)
from app.models.rakuten_product import RakutenProduct
from app.models.rakuten_order import RakutenOrderHistory
from app.models.rakuten_settings import RakutenSettings
from app.models.inventory_reflection_log import InventoryReflectionLog
from app.models.rakuten_sales import RakutenSalesImport, RakutenSalesSummary
from app.services.rakuten_calc import calc_rakuten_order, RakutenCalcSettings
from app.services.rakuten_sales_import import (
    build_sales_summary,
    find_table_rows,
    read_upload_table,
)

router = APIRouter(prefix="/rakuten", tags=["rakuten"])


def _log_inventory_reflection(db: Session, **kwargs) -> InventoryReflectionLog:
    log = InventoryReflectionLog(**kwargs)
    db.add(log)
    return log


@router.get("/inventory-reflection-logs")
def get_inventory_reflection_logs(limit: int = 100, db: Session = Depends(get_db)):
    limit = max(1, min(limit, 500))
    logs = (
        db.query(InventoryReflectionLog)
        .order_by(InventoryReflectionLog.created_at.desc(), InventoryReflectionLog.id.desc())
        .limit(limit)
        .all()
    )
    return {
        "logs": [
            {
                "id": l.id,
                "event_id": l.event_id,
                "source": l.source,
                "source_label": l.source_label,
                "source_id": l.source_id,
                "source_ref": l.source_ref,
                "sku": l.sku,
                "name": l.name,
                "supplier": l.supplier,
                "received_qty": l.received_qty,
                "stock_before": l.stock_before,
                "stock_after": l.stock_after,
                "inbound_before": l.inbound_before,
                "inbound_after": l.inbound_after,
                "standard_stock_before": l.standard_stock_before,
                "standard_stock_after": l.standard_stock_after,
                "rms_push_items": l.rms_push_items,
                "note": l.note,
                "created_at": l.created_at.isoformat() if l.created_at else None,
            }
            for l in logs
        ]
    }


@router.post("/migrate-show-in-orders")
def migrate_show_in_orders(db: Session = Depends(get_db)):
    """一時マイグレーション: show_in_ordersカラム追加＋初期値設定（実行後に削除予定）"""
    from sqlalchemy import text
    try:
        db.execute(text("ALTER TABLE rakuten_products ADD COLUMN show_in_orders BOOLEAN DEFAULT TRUE"))
        db.commit()
        added = True
    except Exception as e:
        db.rollback()
        if "already exists" in str(e).lower() or "duplicate" in str(e).lower():
            added = False
        else:
            raise HTTPException(500, str(e))
    count = db.query(RakutenProduct).filter(RakutenProduct.is_component == True).count()
    db.execute(text("UPDATE rakuten_products SET show_in_orders = FALSE WHERE is_component = TRUE"))
    db.commit()
    return {"column_added": added, "updated_count": count}


def _clean_set_components(sc: Optional[str]) -> Optional[str]:
    """set_componentsのJSON文字列から空エントリ（sku/buy_url/supplier_specが全て空）を除去。全削除なら None を返す。"""
    if not sc:
        return sc
    try:
        items = json.loads(sc)
        filtered = [it for it in items if it.get("sku") or it.get("buy_url") or it.get("supplier_spec")]
        return json.dumps(filtered, ensure_ascii=False) if filtered else None
    except Exception:
        return sc


def _is_taotaro_supplier(supplier: Optional[str]) -> bool:
    value = (supplier or "").strip().lower().replace(" ", "").replace("　", "")
    return value in {"タオタロウ", "タオ太郎", "taotaro"} or "タオタロウ" in value or "タオ太郎" in value


def _is_manufacturer_product(p: RakutenProduct) -> bool:
    supplier = (p.supplier or "").strip()
    if supplier:
        return not _is_taotaro_supplier(supplier)
    sku = (p.sku or "").strip().lower()
    return bool(re.match(r"^s\d", sku) or re.match(r"^\d+$", sku))


def _parse_components_for_stock(p: RakutenProduct):
    try:
        return json.loads(p.set_components or "[]")
    except Exception:
        return []


def _recalc_dependent_set_stock(all_products: list[RakutenProduct], sku_stock: dict, updated_skus: set[str]):
    share_counts = build_component_share_counts(all_products)
    for p in all_products:
        comps = _parse_components_for_stock(p)
        if not comps or not any(c.get("sku") in updated_skus for c in comps):
            continue
        req: dict[str, int] = {}
        for c in comps:
            c_sku = c.get("sku")
            c_qty = c.get("qty") or 1
            if c_sku:
                req[c_sku] = req.get(c_sku, 0) + c_qty
        set_qty = None
        for c_sku, c_qty in req.items():
            avail = calc_set_avail(sku_stock.get(c_sku, 0), c_qty, share_counts.get(c_sku, 0))
            set_qty = avail if set_qty is None else min(set_qty, avail)
        if set_qty is not None:
            p.stock = set_qty
            sku_stock[p.sku] = set_qty


def _build_rms_stock_items(all_products: list[RakutenProduct], sku_stock: dict, updated_skus: set[str]):
    rms_items = []
    for p in all_products:
        # RMSにページが無い内部SKU（y91_case等）はpush対象外
        if p.is_component and not p.rakuten_item_url:
            continue
        sku = (p.sku or "").strip()
        if not sku or not re.match(r'^[a-zA-Z0-9_\-]+$', sku):
            continue
        comps = _parse_components_for_stock(p)
        if sku in updated_skus or (comps and any(c.get("sku") in updated_skus for c in comps)):
            manage_number = p.rakuten_item_url or sku.split("_")[0]
            rms_items.append({"manage_number": manage_number, "variant_id": sku, "quantity": sku_stock.get(sku, 0)})
    return rms_items



# ============================================================
# Settings
# ============================================================

class RakutenSettingsSchema(BaseModel):
    lead_days:          int   = 20
    target_days:        int   = 30
    safety_stock_rate:  float = 0.10
    threshold_days:     int   = 60
    order_qty_cap:      int   = 3
    super_sale_enabled: bool  = False
    super_sale_mode:    str   = 'A'
    super_sale_start:   Optional[date] = None
    super_sale_end:     Optional[date] = None
    commission_rate:    float = 0.09
    default_shipping_fee: int = 180
    rms_service_secret: Optional[str] = None
    rms_license_key:    Optional[str] = None
    rms_key_expires_at: Optional[date] = None

    class Config:
        from_attributes = True

def _get_or_create_settings(db: Session) -> RakutenSettings:
    row = db.query(RakutenSettings).first()
    if not row:
        row = RakutenSettings()
        db.add(row)
        db.commit()
        db.refresh(row)
    return row

@router.get("/settings", response_model=RakutenSettingsSchema)
def get_settings(db: Session = Depends(get_db)):
    return _get_or_create_settings(db)

@router.put("/settings", response_model=RakutenSettingsSchema)
def update_settings(data: RakutenSettingsSchema, db: Session = Depends(get_db)):
    row = _get_or_create_settings(db)
    for k, v in data.model_dump().items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


# ============================================================
# Sales Management（月次売上管理・手動アップロード）
# ============================================================

def _validate_sales_period(period: str) -> str:
    value = (period or "").strip()
    if not re.match(r"^\d{4}-\d{2}$", value):
        raise HTTPException(400, "対象月は YYYY-MM 形式で指定してください。")
    return value


async def _read_sales_upload(file: UploadFile | None, candidates: list[list[str]]) -> tuple[list[dict], str | None]:
    if not file or not file.filename:
        return [], None
    data = await file.read()
    try:
        rows = read_upload_table(file.filename, data)
        last_error = None
        for required in candidates:
            try:
                return find_table_rows(rows, required), file.filename
            except Exception as e:
                last_error = e
        raise last_error or ValueError("テーブルを読み取れませんでした")
    except Exception as e:
        raise HTTPException(400, f"{file.filename} の読み込みに失敗しました: {e}")


def _sales_summary_out(row: RakutenSalesSummary) -> dict:
    return {
        "id": row.id,
        "period": row.period,
        "level": row.level,
        "product_key": row.product_key,
        "sku_key": row.sku_key,
        "product_name": row.product_name,
        "units": row.units or 0,
        "sales": row.sales or 0,
        "point_cost": row.point_cost or 0,
        "all_coupon": row.all_coupon or 0,
        "store_coupon": row.store_coupon or 0,
        "coupon_fee": row.coupon_fee or 0,
        "rpp_cost": row.rpp_cost or 0,
        "coupon_ad_cost": row.coupon_ad_cost or 0,
        "affiliate_cost": row.affiliate_cost or 0,
        "affiliate_fee": row.affiliate_fee or 0,
        "sales_store_coupon_excluded": row.sales_store_coupon_excluded or 0,
        "sales_all_coupon_excluded": row.sales_all_coupon_excluded or 0,
        "pc_sales": row.pc_sales or 0,
        "mobile_sales": row.mobile_sales or 0,
        "platform_fee": row.platform_fee or 0,
        "platform_fee_rate": row.platform_fee_rate,
        "shipping_cost": row.shipping_cost or 0,
        "product_cost": row.product_cost or 0,
        "profit": row.profit or 0,
        "profit_rate": row.profit_rate,
        "rpp_rate": row.rpp_rate,
        "ad_rate": row.ad_rate,
    }


def _sales_import_out(row: RakutenSalesImport) -> dict:
    return {
        "id": row.id,
        "period": row.period,
        "order_file_name": row.order_file_name,
        "rpp_file_name": row.rpp_file_name,
        "coupon_ad_file_name": row.coupon_ad_file_name,
        "affiliate_file_name": row.affiliate_file_name,
        "order_rows": row.order_rows or 0,
        "rpp_rows": row.rpp_rows or 0,
        "coupon_ad_rows": row.coupon_ad_rows or 0,
        "affiliate_rows": row.affiliate_rows or 0,
        "total_units": row.total_units or 0,
        "total_sales": row.total_sales or 0,
        "total_profit": row.total_profit or 0,
        "status": row.status,
        "message": row.message,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


@router.get("/sales/months")
def list_sales_months(db: Session = Depends(get_db)):
    rows = (
        db.query(RakutenSalesImport)
        .order_by(RakutenSalesImport.period.desc(), RakutenSalesImport.id.desc())
        .all()
    )
    return {"months": [_sales_import_out(r) for r in rows]}


@router.get("/sales/summary")
def get_sales_summary(period: str, level: str = "parent", db: Session = Depends(get_db)):
    period = _validate_sales_period(period)
    if level not in {"parent", "sku"}:
        raise HTTPException(400, "level は parent または sku を指定してください。")
    info = db.query(RakutenSalesImport).filter(RakutenSalesImport.period == period).first()
    rows = (
        db.query(RakutenSalesSummary)
        .filter(RakutenSalesSummary.period == period, RakutenSalesSummary.level == level)
        .order_by(RakutenSalesSummary.sales.desc(), RakutenSalesSummary.product_key.asc(), RakutenSalesSummary.sku_key.asc())
        .all()
    )
    totals = {
        "units": round(sum((r.units or 0) for r in rows), 2),
        "sales": round(sum((r.sales or 0) for r in rows), 2),
        "profit": round(sum((r.profit or 0) for r in rows), 2),
    }
    totals["profit_rate"] = round(totals["profit"] / totals["sales"] * 100, 2) if totals["sales"] else None
    return {
        "import": _sales_import_out(info) if info else None,
        "totals": totals,
        "rows": [_sales_summary_out(r) for r in rows],
    }


@router.post("/sales/import")
async def import_sales_month(
    period: str = Form(...),
    order_file: list[UploadFile] = File(...),
    rpp_file: Optional[UploadFile] = File(None),
    coupon_ad_file: Optional[UploadFile] = File(None),
    affiliate_file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    period = _validate_sales_period(period)
    order_rows = []
    order_names = []
    for f in order_file:
        rows, name = await _read_sales_upload(
            f,
            [["注文番号", "ステータス", "商品管理番号", "単価", "個数"]],
        )
        order_rows.extend(rows)
        if name:
            order_names.append(name)
    order_name = " + ".join(order_names) if order_names else None
    rpp_rows, rpp_name = await _read_sales_upload(
        rpp_file,
        [["商品管理番号", "実績額(合計)"], ["商品管理番号", "実績額"], ["商品ページURL", "実績額(合計)"]],
    )
    coupon_ad_rows, coupon_ad_name = await _read_sales_upload(
        coupon_ad_file,
        [
            ["商品管理番号", "実績額(合計)"],
            ["商品管理番号", "実績額"],
            ["商品管理番号", "広告費"],
            ["商品管理番号（URL）", "実績額"],
            ["商品ページURL", "実績額"],
        ],
    )
    affiliate_rows, affiliate_name = await _read_sales_upload(
        affiliate_file,
        [["成果発生日時", "商品管理番号", "成果報酬"], ["date", "item_mng_id", "rewards"], ["受注番号", "商品管理番号", "成果報酬"]],
    )

    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    settings = _get_or_create_settings(db)
    built = build_sales_summary(
        period=period,
        products=products,
        settings=settings,
        order_rows=order_rows,
        rpp_rows=rpp_rows,
        coupon_ad_rows=coupon_ad_rows,
        affiliate_rows=affiliate_rows,
    )

    try:
        existing = db.query(RakutenSalesImport).filter(RakutenSalesImport.period == period).first()
        if existing:
            info = existing
        else:
            info = RakutenSalesImport(period=period)
            db.add(info)
        info.order_file_name = order_name
        info.rpp_file_name = rpp_name
        info.coupon_ad_file_name = coupon_ad_name
        info.affiliate_file_name = affiliate_name
        info.order_rows = len(order_rows)
        info.rpp_rows = len(rpp_rows)
        info.coupon_ad_rows = len(coupon_ad_rows)
        info.affiliate_rows = len(affiliate_rows)
        info.total_units = built["totals"]["units"]
        info.total_sales = built["totals"]["sales"]
        info.total_profit = built["totals"]["profit"]
        info.status = "completed"
        info.message = f"受注スキップ {built['skipped_orders']}件"

        db.query(RakutenSalesSummary).filter(RakutenSalesSummary.period == period).delete(synchronize_session=False)
        for data in built["parent_rows"] + built["sku_rows"]:
            db.add(RakutenSalesSummary(**data))
        db.commit()
        db.refresh(info)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"売上集計の保存に失敗しました: {e}")

    return {
        "import": _sales_import_out(info),
        "totals": built["totals"],
        "parent_count": len(built["parent_rows"]),
        "sku_count": len(built["sku_rows"]),
        "skipped_orders": built["skipped_orders"],
    }


@router.patch("/sales/change-period")
async def change_sales_period(
    from_period: str = Query(...),
    to_period: str = Query(...),
    db: Session = Depends(get_db),
):
    from_period = _validate_sales_period(from_period)
    to_period = _validate_sales_period(to_period)
    existing_to = db.query(RakutenSalesImport).filter(RakutenSalesImport.period == to_period).first()
    if existing_to:
        raise HTTPException(400, f"{to_period} には既にデータがあります。先に削除してください。")
    info = db.query(RakutenSalesImport).filter(RakutenSalesImport.period == from_period).first()
    if not info:
        raise HTTPException(404, f"{from_period} のデータが見つかりません")
    info.period = to_period
    updated = db.query(RakutenSalesSummary).filter(RakutenSalesSummary.period == from_period).update(
        {"period": to_period}, synchronize_session=False
    )
    db.commit()
    return {"message": f"{from_period} → {to_period} に変更しました（{updated}行）"}


# ============================================================
# Products (商品マスタ)
# ============================================================

class RakutenProductIn(BaseModel):
    sku:              str
    name:             Optional[str] = None
    jan_code:         Optional[str] = None
    buy_url:          Optional[str] = None
    supplier_spec:    Optional[str] = None
    price:            Optional[float] = None
    spec:             Optional[str] = None
    set_size:         int = 1
    rakuten_item_url: Optional[str] = None
    rakuten_sku_id:   Optional[str] = None
    supplier:         Optional[str] = None
    standard_stock:   int = 0
    stock:            int = 0
    inbound:          int = 0
    sales_30_recent:  int = 0
    sales_30_prev:    int = 0
    cost_jpy:         Optional[float] = None
    selling_price:    Optional[float] = None
    shipping_fee:     int = 180
    customer_memo:    Optional[str] = None
    notes:            Optional[str] = None
    memo:             Optional[str] = None
    set_components:   Optional[str] = None  # JSON文字列（在庫連動用）
    purchase_components: Optional[str] = None  # JSON文字列（発注用付属品・在庫連動しない）
    is_component:     bool = False          # 単品（セット構成用内部管理）フラグ
    is_active:        bool = True

class RakutenProductOut(RakutenProductIn):
    id:               int
    sales_updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True

@router.get("/products")
def list_products(db: Session = Depends(get_db)):
    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).order_by(RakutenProduct.sku.asc()).all()
    return [RakutenProductOut.model_validate(p).model_dump() for p in products]

@router.get("/stock")
def list_stock(db: Session = Depends(get_db)):
    """在庫・損益一覧（バリエーション商品のみ、is_component=Falseを対象）"""
    settings = _get_or_create_settings(db)
    commission_rate = settings.commission_rate or 0.09
    products = (
        db.query(RakutenProduct)
        .filter(RakutenProduct.is_active == True, RakutenProduct.is_component == False)
        .order_by(RakutenProduct.sku.asc())
        .all()
    )
    ordered1_by_sku, ordered2_by_sku = _ordered_by_sku_stage(db)
    result = []
    for p in products:
        order_managed = bool((p.buy_url or "").strip()) and not _is_manufacturer_product(p)
        inbound = ordered1_by_sku.get(p.sku, 0) if order_managed else (p.inbound or 0)
        standard_stock = ordered2_by_sku.get(p.sku, 0) if order_managed else (p.standard_stock or 0)
        selling_price = p.selling_price
        cost_jpy = p.cost_jpy
        shipping_fee = p.shipping_fee if p.shipping_fee is not None else 180
        commission = round(selling_price * commission_rate, 0) if selling_price else None
        profit = round(selling_price - (cost_jpy or 0) - (commission or 0) - shipping_fee, 0) if selling_price else None
        profit_rate = round(profit / selling_price * 100, 1) if (selling_price and profit is not None) else None
        result.append({
            "id": p.id,
            "sku": p.sku,
            "name": p.name,
            "spec": p.spec,
            "customer_memo": p.customer_memo,
            "stock": p.stock,
            "inbound": inbound,
            "standard_stock": standard_stock,
            "sales_30_recent": p.sales_30_recent,
            "sales_30_prev": p.sales_30_prev,
            "selling_price": selling_price,
            "cost_jpy": cost_jpy,
            "shipping_fee": shipping_fee,
            "commission": commission,
            "commission_rate": commission_rate,
            "profit": profit,
            "profit_rate": profit_rate,
            "notes": p.notes,
            "supplier": p.supplier,
            "set_components": p.set_components,
            "is_manufacturer": _is_manufacturer_product(p),
        })
    return result

@router.post("/products")
def create_product(data: RakutenProductIn, db: Session = Depends(get_db)):
    data.set_components = _clean_set_components(data.set_components)
    existing = db.query(RakutenProduct).filter(RakutenProduct.sku == data.sku).first()
    if existing:
        if existing.is_active:
            raise HTTPException(400, "SKUが既に存在します")
        # 論理削除済みの場合は復活させて更新
        for k, v in data.model_dump().items():
            setattr(existing, k, v)
        existing.is_active = True
        db.commit()
        db.refresh(existing)
        return RakutenProductOut.model_validate(existing).model_dump()
    p = RakutenProduct(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return RakutenProductOut.model_validate(p).model_dump()

@router.put("/products/{product_id}")
async def update_product(product_id: int, data: RakutenProductIn, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    if "set_components" in data.model_fields_set:
        data.set_components = _clean_set_components(data.set_components)
    p = db.query(RakutenProduct).filter(RakutenProduct.id == product_id).first()
    if not p:
        raise HTTPException(404, "商品が見つかりません")
    dup = db.query(RakutenProduct).filter(
        RakutenProduct.sku == data.sku, RakutenProduct.id != product_id
    ).first()
    if dup:
        raise HTTPException(400, "SKUが既に存在します")

    old_stock = p.stock
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)

    # 在庫数が送信された場合はRMSにも反映（値が同じでもセット商品の再計算が必要なため）
    if "stock" in data.model_fields_set and p.sku:
        try:
            settings = db.query(RakutenSettings).first()
            if settings and settings.rms_service_secret and settings.rms_license_key:
                from app.services.rakuten_rms import push_inventory_to_rms
                rms_items = []

                # 自分自身をRMSに反映
                manage_number = p.rakuten_item_url or p.sku.split("_")[0]
                rms_items.append({"manage_number": manage_number, "variant_id": p.sku, "quantity": p.stock})

                # この商品を参照しているセット商品の在庫も自動計算して反映（is_component問わず）
                set_products = db.query(RakutenProduct).filter(
                    RakutenProduct.is_active == True,
                    RakutenProduct.set_components != None,
                ).all()

                # 全商品の現在在庫をまとめて取得
                all_skus = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
                sku_stock = {cp.sku: (cp.stock or 0) for cp in all_skus}
                sku_stock[p.sku] = p.stock  # 今回更新した値で上書き
                share_counts = build_component_share_counts(set_products)

                for sp in set_products:
                    if sp.id == p.id:
                        continue
                    try:
                        comps = json.loads(sp.set_components or "[]")
                    except Exception:
                        continue
                    if not comps:
                        continue
                    # このセット商品が今回更新した商品を含む場合のみ計算
                    if not any(c.get("sku") == p.sku for c in comps):
                        continue
                    # セット在庫 = min(構成品在庫 ÷ 使用数) の切り捨て
                    set_qty = None
                    for c in comps:
                        c_sku = c.get("sku")
                        c_qty = c.get("qty") or 1
                        if not c_sku:
                            continue
                        avail = calc_set_avail(sku_stock.get(c_sku, 0), c_qty, share_counts.get(c_sku, 0))
                        set_qty = avail if set_qty is None else min(set_qty, avail)

                    if set_qty is not None:
                        sp.stock = set_qty
                        sp_manage = sp.rakuten_item_url or sp.sku.split("_")[0]
                        rms_items.append({"manage_number": sp_manage, "variant_id": sp.sku, "quantity": set_qty})

                db.commit()

                # RMSへのPUTは時間がかかるため、レスポンスを待たせずバックグラウンドで実行する
                if rms_items:
                    background_tasks.add_task(
                        push_inventory_to_rms,
                        settings.rms_service_secret, settings.rms_license_key, rms_items,
                    )
        except Exception as e:
            import logging
            logging.getLogger("rakuten").warning(f"RMS在庫反映エラー ({p.sku}): {e}")

    return RakutenProductOut.model_validate(p).model_dump()

@router.post("/products/bulk-update-stock")
async def bulk_update_stock(body: dict, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """複数商品の在庫をまとめて更新してRMSに一括反映
    body: {"updates": [{"id": 1, "stock": 10, "inbound": 0, "standard_stock": 5}, ...]}
    RMSへの在庫反映はバックグラウンドで実行し、保存レスポンスは即座に返す。
    """
    updates = body.get("updates", [])
    if not updates:
        return {"ok": True, "updated": 0}

    settings = _get_or_create_settings(db)
    all_products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    id_to_product = {p.id: p for p in all_products}
    sku_stock = {p.sku: (p.stock or 0) for p in all_products}

    # Step1: DB更新（全件まとめて）
    updated_skus = set()
    for u in updates:
        p = id_to_product.get(u.get("id"))
        if not p:
            continue
        if "stock" in u:
            p.stock = int(u["stock"])
            sku_stock[p.sku] = p.stock
            updated_skus.add(p.sku)
        if "inbound" in u:
            p.inbound = int(u["inbound"])
        if "standard_stock" in u:
            p.standard_stock = int(u["standard_stock"])

    # Step2: セット商品の在庫を構成品から再計算
    def _parse(p):
        try: return json.loads(p.set_components or "[]")
        except: return []

    rms_items = []
    share_counts = build_component_share_counts(all_products)
    for p in all_products:
        comps = _parse(p)
        if not comps:
            continue
        if not any(c.get("sku") in updated_skus for c in comps):
            continue
        req: dict[str, int] = {}
        for c in comps:
            c_sku = c.get("sku")
            c_qty = c.get("qty") or 1
            if c_sku:
                req[c_sku] = req.get(c_sku, 0) + c_qty
        set_qty = None
        for c_sku, c_qty in req.items():
            avail = calc_set_avail(sku_stock.get(c_sku, 0), c_qty, share_counts.get(c_sku, 0))
            set_qty = avail if set_qty is None else min(set_qty, avail)
        if set_qty is not None:
            p.stock = set_qty
            sku_stock[p.sku] = set_qty

    db.commit()

    # Step3: RMSに反映するitemsを組み立て（更新したSKU＋影響したセット商品）。
    # 実在庫を更新していない場合（発注済1/2のみ等）はupdated_skusが空なのでpushは発生しない。
    if settings and settings.rms_service_secret and settings.rms_license_key and updated_skus:
        for p in all_products:
            if p.sku in updated_skus or (p.set_components and any(c.get("sku") in updated_skus for c in _parse(p))):
                manage_number = p.rakuten_item_url or p.sku.split("_")[0]
                rms_items.append({"manage_number": manage_number, "variant_id": p.sku, "quantity": sku_stock.get(p.sku, 0)})
        if rms_items:
            # RMSへのPUTは時間がかかるため、保存レスポンスを待たせずバックグラウンドで実行する
            from app.services.rakuten_rms import push_inventory_to_rms
            background_tasks.add_task(
                push_inventory_to_rms,
                settings.rms_service_secret, settings.rms_license_key, rms_items,
            )

    return {"ok": True, "updated": len(updated_skus), "rms_pushed": len(rms_items)}


@router.post("/products/{product_id}/receive-manufacturer")
async def receive_manufacturer_stock(product_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """メーカー品の入荷処理。
    発注済1(inbound)を実在庫へ加算し、発注済2(旧standard_stock)を発注済1へ繰り上げる。
    輸入品は発注管理・インボイス取込が正規ルートなので、この操作では扱わない。
    """
    p = db.query(RakutenProduct).filter(
        RakutenProduct.id == product_id,
        RakutenProduct.is_active == True,
    ).first()
    if not p:
        raise HTTPException(404, "商品が見つかりません")
    if not _is_manufacturer_product(p):
        raise HTTPException(400, "メーカー品のみ在庫損益から入荷できます")

    received_qty = p.inbound or 0
    if received_qty <= 0:
        raise HTTPException(400, "発注済1が0のため入荷できません")

    next_inbound = p.standard_stock or 0
    before = {
        "stock": p.stock or 0,
        "inbound": p.inbound or 0,
        "standard_stock": p.standard_stock or 0,
    }

    all_products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    sku_stock = {item.sku: (item.stock or 0) for item in all_products}

    p.stock = (p.stock or 0) + received_qty
    p.inbound = next_inbound
    p.standard_stock = 0
    sku_stock[p.sku] = p.stock
    updated_skus = {p.sku}

    _recalc_dependent_set_stock(all_products, sku_stock, updated_skus)
    event_id = str(uuid.uuid4())
    _log_inventory_reflection(
        db,
        event_id=event_id,
        source="manufacturer_receive",
        source_label="メーカー入荷",
        source_id=p.id,
        source_ref="在庫・損益",
        sku=p.sku,
        name=p.name,
        supplier=p.supplier,
        received_qty=received_qty,
        stock_before=before["stock"],
        stock_after=p.stock or 0,
        inbound_before=before["inbound"],
        inbound_after=p.inbound or 0,
        standard_stock_before=before["standard_stock"],
        standard_stock_after=p.standard_stock or 0,
        rms_push_items=0,
        note="在庫・損益の入荷ボタンから反映",
    )
    db.commit()
    db.refresh(p)

    settings = _get_or_create_settings(db)
    rms_items = []
    if settings and settings.rms_service_secret and settings.rms_license_key:
        rms_items = _build_rms_stock_items(all_products, sku_stock, updated_skus)
        if rms_items:
            from app.services.rakuten_rms import push_inventory_to_rms
            background_tasks.add_task(
                push_inventory_to_rms,
                settings.rms_service_secret, settings.rms_license_key, rms_items,
            )
            db.query(InventoryReflectionLog).filter(
                InventoryReflectionLog.event_id == event_id,
            ).update({"rms_push_items": len(rms_items)})
            db.commit()

    return {
        "ok": True,
        "sku": p.sku,
        "received_qty": received_qty,
        "before": before,
        "after": {
            "stock": p.stock or 0,
            "inbound": p.inbound or 0,
            "standard_stock": p.standard_stock or 0,
        },
        "rms_pushed": len(rms_items),
    }


@router.post("/products/bulk-set-components")
def bulk_set_components(body: dict, db: Session = Depends(get_db)):
    """SKUをキー、set_componentsをJSONとして受け取り一括更新"""
    # body: {"updates": [{"sku": "y76_b-b", "set_components": "[...]"}, ...]}
    updates = body.get("updates", [])
    ok = 0
    for item in updates:
        sku = item.get("sku")
        comps = item.get("set_components")
        if not sku or comps is None:
            continue
        p = db.query(RakutenProduct).filter(RakutenProduct.sku == sku, RakutenProduct.is_active == True).first()
        if p:
            p.set_components = comps
            ok += 1
    db.commit()
    return {"updated": ok}

@router.delete("/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    p = db.query(RakutenProduct).filter(RakutenProduct.id == product_id).first()
    if not p:
        raise HTTPException(404)
    p.is_active = False
    db.commit()
    return {"ok": True}

@router.patch("/products/{product_id}/stock")
def update_stock(product_id: int, body: dict, db: Session = Depends(get_db)):
    p = db.query(RakutenProduct).filter(RakutenProduct.id == product_id).first()
    if not p:
        raise HTTPException(404)
    if "stock" in body:
        p.stock = body["stock"]
    if "inbound" in body:
        p.inbound = body["inbound"]
    if "sales_30_recent" in body:
        p.sales_30_recent = body["sales_30_recent"]
    if "sales_30_prev" in body:
        p.sales_30_prev = body["sales_30_prev"]
        p.sales_updated_at = _now_jst()
    db.commit()
    return {"ok": True}


# ============================================================
# Order Recommendations（発注推奨リスト）
# ============================================================

def _ordered_by_sku_stage(db):
    """未納品の発注数をSKU×ステージ（発注済1/発注済2）ごとに集計"""
    rows = (
        db.query(RakutenOrderHistory.sku, RakutenOrderHistory.stage, func.sum(RakutenOrderHistory.qty))
        .filter(RakutenOrderHistory.is_delivered == False, RakutenOrderHistory.is_deleted == False)
        .group_by(RakutenOrderHistory.sku, RakutenOrderHistory.stage)
        .all()
    )
    o1, o2 = {}, {}
    for sku, stage, qty in rows:
        target = o2 if stage == 2 else o1  # stageがNULLの既存データは発注済1扱い
        target[sku] = target.get(sku, 0) + (qty or 0)
    return o1, o2


def _daily_avg_from_table(db) -> dict[str, dict]:
    """rakuten_daily_salesから7日・30日のSKU別日販を算出（在庫切れ日を除外）。
    戻り値: {sku: {"avg_7": float, "avg_30": float}}
    """
    from app.models.rakuten_daily_sales import RakutenDailySales
    today = date.today()
    cutoff_30 = today - timedelta(days=30)
    cutoff_7 = today - timedelta(days=7)
    rows = db.query(
        RakutenDailySales.sku,
        RakutenDailySales.sale_date,
        RakutenDailySales.qty,
        RakutenDailySales.is_stockout,
    ).filter(RakutenDailySales.sale_date >= cutoff_30).all()
    if not rows:
        return {}
    sku_data: dict[str, dict] = {}
    for sku, sale_date, qty, is_stockout in rows:
        if sku not in sku_data:
            sku_data[sku] = {"sum_7": 0, "sum_30": 0, "days_7": 0, "days_30": 0}
        if not is_stockout:
            sku_data[sku]["sum_30"] += qty or 0
            sku_data[sku]["days_30"] += 1
            if sale_date >= cutoff_7:
                sku_data[sku]["sum_7"] += qty or 0
                sku_data[sku]["days_7"] += 1
        elif sale_date >= cutoff_7:
            pass  # stockout日は7日カウントにも含めない
    result = {}
    for sku, d in sku_data.items():
        eff_7 = max(d["days_7"], 1)
        eff_30 = max(d["days_30"], 1)
        result[sku] = {
            "avg_7": round(d["sum_7"] / eff_7, 2),
            "avg_30": round(d["sum_30"] / eff_30, 2),
        }
    return result


@router.get("/orders/recommendations")
def get_recommendations(db: Session = Depends(get_db)):
    settings_row = _get_or_create_settings(db)
    sku_daily_avgs = _daily_avg_from_table(db)
    s = RakutenCalcSettings(
        lead_days=settings_row.lead_days,
        target_days=settings_row.target_days,
        safety_stock_rate=settings_row.safety_stock_rate,
        threshold_days=settings_row.threshold_days,
    )

    # 発注済み（未納品）の数量をSKU×ステージごとに集計
    ordered1_by_sku, ordered2_by_sku = _ordered_by_sku_stage(db)

    # 全商品を取得
    all_products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    sku_to_product = {p.sku: p for p in all_products}

    # 「親発注品」= is_component=False かつ buy_url あり かつ set_components あり
    # （例: y34=3本セット、y15=天使の羽+風船セット）→ セットごと仕入れる商品
    parent_orders: dict[str, RakutenProduct] = {
        p.sku: p for p in all_products
        if not p.is_component
        and (p.buy_url or "").strip()
        and p.set_components
    }

    # 親発注品のset_components内SKU → 個別に推奨リストへ出さない
    parent_comp_skus: set[str] = set()
    for p in parent_orders.values():
        try:
            comps = json.loads(p.set_components or "[]")
            parent_comp_skus.update(c["sku"] for c in comps if c.get("sku"))
        except Exception:
            pass

    # セット商品の販売実績を構成単品SKUへ按分（親発注品は除く）
    # バリエーション（例: y76_b-w）が参照する単品（例: y76_black）は、
    # is_component=False（一覧に表示する単品）でも日販計算には合算する
    unit_sales: dict[str, dict] = {}
    referenced_skus: set[str] = set()

    for p in all_products:
        if not p.set_components:
            continue
        if p.sku in parent_orders:
            continue  # 親発注品は後で別処理
        try:
            comps = json.loads(p.set_components or "[]")
        except Exception:
            comps = []
        for c in comps:
            unit_sku = c.get("sku", "")
            qty = c.get("qty", 1) or 1
            if not unit_sku:
                continue
            referenced_skus.add(unit_sku)
            if unit_sku not in unit_sales:
                unit_sales[unit_sku] = {"recent": 0, "prev": 0}
            unit_sales[unit_sku]["recent"] += (p.sales_30_recent or 0) * qty
            unit_sales[unit_sku]["prev"]   += (p.sales_30_prev   or 0) * qty

    # 親発注品の販売数 = 親自身の直販 + コンポーネントSKUの直販合計
    for p_sku, p in parent_orders.items():
        try:
            comps = json.loads(p.set_components or "[]")
        except Exception:
            comps = []
        comp_recent = sum(
            (sku_to_product[c["sku"]].sales_30_recent or 0) * (c.get("qty", 1) or 1)
            for c in comps if c.get("sku") and c["sku"] in sku_to_product
        )
        comp_prev = sum(
            (sku_to_product[c["sku"]].sales_30_prev or 0) * (c.get("qty", 1) or 1)
            for c in comps if c.get("sku") and c["sku"] in sku_to_product
        )
        unit_sales[p_sku] = {
            "recent": (p.sales_30_recent or 0) + comp_recent,
            "prev":   (p.sales_30_prev   or 0) + comp_prev,
        }

    # buy_url あり + 親発注品のコンポーネントでない + (内部管理SKU or バリエーションから参照される単品 or 通常単品)
    singles = [
        p for p in all_products
        if (p.buy_url or "").strip()
        and p.sku not in parent_comp_skus
        and p.sku not in parent_orders
        and not (not p.is_component and p.set_components)  # セット販売商品（parent_ordersに入らなかったもの）は除外
    ]

    # 通常単品 + 親発注品を合わせて計算
    all_order_items = singles + list(parent_orders.values())

    items = []
    for p in all_order_items:
        ordered_1 = ordered1_by_sku.get(p.sku, 0)
        ordered_2 = ordered2_by_sku.get(p.sku, 0)
        ordered = ordered_1 + ordered_2
        agg = unit_sales.get(p.sku, {})
        sales_recent = agg.get("recent", 0)
        sales_prev   = agg.get("prev",   0)
        da = sku_daily_avgs.get(p.sku, {})
        calc = calc_rakuten_order(
            stock=p.stock or 0,
            inbound=0,
            ordered=ordered,
            sales_30_recent=sales_recent,
            sales_30_prev=sales_prev,
            super_sale_qty=0,
            sales_90=p.sales_90 or 0,
            stockout_days_90=p.stockout_days_90 or 0,
            daily_avg_7=da.get("avg_7", 0),
            daily_avg_30=da.get("avg_30", 0),
            s=s,
        )
        try:
            sc_parsed = json.loads(p.set_components) if p.set_components else []
        except Exception:
            sc_parsed = []
        if not calc.needs_order:
            continue
        items.append({
            "product_id":      p.id,
            "sku":             p.sku or "",
            "name":            p.name or "",
            "jan_code":        p.jan_code or "",
            "buy_url":         p.buy_url or "",
            "set_size":        p.set_size or 1,
            "set_components":  sc_parsed,
            "stock":           p.stock or 0,
            "inbound":         ordered,
            "inbound_1":       ordered_1,
            "inbound_2":       ordered_2,
            "ordered":         ordered,
            "ordered_1":       ordered_1,
            "ordered_2":       ordered_2,
            "total_stock":     calc.total_stock,
            "daily_avg":       calc.daily_avg,
            "daily_avg_7":     calc.daily_avg_7,
            "daily_avg_30":    calc.daily_avg_30,
            "days_left":       calc.days_left,
            "growth_rate":     calc.growth_rate,
            "predicted_30":    calc.predicted_30,
            "lead_sales":      calc.lead_sales,
            "safety_stock":    calc.safety_stock,
            "order_qty":       calc.order_qty,
            "needs_order":     calc.needs_order,
            "sales_30_recent": sales_recent,
            "sales_30_prev":   sales_prev,
        })

    return {"items": items, "settings": RakutenSettingsSchema.model_validate(settings_row)}


@router.get("/orders/all-products")
def get_all_products_order(db: Session = Depends(get_db)):
    """全商品（is_component=False）を発注推奨リストと同じ形式で返す"""
    settings_row = _get_or_create_settings(db)
    sku_daily_avgs = _daily_avg_from_table(db)
    s = RakutenCalcSettings(
        lead_days=settings_row.lead_days,
        target_days=settings_row.target_days,
        safety_stock_rate=settings_row.safety_stock_rate,
        threshold_days=settings_row.threshold_days,
    )
    ordered1_by_sku, ordered2_by_sku = _ordered_by_sku_stage(db)
    all_products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
    ).all()

    internalSkus = set(p.sku for p in all_products if p.is_component)

    # セット商品の販売実績を構成単品SKUへ按分（recommendationsと同じロジック）
    unit_sales: dict[str, dict] = {}
    for p in all_products:
        if not p.set_components:
            continue
        try:
            comps = json.loads(p.set_components or "[]")
        except Exception:
            comps = []
        for c in comps:
            unit_sku = c.get("sku", "")
            qty = c.get("qty", 1) or 1
            if not unit_sku:
                continue
            if unit_sku not in unit_sales:
                unit_sales[unit_sku] = {"recent": 0, "prev": 0}
            unit_sales[unit_sku]["recent"] += (p.sales_30_recent or 0) * qty
            unit_sales[unit_sku]["prev"]   += (p.sales_30_prev   or 0) * qty

    # is_component=False・buy_urlあり（内部SKUのみ除外、セット組・本体はすべて表示）
    targets = sorted(
        [p for p in all_products if not p.is_component and (p.buy_url or "").strip()],
        key=lambda p: p.sku or ""
    )
    items = []
    for p in targets:
        ordered_1 = ordered1_by_sku.get(p.sku, 0)
        ordered_2 = ordered2_by_sku.get(p.sku, 0)
        ordered = ordered_1 + ordered_2
        agg = unit_sales.get(p.sku, {})
        sales_recent = agg.get("recent", 0) or (p.sales_30_recent or 0)
        sales_prev   = agg.get("prev",   0) or (p.sales_30_prev   or 0)
        da = sku_daily_avgs.get(p.sku, {})
        calc = calc_rakuten_order(
            stock=p.stock or 0,
            inbound=0,
            ordered=ordered,
            sales_30_recent=sales_recent,
            sales_30_prev=sales_prev,
            super_sale_qty=0,
            sales_90=getattr(p, 'sales_90', None) or 0,
            stockout_days_90=getattr(p, 'stockout_days_90', None) or 0,
            daily_avg_7=da.get("avg_7", 0),
            daily_avg_30=da.get("avg_30", 0),
            s=s,
        )
        sc_parsed = []
        if p.set_components:
            try:
                sc_parsed = json.loads(p.set_components) if isinstance(p.set_components, str) else (p.set_components or [])
            except Exception:
                sc_parsed = []
        items.append({
            "product_id":      p.id,
            "sku":             p.sku or "",
            "name":            p.name or "",
            "buy_url":         p.buy_url or "",
            "set_components":  sc_parsed,
            "stock":           p.stock or 0,
            "inbound":         ordered,
            "inbound_1":       ordered_1,
            "inbound_2":       ordered_2,
            "ordered":         ordered,
            "ordered_1":       ordered_1,
            "ordered_2":       ordered_2,
            "total_stock":     calc.total_stock,
            "daily_avg":       calc.daily_avg,
            "daily_avg_7":     calc.daily_avg_7,
            "daily_avg_30":    calc.daily_avg_30,
            "days_left":       calc.days_left,
            "growth_rate":     calc.growth_rate,
            "order_qty":       calc.order_qty,
            "needs_order":     calc.needs_order,
            "sales_30_recent": sales_recent,
            "sales_30_prev":   sales_prev,
        })
    return {"items": items, "settings": RakutenSettingsSchema.model_validate(settings_row)}


# ============================================================
# Order History（発注済みリスト）
# ============================================================

class RakutenOrderIn(BaseModel):
    sku:       str
    name:      Optional[str] = None
    qty:       int
    stage:     Optional[int] = None  # 未指定なら自動振り分け（未納品ありなら発注済2）
    ordered_at: Optional[date] = None
    memo:      Optional[str] = None

class RakutenOrderOut(BaseModel):
    id:          int
    sku:         str
    name:        Optional[str] = None
    qty:         int
    stage:       Optional[int] = 1
    ordered_at:  Optional[date] = None
    is_delivered: bool
    memo:        Optional[str] = None

    class Config:
        from_attributes = True

@router.get("/orders/history", response_model=List[RakutenOrderOut])
def list_orders(db: Session = Depends(get_db)):
    return (
        db.query(RakutenOrderHistory)
        .filter(
            RakutenOrderHistory.is_deleted == False,
            RakutenOrderHistory.is_delivered == False,
        )
        .order_by(RakutenOrderHistory.ordered_at.desc())
        .all()
    )

@router.post("/orders/history", response_model=RakutenOrderOut)
def create_order(data: RakutenOrderIn, db: Session = Depends(get_db)):
    if data.stage in (1, 2):
        stage = data.stage
    else:
        # 同一SKUに未納品の発注が残っていれば追加発注 → 発注済2
        has_pending = db.query(RakutenOrderHistory).filter(
            RakutenOrderHistory.sku == data.sku,
            RakutenOrderHistory.is_deleted == False,
            RakutenOrderHistory.is_delivered == False,
        ).first() is not None
        stage = 2 if has_pending else 1
    o = RakutenOrderHistory(
        sku=data.sku,
        name=data.name,
        qty=data.qty,
        stage=stage,
        ordered_at=data.ordered_at or date.today(),
        memo=data.memo,
    )
    db.add(o)
    db.commit()
    db.refresh(o)
    return o

@router.patch("/orders/history/{order_id}/stage")
def update_order_stage(order_id: int, body: dict, db: Session = Depends(get_db)):
    """発注済1⇔発注済2の切り替え"""
    stage = body.get("stage")
    if stage not in (1, 2):
        raise HTTPException(400, "stageは1か2を指定してください")
    o = db.query(RakutenOrderHistory).filter(RakutenOrderHistory.id == order_id).first()
    if not o:
        raise HTTPException(404)
    o.stage = stage
    db.commit()
    return {"ok": True, "stage": stage}

@router.patch("/orders/history/{order_id}/qty")
def update_order_qty(order_id: int, body: dict, db: Session = Depends(get_db)):
    qty = body.get("qty")
    if qty is None or not isinstance(qty, (int, float)) or int(qty) < 0:
        raise HTTPException(400, "qtyは0以上の整数を指定してください")
    o = db.query(RakutenOrderHistory).filter(RakutenOrderHistory.id == order_id).first()
    if not o:
        raise HTTPException(404)
    o.qty = int(qty)
    db.commit()
    return {"ok": True, "qty": o.qty}

@router.patch("/orders/history/{order_id}/deliver")
def mark_delivered(order_id: int, db: Session = Depends(get_db)):
    o = db.query(RakutenOrderHistory).filter(RakutenOrderHistory.id == order_id).first()
    if not o:
        raise HTTPException(404)
    o.is_delivered = True
    db.commit()
    return {"ok": True}

@router.delete("/orders/history/{order_id}")
def delete_order(order_id: int, db: Session = Depends(get_db)):
    o = db.query(RakutenOrderHistory).filter(RakutenOrderHistory.id == order_id).first()
    if not o:
        raise HTTPException(404)
    o.is_deleted = True
    db.commit()
    return {"ok": True}


@router.post("/orders/migrate-legacy-inbound")
def migrate_legacy_inbound(body: Optional[dict] = None, db: Session = Depends(get_db)):
    """商品マスタに残っている旧発注済1/2を発注済みリストへ移行する。
    body: {"dry_run": true} ならプレビューのみ。{"dry_run": false} で移行して旧値を0にする。
    """
    body = body or {}
    dry_run = body.get("dry_run", True)
    ordered_at = date.today()
    if body.get("ordered_at"):
        try:
            ordered_at = date.fromisoformat(str(body.get("ordered_at"))[:10])
        except Exception:
            ordered_at = date.today()

    products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
        RakutenProduct.is_component == False,
    ).order_by(RakutenProduct.sku.asc()).all()

    rows = []
    for p in products:
        if not (p.buy_url or "").strip() or _is_manufacturer_product(p):
            continue
        inbound_1 = p.inbound or 0
        inbound_2 = p.standard_stock or 0
        if inbound_1 <= 0 and inbound_2 <= 0:
            continue
        rows.append({
            "id": p.id,
            "sku": p.sku,
            "name": p.name,
            "inbound_1": inbound_1,
            "inbound_2": inbound_2,
        })
        if dry_run:
            continue
        if inbound_1 > 0:
            db.add(RakutenOrderHistory(
                sku=p.sku,
                name=p.name,
                qty=inbound_1,
                stage=1,
                ordered_at=ordered_at,
                memo=body.get("memo") or "旧発注済1から移行",
            ))
        if inbound_2 > 0:
            db.add(RakutenOrderHistory(
                sku=p.sku,
                name=p.name,
                qty=inbound_2,
                stage=2,
                ordered_at=ordered_at,
                memo=body.get("memo") or "旧発注済2から移行",
            ))
        p.inbound = 0
        p.standard_stock = 0

    if not dry_run:
        db.commit()

    return {
        "dry_run": dry_run,
        "count": len(rows),
        "total_inbound_1": sum(r["inbound_1"] for r in rows),
        "total_inbound_2": sum(r["inbound_2"] for r in rows),
        "rows": rows,
    }


# ============================================================
# Excel発注書ダウンロード（タオタロウ形式）
# ============================================================

@router.post("/orders/excel")
def download_order_excel(body: dict, db: Session = Depends(get_db)):
    """発注リストをタオタロウ形式Excelで出力"""
    from app.services.excel_export import build_rakuten_taotaro_excel
    order_items = body.get("items", [])  # [{sku, qty}, ...]
    record_history = bool(body.get("record_history"))
    ordered_at = date.today()
    if body.get("ordered_at"):
        try:
            ordered_at = date.fromisoformat(str(body.get("ordered_at"))[:10])
        except Exception:
            ordered_at = date.today()

    excel_items = []
    history_items: dict[str, dict] = {}
    for oi in order_items:
        sku = oi.get("sku")
        try:
            qty = int(float(oi.get("qty", 0)))
        except Exception:
            qty = 0
        if not sku or not qty:
            continue
        p = db.query(RakutenProduct).filter(RakutenProduct.sku == sku).first()
        if not p:
            continue
        if record_history:
            h = history_items.setdefault(sku, {"sku": sku, "name": p.name, "qty": 0})
            h["qty"] += qty
        # 本体行（set_componentsありかつspec空の場合はスキップ）
        if not (p.set_components and not (p.spec or "").strip()):
            # 発注数は販売単位（セット数）。タオタロウへ渡す数量は仕入単位（個数）なので
            # セット入数(set_size)を掛ける（例: y79 2枚セット×40 → 80個）。
            # セット商品(set_components)は構成品行の qty×comp_qty 側で換算されるため、
            # この掛け算は単体商品（set_sizeで管理・セット商品はset_size=1運用）にだけ効く
            excel_items.append({
                "buy_url":       p.buy_url or "",
                "supplier_spec": getattr(p, "supplier_spec", "") or "",
                "spec":          p.spec or "",
                "qty":           qty * (p.set_size or 1),
                "price":         p.price or 0,
                "customer_memo": p.customer_memo or "",
                "notes":         p.notes or "",
            })
        # set_components + purchase_componentsを展開して追加行として出力
        try:
            comps = json.loads(p.set_components or "[]")
        except Exception:
            comps = []
        try:
            pcomps = json.loads(getattr(p, 'purchase_components', None) or "[]")
        except Exception:
            pcomps = []
        for comp in comps + pcomps:
            comp_sku = comp.get("sku")
            comp_qty = comp.get("qty", 1)
            # set_components内に直接情報がある場合はそちらを使う
            comp_url  = comp.get("buy_url", "")
            comp_spec = comp.get("supplier_spec", "")
            comp_price = comp.get("price", None)
            # なければ商品マスタから補完（is_componentのもののみ）
            if not comp_url or not comp_spec or comp_price is None:
                c = db.query(RakutenProduct).filter(RakutenProduct.sku == comp_sku).first() if comp_sku else None
                if c and c.is_component:
                    comp_url   = comp_url   or c.buy_url or ""
                    comp_spec  = comp_spec  or getattr(c, "supplier_spec", "") or ""
                    comp_price = comp_price if comp_price is not None else (c.price or 0)
                elif not comp_url:
                    continue  # URLも商品マスタもなければスキップ
            excel_items.append({
                "buy_url":       comp_url,
                "supplier_spec": comp_spec,
                "spec":          "",
                "qty":           qty * comp_qty,
                "price":         comp_price or 0,
                "customer_memo": comp.get("customer_memo", ""),
                "notes":         comp.get("notes", ""),
            })

    xls = build_rakuten_taotaro_excel(excel_items)
    if record_history and history_items:
        for item in history_items.values():
            has_pending = db.query(RakutenOrderHistory).filter(
                RakutenOrderHistory.sku == item["sku"],
                RakutenOrderHistory.is_deleted == False,
                RakutenOrderHistory.is_delivered == False,
            ).first() is not None
            db.add(RakutenOrderHistory(
                sku=item["sku"],
                name=item["name"],
                qty=item["qty"],
                stage=2 if has_pending else 1,
                ordered_at=ordered_at,
                memo=body.get("memo") or "発注Excelから登録",
            ))
        db.commit()
    return StreamingResponse(
        io.BytesIO(xls),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rakuten_order.xlsx"},
    )


# ============================================================
# CSV インポート / テンプレートDL
# ============================================================

CSV_COLUMNS = [
    "sku", "name", "jan_code", "spec", "buy_url", "supplier_spec", "price",
    "set_size", "rakuten_item_url", "rakuten_sku_id", "supplier", "standard_stock",
    "stock", "inbound", "sales_30_recent", "sales_30_prev",
    "customer_memo", "notes", "memo", "set_components", "purchase_components", "is_component",
]

CSV_COLUMN_LABELS = {
    "sku":              "商品管理番号(URL)※必須",
    "name":             "商品名",
    "jan_code":         "JANコード",
    "spec":             "システム連携用SKU番号",
    "buy_url":          "仕入れURL",
    "supplier_spec":    "仕入れ仕様(中国語)",
    "price":            "単価(元)",
    "set_size":         "セット入数",
    "rakuten_item_url": "在庫管理番号",
    "rakuten_sku_id":   "楽天SKU管理番号",
    "supplier":         "仕入先",
    "standard_stock":   "発注済2",
    "stock":            "実在庫(手持ち)",
    "inbound":          "発注済1",
    "sales_30_recent":  "直近30日販売数",
    "sales_30_prev":    "60日前〜31日前の販売数",
    "customer_memo":    "お客様専用メモ",
    "notes":            "備考",
    "memo":             "内部メモ",
    "set_components":   "セット構成JSON(在庫連動用)",
    "purchase_components": "発注用付属品JSON(在庫連動しない)",
    "is_component":     "単品フラグ(TRUE/FALSE)",
}

@router.get("/products/csv/template")
def download_csv_template():
    """CSVテンプレートをダウンロード"""
    output = io.StringIO()
    writer = csv.writer(output)
    # ヘッダー行（日本語ラベル）
    writer.writerow([CSV_COLUMN_LABELS[c] for c in CSV_COLUMNS])
    # サンプル行
    writer.writerow([
        "ITEM-001", "サンプル商品A", "4900000000001", "レッド",
        "https://item.taobao.com/xxx", "12.5",
        "1", "https://item.rakuten.co.jp/shop/xxx/", "12345678-A", "タオタロウ", "50",
        "100", "0", "45", "40",
        "お客様専用メモ例", "備考例", "内部メモ例", "", "", "FALSE",
    ])
    output.seek(0)
    # BOM付きUTF-8でExcelで文字化けしないように
    content = "﻿" + output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=rakuten_products_template.csv"},
    )

@router.get("/products/csv/export")
def export_products_csv(db: Session = Depends(get_db)):
    """現在の商品マスタをCSVエクスポート"""
    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).order_by(RakutenProduct.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([CSV_COLUMN_LABELS[c] for c in CSV_COLUMNS])
    for p in products:
        writer.writerow([
            p.sku or "", p.name or "", p.jan_code or "", p.spec or "",
            p.buy_url or "", getattr(p, "supplier_spec", "") or "",
            p.price if p.price is not None else "",
            p.set_size or 1,
            getattr(p, 'rakuten_item_url', '') or "",
            getattr(p, 'rakuten_sku_id', '') or "",
            getattr(p, 'supplier', '') or "",
            getattr(p, 'standard_stock', 0) or 0,
            p.stock or 0, p.inbound or 0,
            p.sales_30_recent or 0, p.sales_30_prev or 0,
            getattr(p, 'customer_memo', '') or "",
            getattr(p, 'notes', '') or "",
            p.memo or "",
            p.set_components or "",
            getattr(p, 'purchase_components', '') or "",
            "TRUE" if p.is_component else "FALSE",
        ])
    output.seek(0)
    content = "﻿" + output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=rakuten_products.csv"},
    )

@router.post("/products/csv/import")
def import_products_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """CSVをアップロードして商品を一括登録・更新"""
    try:
        raw = file.file.read()
        # BOM除去 + デコード（UTF-8 / Shift-JIS 両対応）
        for enc in ("utf-8-sig", "shift_jis", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        else:
            raise HTTPException(400, "文字コードの読み取りに失敗しました（UTF-8またはShift-JISで保存してください）")
    except Exception as e:
        raise HTTPException(400, f"ファイル読み込みエラー: {e}")

    reader = csv.DictReader(io.StringIO(text))

    # ヘッダーを内部キー名にマッピング（日本語ラベル or 英語キー どちらでも受け付ける）
    label_to_key = {v: k for k, v in CSV_COLUMN_LABELS.items()}
    label_to_key["規定在庫数"] = "standard_stock"
    # 旧ラベル（輸送中系）のCSVも引き続き取り込めるようにエイリアスを残す
    label_to_key["輸送中"] = "inbound"
    label_to_key["輸送中1"] = "inbound"
    label_to_key["輸送中2"] = "standard_stock"
    label_to_key.update({k: k for k in CSV_COLUMNS})  # 英語キーもOK

    created = updated = skipped = 0
    errors = []

    for i, row in enumerate(reader, start=2):  # 2行目から（1行目はヘッダー）
        # ヘッダーを正規化
        normalized = {}
        for col, val in row.items():
            key = label_to_key.get((col or "").strip())
            if key:
                normalized[key] = (val or "").strip()

        sku = normalized.get("sku", "")
        if not sku:
            errors.append(f"{i}行目: SKUが空のためスキップ")
            skipped += 1
            continue

        def to_int(v, default=0):
            try: return int(float(v)) if v else default
            except: return default

        def to_float(v, default=None):
            try: return float(v) if v else default
            except: return default

        is_comp_raw = normalized.get("is_component", "").upper()
        is_component = is_comp_raw in ("TRUE", "1", "YES", "はい")
        data = {
            "name":             normalized.get("name") or None,
            "jan_code":         normalized.get("jan_code") or None,
            "spec":             normalized.get("spec") or None,
            "buy_url":          normalized.get("buy_url") or None,
            "supplier_spec":    normalized.get("supplier_spec") or None,
            "price":            to_float(normalized.get("price")),
            "set_size":         to_int(normalized.get("set_size"), 1),
            "rakuten_item_url": normalized.get("rakuten_item_url") or None,
            "rakuten_sku_id":   normalized.get("rakuten_sku_id") or None,
            "supplier":         normalized.get("supplier") or None,
            "standard_stock":   to_int(normalized.get("standard_stock"), 0),
            "stock":            to_int(normalized.get("stock"), 0),
            "inbound":          to_int(normalized.get("inbound"), 0),
            "sales_30_recent":  to_int(normalized.get("sales_30_recent"), 0),
            "sales_30_prev":    to_int(normalized.get("sales_30_prev"), 0),
            "customer_memo":    normalized.get("customer_memo") or None,
            "notes":            normalized.get("notes") or None,
            "memo":             normalized.get("memo") or None,
            "set_components":   normalized.get("set_components") or None,
            "purchase_components": normalized.get("purchase_components") or None,
            "is_component":     is_component,
        }

        existing = db.query(RakutenProduct).filter(RakutenProduct.sku == sku).first()
        if existing:
            for k, v in data.items():
                setattr(existing, k, v)
            updated += 1
        else:
            p = RakutenProduct(sku=sku, **data)
            db.add(p)
            created += 1

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


# ============================================================
# 仕入れ管理（インボイスExcel読み込み）
# ============================================================

class RakutenInvoiceItemIn(BaseModel):
    sku: str
    name_jp: str = ""
    qty: int
    unit_price_cny: float
    buy_url: str = ""
    asin_memo: str = ""  # J列（ASIN/商品番号）：商品内訳メモ
    permit_col: int | None = None  # 手動で指定した申告欄番号（1始まり）

class PermitColumnIn(BaseModel):
    col_no: int
    item_name: str = ""
    hs_code: str = ""
    cif_jpy: int = 0
    tariff_rate: float = 0.0
    tariff_rate_str: str = ""
    duty_jpy: int = 0
    bpr_coeff: float = 0.0

class RakutenInvoiceIn(BaseModel):
    invoice_no: str = ""
    invoice_date: str = ""
    exchange_rate: float
    domestic_freight: float = 0
    international_freight: float = 0
    import_tax_jpy: float = 0  # 輸入税合計（円）：関税＋消費税＋地方消費税
    items: List[RakutenInvoiceItemIn]
    permit_columns: List[PermitColumnIn] = []  # 許可書の申告欄情報（空=従来の一律按分）


def _num(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return default


def _text(value) -> str:
    return str(value or "").strip()


def _code(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip()


def _row_value(row, index, default=None):
    return row[index] if len(row) > index else default


def _url_key(url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    offer = re.search(r"/offer/(\d+)\.html", value)
    if offer:
        return f"1688:{offer.group(1)}"
    item = re.search(r"[?&]id=(\d+)", value)
    if item:
        return f"id:{item.group(1)}"
    return value.split("#", 1)[0].rstrip("/")


def _invoice_sheet(wb):
    for ws in wb.worksheets:
        if str(ws.title).strip().lower() in {"发票", "發票", "invoice"}:
            return ws
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
            joined = " ".join(_text(c) for c in row)
            if "Name of Commodity" in joined:
                return ws
    return wb.active


def _invoice_data_start(ws):
    candidate = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        cells = [_text(c) for c in row]
        joined = " ".join(cells)
        if "Name of Commodity" in joined:
            candidate = i + 2
        if (
            any(c.upper() == "PCS" for c in cells)
            and any("Unit Price" in c or "单价" in c or "單價" in c for c in cells)
            and any("TOTAL Price" in c or "总价" in c or "總價" in c for c in cells)
        ):
            return i + 1
    return candidate


def _invoice_fee_amount(row):
    for value in reversed(row):
        amount = _num(value, None)
        if amount is not None:
            return amount
    return 0.0


def _parse_rakuten_invoice_workbook(wb):
    ws = _invoice_sheet(wb)

    invoice_no = ""
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for cell in row:
            value = _text(cell)
            if value.startswith("VIP") or value.startswith("GBVIP") or "VIP" in value:
                invoice_no = value
                break
        if invoice_no:
            break

    data_start = _invoice_data_start(ws)
    if not data_start:
        raise HTTPException(400, "インボイスの商品データが見つかりません")

    items = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        first = _text(_row_value(row, 0))
        if first.startswith("TOTAL") or first.startswith("MADE IN"):
            break

        qty = _num(_row_value(row, 6))
        unit_price = _num(_row_value(row, 7))
        total_price = _num(_row_value(row, 8))
        if qty <= 0 or unit_price <= 0:
            continue

        invoice_code = _code(_row_value(row, 9))
        fallback_sku = _code(_row_value(row, 12)) or invoice_code
        # 商品URLはK列(SKU/URL)が基本だが、便によってはL列(1688链接)にしか
        # 記入されないことがあるため、URLらしい方を採用する
        buy_url = _text(_row_value(row, 10))
        if "http" not in buy_url:
            alt_url = _text(_row_value(row, 11))
            if "http" in alt_url:
                buy_url = alt_url
        items.append({
            "sku": fallback_sku,
            "name_jp": _text(_row_value(row, 2)) or _text(_row_value(row, 1)),
            "qty": int(qty),
            "unit_price_cny": unit_price,
            "total_price_cny": total_price or round(qty * unit_price, 2),
            "buy_url": buy_url,
            "asin_memo": invoice_code,
        })

    if not items:
        raise HTTPException(400, "インボイスの商品データが見つかりません")

    added_value = domestic_freight = international_freight = 0.0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        joined = " ".join(_text(c) for c in row)
        if "Added Value" in joined or "增值" in joined or "増値" in joined:
            added_value += _invoice_fee_amount(row)
        elif "Domestic Freight" in joined or "国内运费" in joined or "国内送料" in joined:
            domestic_freight += _invoice_fee_amount(row)
        elif "International" in joined and ("Freight" in joined or "运费" in joined or "送料" in joined):
            international_freight += _invoice_fee_amount(row)

    return {
        "invoice_no": invoice_no,
        "domestic_freight": round(domestic_freight + added_value, 2),
        "international_freight": round(international_freight, 2),
        "items": items,
    }


def _permit_text(content: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        raise HTTPException(400, f"輸入許可書読み込みエラー: {str(e)}")


def _permit_values(text: str):
    cny = 0.0
    m = re.search(r"仕入書価格\s+[A-Z]\s+-\s+CIF\s+-\s+CNY\s+-\s+([\d,\.]+)", text)
    if not m:
        m = re.search(r"CIF\s*-\s*CNY\s*-\s*([\d,\.]+)", text)
    if m:
        cny = float(m.group(1).replace(",", ""))

    exchange_rate = 0.0
    m = re.search(r"通貨レート\s+CNY\s*-\s*([0-9]+(?:\.[0-9]+)?)", text)
    if not m:
        m = re.search(r"CNY\s*-\s*([0-9]+(?:\.[0-9]+)?)", text)
    if m:
        exchange_rate = float(m.group(1))

    tax_total = 0
    m = re.search(r"納税額合計\s*[\\¥￥]?\s*([0-9,]+)", text)
    if m:
        tax_total = int(m.group(1).replace(",", ""))

    tax_values = sorted(
        [int(v.replace(",", "")) for v in re.findall(r"[\\¥￥]\s*([0-9]{1,3}(?:,[0-9]{3})+)", text)],
        reverse=True,
    )
    if not tax_total and tax_values:
        tax_total = max(tax_values)

    return {
        "permit_cny": cny,
        "exchange_rate": exchange_rate,
        "import_tax_jpy": tax_total,
        "tax_breakdown": tax_values[:10],
    }


def _parse_permit_columns(text: str) -> list[dict]:
    """輸入許可書から申告欄ごとの関税率・BPR按分係数・品名・税表番号を抽出する。"""
    columns = []
    for m in re.finditer(r"＜\s*(\d+)\s*欄＞", text):
        col_no = int(m.group(1))
        after = text[m.end():m.end() + 800]

        item_name = ""
        n = re.search(r"品名\s+(.+?)(?:\s+数量|$)", after)
        if n:
            item_name = n.group(1).strip()

        hs_code = ""
        n = re.search(r"税表番号\s+([0-9]+(?:\.[0-9]+)?)", after)
        if n:
            hs_code = n.group(1)

        cif_jpy = 0
        n = re.search(r"申告価格（ＣＩＦ）\s*[\\¥￥]?\s*([0-9,]+)", after)
        if n:
            cif_jpy = int(n.group(1).replace(",", ""))

        tariff_rate = 0.0
        tariff_rate_str = ""
        n = re.search(r"関税率\s+[A-Z]?\s*(\S+)", after)
        if n:
            rate_text = n.group(1).strip()
            tariff_rate_str = rate_text
            if rate_text.upper() == "FREE":
                tariff_rate = 0.0
            else:
                pct = re.search(r"([0-9]+(?:\.[0-9]+)?)%", rate_text)
                if pct:
                    tariff_rate = float(pct.group(1))
                else:
                    try:
                        tariff_rate = float(rate_text.replace("%", ""))
                    except ValueError:
                        pass

        duty_jpy = 0
        n = re.search(r"関税額\s*[\\¥￥]?\s*([0-9,]+)", after)
        if n:
            duty_jpy = int(n.group(1).replace(",", ""))

        bpr_coeff = 0.0
        n = re.search(r"ＢＰＲ按分係数\s+([0-9,]+(?:\.[0-9]+)?)", after)
        if n:
            bpr_coeff = float(n.group(1).replace(",", ""))

        columns.append({
            "col_no": col_no,
            "item_name": item_name,
            "hs_code": hs_code,
            "cif_jpy": cif_jpy,
            "tariff_rate": tariff_rate,
            "tariff_rate_str": tariff_rate_str,
            "duty_jpy": duty_jpy,
            "bpr_coeff": bpr_coeff,
        })
    return columns


def _match_items_to_columns(
    items: list[dict], columns: list[dict]
) -> list[int | None]:
    """インボイス商品を許可書の申告欄にマッチングする。
    BPR按分係数 = 商品金額合計（元）なので、金額の組合せで欄を特定する。
    戻り値: 各商品に対応するcolumnsのインデックス（マッチしない場合None）。
    """
    if not columns:
        return [None] * len(items)
    if len(columns) == 1:
        return [0] * len(items)

    item_amounts = [round(it.get("total_price_cny", 0) or (it.get("qty", 0) * it.get("unit_price_cny", 0)), 2) for it in items]

    # 各欄のBPR按分係数（=その欄に属する商品のCNY合計）
    col_targets = [c["bpr_coeff"] for c in columns]

    # 2欄の場合: 各商品の金額を足し合わせて、どの欄のBPR按分係数に近いかで分ける
    # N欄の場合もグリーディに割り当て
    n = len(items)
    assignments: list[int | None] = [None] * n

    def _find_subset_for_target(indices: list[int], target: float) -> list[int] | None:
        """indicesの中からitem_amountsの合計がtargetに一致する部分集合を探す。"""
        k = len(indices)
        if k <= 25:
            for mask in range(1 << k):
                total = sum(item_amounts[indices[j]] for j in range(k) if mask & (1 << j))
                if abs(total - target) <= 1.0:
                    return [indices[j] for j in range(k) if mask & (1 << j)]
            return None
        # 商品数が多い場合: meet-in-the-middle (2^13 * 2 ≈ 16K)
        half = k // 2
        left_indices = indices[:half]
        right_indices = indices[half:]
        left_sums: dict[float, int] = {}
        for mask in range(1 << len(left_indices)):
            total = sum(item_amounts[left_indices[j]] for j in range(len(left_indices)) if mask & (1 << j))
            rounded = round(total, 2)
            left_sums[rounded] = mask
        for rmask in range(1 << len(right_indices)):
            rtotal = sum(item_amounts[right_indices[j]] for j in range(len(right_indices)) if rmask & (1 << j))
            need = round(target - rtotal, 2)
            for delta in [0, 0.01, -0.01, 0.02, -0.02]:
                lmask = left_sums.get(round(need + delta, 2))
                if lmask is not None:
                    result = [left_indices[j] for j in range(len(left_indices)) if lmask & (1 << j)]
                    result += [right_indices[j] for j in range(len(right_indices)) if rmask & (1 << j)]
                    return result
        return None

    remaining = list(range(n))
    # 欄を小さいBPR順に処理（小さい方がマッチしやすい）、最後の欄は残り全部
    sorted_cols = sorted(range(len(columns)), key=lambda ci: col_targets[ci])
    for idx, ci in enumerate(sorted_cols):
        if idx == len(sorted_cols) - 1:
            for i in remaining:
                assignments[i] = ci
            break
        matched = _find_subset_for_target(remaining, col_targets[ci])
        if matched is not None:
            for i in matched:
                assignments[i] = ci
                remaining.remove(i)
        # マッチしなかった場合はスキップして最後の欄に回す

    return assignments


def _rakuten_products_by_unique_url(db: Session):
    rows = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    grouped = {}
    for p in rows:
        key = _url_key(p.buy_url or "")
        if key:
            grouped.setdefault(key, []).append(p)
    return {key: products[0] for key, products in grouped.items() if len(products) == 1}


def _find_invoice_product(db: Session, item: RakutenInvoiceItemIn):
    product = db.query(RakutenProduct).filter(
        RakutenProduct.sku == item.sku,
        RakutenProduct.is_active == True,
    ).first()
    if product:
        return product
    key = _url_key(item.buy_url)
    if not key:
        return None
    matches = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
        RakutenProduct.buy_url != None,
    ).all()
    matched = [p for p in matches if _url_key(p.buy_url or "") == key]
    return matched[0] if len(matched) == 1 else None


@router.post("/invoices/validate-pair")
async def rakuten_validate_pair(
    invoice_file: UploadFile = File(...),
    permit_file: UploadFile = File(...),
):
    """インボイスXLSと輸入許可書PDFのCNY合計が一致するか検証する"""
    import openpyxl
    inv_content = await invoice_file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(inv_content))
    except Exception as e:
        raise HTTPException(400, f"インボイス読み込みエラー: {str(e)}")

    parsed = _parse_rakuten_invoice_workbook(wb)
    goods_cny = sum((item["total_price_cny"] or item["qty"] * item["unit_price_cny"]) for item in parsed["items"])
    with_fees = goods_cny + (parsed["domestic_freight"] or 0) + (parsed["international_freight"] or 0)

    permit_content = await permit_file.read()
    permit = _permit_values(_permit_text(permit_content))
    permit_cny = permit["permit_cny"]

    # 輸入許可書のCIFは商品代のみ(旧)／商品代＋諸費用(新)の両パターンがあるため近い方で判定
    goods_cny = round(goods_cny, 2)
    with_fees = round(with_fees, 2)
    if abs(with_fees - permit_cny) <= abs(goods_cny - permit_cny):
        total_cny = with_fees
    else:
        total_cny = goods_cny
    diff = abs(total_cny - permit_cny)
    ok = diff <= 1.0

    return {
        "ok": ok,
        "invoice_cny": total_cny,
        "permit_cny": permit_cny,
        "diff": round(diff, 2),
        "message": "照合OK" if ok else f"金額不一致（インボイス: {total_cny}元、輸入許可書: {permit_cny}元、差額: {round(diff,2)}元）",
    }


@router.post("/invoices/parse-pdf")
async def rakuten_parse_pdf(file: UploadFile = File(...)):
    """輸入許可証PDFから納税額合計・為替レート・申告欄ごとの関税率を抽出"""
    content = await file.read()
    text = _permit_text(content)
    values = _permit_values(text)
    permit_columns = _parse_permit_columns(text)
    return {
        "import_tax_jpy": values["import_tax_jpy"],
        "exchange_rate": values["exchange_rate"],
        "tax_breakdown": values["tax_breakdown"],
        "permit_columns": permit_columns,
    }


@router.post("/invoices/parse-excel")
async def rakuten_parse_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """タオタロウ形式ExcelをパースしてSKU・単価を返す"""
    import openpyxl
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Excel読み込みエラー: {str(e)}")

    parsed = _parse_rakuten_invoice_workbook(wb)
    unique_by_url = _rakuten_products_by_unique_url(db)

    # 同じ便の配送依頼（追跡番号＝インボイス番号）で照合済みの明細から商品を引き継ぐ。
    # 配送依頼側は色・仕様も使って照合しているため、色違いで同じURLの商品
    # （インボイスのURLだけでは区別できない行）もここで解決できる。
    # キーは (URL, 数量)。同URL同数量の色違いは単価も原価も同じため、残りを順に割り当てる。
    from app.models.shipment_order import ShipmentOrder, ShipmentOrderItem
    ship_pool: dict = {}
    if parsed["invoice_no"]:
        ship_rows = (
            db.query(ShipmentOrderItem)
            .join(ShipmentOrder, ShipmentOrderItem.shipment_order_id == ShipmentOrder.id)
            .filter(
                ShipmentOrder.tracking_no == parsed["invoice_no"],
                ShipmentOrderItem.product_id.isnot(None),
            )
            .all()
        )
        for si in ship_rows:
            key = (_url_key(si.buy_url or ""), int(si.qty or 0))
            ship_pool.setdefault(key, [])
            if si.product_id not in ship_pool[key]:
                ship_pool[key].append(si.product_id)

    matched = unmatched = 0
    for item in parsed["items"]:
        if db.query(RakutenProduct).filter(
            RakutenProduct.sku == item["sku"],
            RakutenProduct.is_active == True,
        ).first():
            matched += 1
            continue
        product = unique_by_url.get(_url_key(item.get("buy_url", "")))
        if not product and ship_pool:
            key = (_url_key(item.get("buy_url", "")), int(item["qty"]))
            pids = ship_pool.get(key) or []
            if pids:
                pid = pids.pop(0)  # 同キーが複数行あるときは1件ずつ消費
                product = db.query(RakutenProduct).filter(
                    RakutenProduct.id == pid, RakutenProduct.is_active == True
                ).first()
        if product:
            original_code = item.get("asin_memo") or item["sku"]
            item["sku"] = product.sku
            item["name_jp"] = product.name or item["name_jp"]
            item["asin_memo"] = f"{original_code} -> {product.sku}"
            matched += 1
        else:
            # 照合できなかった行はSKUを空にして画面で手動選択させる
            # （订单号のままだと保存時に静かにスキップされて紛らわしい）
            item["sku"] = ""
            unmatched += 1

    parsed["matched"] = matched
    parsed["unmatched"] = unmatched
    return parsed

def _calc_tariff_tax(
    items_with_totals: list[tuple[int, float]],
    data: "RakutenInvoiceIn",
    columns: list["PermitColumnIn"],
) -> dict[int, dict]:
    """税率別計算: 各商品インデックスに対する関税・消費税・地方消費税を返す。
    items_with_totals: [(item_index, item_total_cny), ...]
    戻り値: {item_index: {tariff_rate, duty_jpy, consumption_tax_jpy, local_tax_jpy, total_tax_jpy, col_no}}
    """
    if not columns:
        return {}

    items_dicts = [
        {"total_price_cny": total, "permit_col": data.items[idx].permit_col}
        for idx, total in items_with_totals
    ]
    col_dicts = [c.model_dump() for c in columns]

    # 手動指定がある商品はそれを使い、残りを自動マッチング
    assignments = [None] * len(items_dicts)
    for i, it in enumerate(items_dicts):
        if it.get("permit_col") is not None:
            col_idx = next((ci for ci, c in enumerate(col_dicts) if c["col_no"] == it["permit_col"]), None)
            if col_idx is not None:
                assignments[i] = col_idx

    unassigned = [i for i, a in enumerate(assignments) if a is None]
    if unassigned:
        auto_items = [items_dicts[i] for i in unassigned]
        # 手動割り当て分を差し引いたBPR按分係数で再計算
        adjusted_cols = []
        for ci, c in enumerate(col_dicts):
            manual_total = sum(items_dicts[i]["total_price_cny"] for i, a in enumerate(assignments) if a == ci)
            adjusted_cols.append({**c, "bpr_coeff": c["bpr_coeff"] - manual_total})
        auto_assignments = _match_items_to_columns(auto_items, adjusted_cols)
        for j, ui in enumerate(unassigned):
            assignments[ui] = auto_assignments[j]

    result = {}
    for i, (item_idx, item_total) in enumerate(items_with_totals):
        col_idx = assignments[i]
        if col_idx is None:
            col_idx = 0
        col = col_dicts[col_idx]
        tariff_rate = col["tariff_rate"]
        item_cif_jpy = round(item_total * data.exchange_rate)
        # 送料按分を加えたCIF相当額に関税率を適用
        total_cny = sum(t for _, t in items_with_totals)
        total_freight = data.domestic_freight + data.international_freight
        freight_alloc_cny = (item_total / total_cny * total_freight) if total_cny > 0 else 0
        cif_with_freight_jpy = (item_total + freight_alloc_cny) * data.exchange_rate

        duty_jpy = round(cif_with_freight_jpy * tariff_rate / 100)
        taxable = cif_with_freight_jpy + duty_jpy
        consumption_tax_jpy = round(taxable * 7.8 / 100)
        local_tax_jpy = round(consumption_tax_jpy * 22 / 78)

        result[item_idx] = {
            "tariff_rate": tariff_rate,
            "tariff_rate_str": col.get("tariff_rate_str", f"{tariff_rate}%"),
            "duty_jpy": duty_jpy,
            "consumption_tax_jpy": consumption_tax_jpy,
            "local_tax_jpy": local_tax_jpy,
            "total_tax_jpy": duty_jpy + consumption_tax_jpy + local_tax_jpy,
            "col_no": col["col_no"],
            "hs_code": col.get("hs_code", ""),
        }
    return result


@router.post("/invoices/calculate")
def rakuten_calculate_cost(data: RakutenInvoiceIn, db: Session = Depends(get_db)):
    total_cny = sum(i.qty * i.unit_price_cny for i in data.items)
    total_freight = data.domestic_freight + data.international_freight
    import_tax_jpy = data.import_tax_jpy or 0
    use_tariff = bool(data.permit_columns)

    # 税率別計算用: 有効な商品のインデックスと金額を収集
    valid_items: list[tuple[int, float]] = []
    for idx, item in enumerate(data.items):
        if not (item.sku or "").strip():
            continue
        valid_items.append((idx, item.qty * item.unit_price_cny))

    tariff_info = {}
    if use_tariff and valid_items:
        tariff_info = _calc_tariff_tax(valid_items, data, data.permit_columns)

    result = []
    skipped = 0
    for idx, item in enumerate(data.items):
        if not (item.sku or "").strip():
            skipped += 1
            continue
        item_total = item.qty * item.unit_price_cny
        freight_alloc = (item_total / total_cny * total_freight) if total_cny > 0 else 0
        product = _find_invoice_product(db, item)
        if not product:
            skipped += 1
            continue
        set_size = product.set_size or 1
        sell_units = item.qty / set_size if item.qty > 0 else 0

        if use_tariff and idx in tariff_info:
            ti = tariff_info[idx]
            tax_alloc_jpy = ti["total_tax_jpy"]
        else:
            ti = None
            tax_alloc_jpy = (item_total / total_cny * import_tax_jpy) if total_cny > 0 else 0

        cost_jpy = (((item_total + freight_alloc) * data.exchange_rate + tax_alloc_jpy) / sell_units) if sell_units > 0 else 0
        customer_memo = product.customer_memo
        result_item = item.model_dump()
        result_item["sku"] = product.sku
        if product.name:
            result_item["name_jp"] = product.name
        row = {
            **result_item,
            "total_price_cny": round(item_total, 2),
            "freight_alloc_cny": round(freight_alloc, 2),
            "tax_alloc_jpy": round(tax_alloc_jpy, 0),
            "cost_jpy": round(cost_jpy, 1),
            "customer_memo": customer_memo,
            "matched_sku": product.sku,
        }
        if ti:
            row["tariff_rate"] = ti["tariff_rate"]
            row["tariff_rate_str"] = ti["tariff_rate_str"]
            row["duty_jpy"] = ti["duty_jpy"]
            row["col_no"] = ti["col_no"]
            row["hs_code"] = ti["hs_code"]
        result.append(row)

    return {
        "items": result,
        "total_cny": round(total_cny, 2),
        "total_freight_cny": round(total_freight, 2),
        "import_tax_jpy": import_tax_jpy,
        "grand_total_jpy": round((total_cny + total_freight) * data.exchange_rate + import_tax_jpy, 0),
        "skipped": skipped,
        "use_tariff": use_tariff,
    }

@router.post("/invoices/save")
def rakuten_save_invoice(data: RakutenInvoiceIn, db: Session = Depends(get_db)):
    total_cny = sum(i.qty * i.unit_price_cny for i in data.items)
    total_freight = data.domestic_freight + data.international_freight
    updated = 0
    skipped = 0
    updated_skus: dict[str, float] = {}  # sku -> cost_jpy
    use_tariff = bool(data.permit_columns)

    import_tax_jpy = data.import_tax_jpy or 0

    valid_items: list[tuple[int, float]] = []
    for idx, item in enumerate(data.items):
        if not (item.sku or "").strip():
            continue
        valid_items.append((idx, item.qty * item.unit_price_cny))

    tariff_info = {}
    if use_tariff and valid_items:
        tariff_info = _calc_tariff_tax(valid_items, data, data.permit_columns)

    for idx, item in enumerate(data.items):
        if not (item.sku or "").strip():
            skipped += 1
            continue
        item_total = item.qty * item.unit_price_cny
        freight_alloc = (item_total / total_cny * total_freight) if total_cny > 0 else 0

        if use_tariff and idx in tariff_info:
            tax_alloc_jpy = tariff_info[idx]["total_tax_jpy"]
        else:
            tax_alloc_jpy = (item_total / total_cny * import_tax_jpy) if total_cny > 0 else 0

        product = _find_invoice_product(db, item)
        if product:
            set_size = product.set_size or 1
            sell_units = item.qty / set_size if item.qty > 0 else 0
            cost_jpy = round((((item_total + freight_alloc) * data.exchange_rate + tax_alloc_jpy) / sell_units), 1) if sell_units > 0 else 0
            product.cost_jpy = cost_jpy
            product.price = item.unit_price_cny if item.unit_price_cny else product.price
            updated_skus[product.sku] = cost_jpy
            updated += 1
        else:
            skipped += 1

    # set_componentsを持つセット商品の原価を自動再計算
    set_products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
        RakutenProduct.is_component == False,
        RakutenProduct.set_components != None,
    ).all()
    for sp in set_products:
        try:
            comps = json.loads(sp.set_components or "[]")
        except Exception:
            continue
        total_cost = 0.0
        all_found = True
        for c in comps:
            comp_sku = c.get("sku", "")
            qty = c.get("qty", 1)
            # 更新されたSKUの原価を優先、なければDB値を使用
            if comp_sku in updated_skus:
                comp_cost = updated_skus[comp_sku]
            else:
                comp_product = db.query(RakutenProduct).filter(
                    RakutenProduct.sku == comp_sku, RakutenProduct.is_active == True
                ).first()
                if comp_product and comp_product.cost_jpy:
                    comp_cost = comp_product.cost_jpy
                else:
                    all_found = False
                    break
            total_cost += comp_cost * qty
        if all_found and total_cost > 0:
            sp.cost_jpy = round(total_cost, 1)
            updated += 1

    db.commit()
    return {"updated": updated, "skipped": skipped}


# ============================================================
# RMS API 連携
# ============================================================

@router.get("/rms/debug-orders")
async def debug_rms_orders(db: Session = Depends(get_db)):
    """デバッグ用: searchOrderの生レスポンスを返す（直近3日）"""
    import json, base64, httpx
    from datetime import datetime, timedelta
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")
    token = base64.b64encode(f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()).decode()
    headers = {"Authorization": f"ESA {token}", "Content-Type": "application/json; charset=utf-8"}
    now = datetime.now()
    body = {
        "dateType": 1,
        "startDatetime": (now - timedelta(days=3)).strftime("%Y-%m-%dT00:00:00+0900"),
        "endDatetime": now.strftime("%Y-%m-%dT23:59:59+0900"),
        "PaginationRequestModel": {"requestRecordsAmount": 10, "requestPage": 1},
    }
    async with httpx.AsyncClient(timeout=30) as client:
        res = await client.post(
            "https://api.rms.rakuten.co.jp/es/2.0/order/searchOrder",
            headers=headers,
            content=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        )
    return {"status": res.status_code, "body": res.json()}


@router.get("/rms/debug-order-detail")
async def debug_rms_order_detail(db: Session = Depends(get_db)):
    """デバッグ用: getOrderの生レスポンスを返す（直近3日の先頭1件）"""
    import json, base64, httpx
    from datetime import datetime, timedelta
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")
    token = base64.b64encode(f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()).decode()
    headers = {"Authorization": f"ESA {token}", "Content-Type": "application/json; charset=utf-8"}
    now = datetime.now()
    # まず注文番号を1件取得
    search_body = {
        "dateType": 1,
        "startDatetime": (now - timedelta(days=3)).strftime("%Y-%m-%dT00:00:00+0900"),
        "endDatetime": now.strftime("%Y-%m-%dT23:59:59+0900"),
        "PaginationRequestModel": {"requestRecordsAmount": 1, "requestPage": 1},
    }
    async with httpx.AsyncClient(timeout=30) as client:
        res = await client.post(
            "https://api.rms.rakuten.co.jp/es/2.0/order/searchOrder",
            headers=headers,
            content=json.dumps(search_body, ensure_ascii=False).encode("utf-8"),
        )
        data = res.json()
    order_numbers = data.get("orderNumberList", [])
    if not order_numbers:
        return {"error": "注文なし"}
    # getOrderで詳細取得
    async with httpx.AsyncClient(timeout=30) as client:
        res2 = await client.post(
            "https://api.rms.rakuten.co.jp/es/2.0/order/getOrder",
            headers=headers,
            content=json.dumps({"orderNumberList": order_numbers[:1], "version": 10}, ensure_ascii=False).encode("utf-8"),
        )
    return {"status": res2.status_code, "body": res2.json()}


@router.get("/rms/debug-inventory")
async def debug_rms_inventory(
    manage_number: str = "s08",
    db: Session = Depends(get_db),
):
    """デバッグ用: 指定manageNumber配下の在庫bulk-get生レスポンスを返す（DB書き換えなし・読み取り専用）。

    ツールが実際に送る形（variantId=DBのsku）と、variantIdを送らずmanageNumberだけ送る形の
    両方で楽天に問い合わせ、楽天が返す実variantIdと数量を生のまま表示する。
    これにより「DBのsku(244)が楽天のvariantId(シトラスミント)と食い違うときpullが取れているか」を確定できる。
    """
    import json, base64, httpx
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")
    token = base64.b64encode(f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()).decode()
    headers = {"Authorization": f"ESA {token}", "Content-Type": "application/json"}
    url = "https://api.rms.rakuten.co.jp/es/2.0/inventories/bulk-get"

    # この manageNumber に紐づくDB商品（送っているvariantId=skuを把握するため）
    db_products = db.query(RakutenProduct).filter(
        RakutenProduct.rakuten_item_url == manage_number,
        RakutenProduct.is_active == True,
    ).all()
    db_rows = [
        {"id": p.id, "sku": p.sku, "rakuten_sku_id": p.rakuten_sku_id, "db_stock": p.stock}
        for p in db_products
    ]
    sent_variant_ids = [p.sku for p in db_products if p.sku]

    async def _call(inventories: list[dict]):
        body = json.dumps({"inventories": inventories}, ensure_ascii=False).encode("utf-8")
        async with httpx.AsyncClient(timeout=30) as client:
            res = await client.post(url, headers=headers, content=body)
        try:
            parsed = res.json()
        except Exception:
            parsed = res.text[:500]
        return {"status": res.status_code, "body": parsed}

    # ① ツールが実際に送る形: variantId=DBのsku（244など）
    as_sent = await _call([{"manageNumber": manage_number, "variantId": vid} for vid in sent_variant_ids]) if sent_variant_ids else {"skipped": "DBにこのmanageNumberのskuなし"}
    # ② variantIdを送らずmanageNumberだけ（楽天が配下全variantを返すか確認）
    mn_only = await _call([{"manageNumber": manage_number}])

    return {
        "manage_number": manage_number,
        "db_products": db_rows,
        "sent_variant_ids": sent_variant_ids,
        "result_as_tool_sends": as_sent,
        "result_manage_number_only": mn_only,
    }


@router.get("/rms/debug-pull-missing")
async def debug_rms_pull_missing(db: Session = Depends(get_db)):
    """デバッグ用: pull対象の全SKUを楽天へ問い合わせ、楽天が返さないSKU（取りこぼし）を一覧する。
    DB書き換えなし・読み取り専用。_pull_rms_stock と同じitems組み立てで、sent/returned/missingの差分を返す。
    """
    import re as _re
    from app.services.rakuten_rms import fetch_inventory_from_rms
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")

    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    sku_to_product = {p.sku: p for p in products}

    # _pull_rms_stock と同一ロジックで送信items を組み立てる
    items = []
    for p in products:
        sku = (p.sku or "").strip()
        if not sku:
            continue
        if p.is_component and not p.rakuten_item_url:
            continue
        if not _re.match(r'^[a-zA-Z0-9_\-]+$', sku):
            continue
        manage_number = (p.rakuten_item_url or sku.split("_")[0]).strip()
        if not manage_number:
            continue
        items.append({"manage_number": manage_number, "variant_id": sku})

    rms_stock = await fetch_inventory_from_rms(
        settings.rms_service_secret, settings.rms_license_key, items
    )

    sent_skus = {it["variant_id"] for it in items}
    returned_skus = set(rms_stock.keys())
    missing_skus = sent_skus - returned_skus
    # 楽天は返したがDBにそのskuが無い（マッチ先がない）ケースも拾う
    unmatched_returned = returned_skus - {p.sku for p in products}

    missing_detail = []
    for sku in sorted(missing_skus):
        p = sku_to_product.get(sku)
        missing_detail.append({
            "sku": sku,
            "manage_number": (p.rakuten_item_url or sku.split("_")[0]) if p else None,
            "rakuten_sku_id": p.rakuten_sku_id if p else None,
            "db_stock": p.stock if p else None,
            "name": p.name if p else None,
        })

    return {
        "sent": len(sent_skus),
        "returned": len(returned_skus),
        "missing_count": len(missing_skus),
        "missing": missing_detail,
        "unmatched_returned": sorted(unmatched_returned),
    }


@router.post("/rms/test")
async def test_rms_connection(db: Session = Depends(get_db)):
    """RMS API 接続テスト"""
    from app.services.rakuten_rms import test_connection
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")
    result = await test_connection(settings.rms_service_secret, settings.rms_license_key)
    if not result["ok"]:
        raise HTTPException(502, f"接続失敗 (HTTP {result.get('status')}): {result.get('detail', '')}")
    return {"ok": True}


_price_sync_status = {"running": False, "result": None}

@router.get("/rms/sync-prices/status")
def get_price_sync_status():
    return _price_sync_status

@router.post("/rms/sync-prices")
async def sync_prices_from_rms(db: Session = Depends(get_db)):
    """RMS Items Search APIから商品の売価をバックグラウンドで取得"""
    import base64, httpx, asyncio
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。")

    if _price_sync_status["running"]:
        return {"ok": True, "message": "同期中です。しばらくお待ちください。", **_price_sync_status}

    token = base64.b64encode(
        f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()
    ).decode()
    headers = {"Authorization": f"ESA {token}"}
    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    product_data = [(p.id, p.sku or "") for p in products]

    from app.core.database import SessionLocal

    async def do_sync():
        _price_sync_status["running"] = True
        _price_sync_status["result"] = None
        sku_price_map = {}
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                offset = 0
                retry = 0
                while True:
                    res = await client.get(
                        "https://api.rms.rakuten.co.jp/es/2.0/items/search",
                        headers=headers,
                        params={"offset": offset},
                    )
                    if res.status_code == 429:
                        if retry >= 5:
                            break
                        await asyncio.sleep(2 ** retry)
                        retry += 1
                        continue
                    if res.status_code != 200:
                        break
                    retry = 0
                    data = res.json()
                    results = data.get("results", [])
                    if not results:
                        break
                    for entry in results:
                        item = entry.get("item", {})
                        for variant_key, variant_data in item.get("variants", {}).items():
                            price = variant_data.get("standardPrice")
                            if price is not None:
                                sku_price_map[variant_key] = float(price)
                            merchant_sku_id = variant_data.get("merchantDefinedSkuId")
                            if merchant_sku_id:
                                sku_price_map[f"__spec__{variant_key}"] = merchant_sku_id
                    total = data.get("numFound", 0)
                    offset += len(results)
                    if offset >= total:
                        break
                    await asyncio.sleep(0.3)

            updated = 0
            session = SessionLocal()
            try:
                for pid, key in product_data:
                    if not key:
                        continue
                    p = session.query(RakutenProduct).filter(RakutenProduct.id == pid).first()
                    if not p:
                        continue
                    if key in sku_price_map:
                        p.selling_price = sku_price_map[key]
                        updated += 1
                    spec_val = sku_price_map.get(f"__spec__{key}")
                    if spec_val and not p.spec:
                        p.spec = spec_val
                session.commit()
            finally:
                session.close()

            _price_sync_status["result"] = {"ok": True, "fetched_variants": len(sku_price_map), "updated_products": updated}
        except Exception as e:
            _price_sync_status["result"] = {"ok": False, "error": str(e)}
        finally:
            _price_sync_status["running"] = False

    asyncio.create_task(do_sync())
    return {"ok": True, "message": "バックグラウンドで売価取得を開始しました。/status で進捗確認できます。"}


@router.post("/rms/sync-sku-mapping")
async def sync_sku_mapping_from_rms(db: Session = Depends(get_db)):
    """RMS Items APIから商品管理番号(rakuten_item_url)とSKU番号(rakuten_sku_id)を一括取得してDBに保存"""
    import base64, httpx, asyncio
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。")

    token = base64.b64encode(
        f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()
    ).decode()
    headers = {"Authorization": f"ESA {token}"}

    # RMSから全商品のmanageNumber・variantId・merchantDefinedSkuIdを取得
    # variant_key(SKU管理番号) -> {manage_number, merchant_sku_id}
    sku_map: dict[str, dict] = {}

    async with httpx.AsyncClient(timeout=60) as client:
        offset = 0
        retry = 0
        while True:
            res = await client.get(
                "https://api.rms.rakuten.co.jp/es/2.0/items/search",
                headers=headers,
                params={"offset": offset},
            )
            if res.status_code == 429:
                if retry >= 5:
                    break
                await asyncio.sleep(2 ** retry)
                retry += 1
                continue
            if res.status_code != 200:
                raise HTTPException(502, f"RMS API エラー: {res.status_code}")
            retry = 0
            data = res.json()
            results = data.get("results", [])
            if not results:
                break
            for entry in results:
                item = entry.get("item", {})
                manage_number = item.get("manageNumber", "")
                for variant_key, variant_data in item.get("variants", {}).items():
                    merchant_sku_id = variant_data.get("merchantDefinedSkuId") or ""
                    sku_map[variant_key] = {
                        "manage_number": manage_number,
                        "merchant_sku_id": merchant_sku_id,
                    }
            total = data.get("numFound", 0)
            offset += len(results)
            if offset >= total:
                break
            await asyncio.sleep(0.3)

    # DBの商品と照合してrakuten_item_url・rakuten_sku_idを更新
    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    updated = 0
    for p in products:
        if not p.sku:
            continue
        info = sku_map.get(p.sku)
        if info:
            p.rakuten_item_url = info["manage_number"]
            if info["merchant_sku_id"]:
                p.rakuten_sku_id = info["merchant_sku_id"]
            updated += 1
    db.commit()

    return {"ok": True, "fetched_variants": len(sku_map), "updated": updated}


@router.get("/rms/item-sample")
async def rms_item_sample(db: Session = Depends(get_db)):
    """Item API 2.0のレスポンス構造確認用（先頭1件取得）"""
    import base64, httpx
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。")
    token = base64.b64encode(f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()).decode()
    headers = {"Authorization": f"ESA {token}"}
    async with httpx.AsyncClient(timeout=30) as client:
        res = await client.get(
            "https://api.rms.rakuten.co.jp/es/2.0/items/search",
            headers=headers,
            params={"offset": 0, "limit": 1},
        )
    return {"status": res.status_code, "body": res.json() if res.status_code == 200 else res.text[:1000]}


@router.post("/rms/import-stock")
async def import_stock_from_rms(db: Session = Depends(get_db)):
    """RMSから在庫数を一括取得してDBのstockに保存"""
    from app.services.rakuten_rms import fetch_inventory_from_rms
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。")

    # 在庫計算には内部SKU（is_component=True: 構成品）も含めた全商品が必要。
    # 内部SKUはRMSに在庫が無い（発注→入荷処理でDBにのみ在庫が入る）ため、
    # RMS問い合わせ対象(items)からは除外しつつ、セット計算用のsku_to_productには含める。
    products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
        RakutenProduct.sku != None,
    ).all()

    import re
    items = []
    sku_to_product = {}
    for p in products:
        sku = (p.sku or "").strip()
        if not sku:
            continue
        sku_to_product[sku] = p
        if p.is_component and not p.rakuten_item_url:
            continue
        # RMSのvariantIdに使えない文字（スペース等）を含むSKUは問い合わせ対象から除外
        if not re.match(r'^[a-zA-Z0-9_\-]+$', sku):
            continue
        # manageNumber: rakuten_item_url優先、なければskuの"_"前
        if p.rakuten_item_url:
            manage_number = p.rakuten_item_url.strip()
        elif "_" in sku:
            manage_number = sku.split("_")[0]
        else:
            manage_number = sku
        items.append({"manage_number": manage_number, "variant_id": sku})

    if not items:
        raise HTTPException(400, "対象商品がありません。")

    try:
        rms_stock = await fetch_inventory_from_rms(
            settings.rms_service_secret,
            settings.rms_license_key,
            items,
        )
    except Exception as e:
        raise HTTPException(502, f"楽天APIエラー: {str(e)}")

    # Step1: RMSから取得できた在庫を上書き（内部SKUはRMS対象外なのでDB在庫を保持）
    updated = 0
    for sku, qty in rms_stock.items():
        p = sku_to_product.get(sku)
        if p:
            p.stock = qty
            updated += 1
    # RMS問い合わせ対象だったのに在庫が返らなかった販売SKU数
    not_found = sum(
        1 for sku, p in sku_to_product.items()
        if not p.is_component and sku not in rms_stock
    )

    # セット販売SKU（is_component=false）の在庫はRMSが基盤（Step1で上書き済み）。
    # 内部SKU（is_component=true）の在庫はRMSに無く、入荷処理で入ったDB在庫をそのまま維持する。
    # → 構成品⇔セット間の在庫計算（旧Step2/Step3）は、RMSの正値を潰すため廃止。
    db.commit()
    not_found_skus = [
        sku for sku, p in sku_to_product.items()
        if not p.is_component and sku not in rms_stock
    ]
    return {
        "ok": True,
        "updated": updated,
        "not_found": not_found,
        "total_from_rms": len(rms_stock),
        "not_found_skus": not_found_skus[:50],
    }


# ===== 売上同期（バックグラウンドジョブ） =====
# 受注60日分の取得は数分かかり1リクエストでは502になるため、
# バックグラウンドで実行し、フロントはjob_idでポーリングする。
_sync_jobs: dict = {}
_sync_jobs_lock = threading.Lock()


def _prune_sync_jobs():
    """古い同期ジョブをメモリから掃除（_sync_jobsの肥大化＝メモリリーク防止）"""
    now = time.time()
    with _sync_jobs_lock:
        for jid in [j for j, v in _sync_jobs.items() if now - v.get("started_at", now) > 3600]:
            _sync_jobs.pop(jid, None)
        if len(_sync_jobs) > 20:
            for jid, _ in sorted(_sync_jobs.items(), key=lambda kv: kv[1].get("started_at", 0))[:-20]:
                _sync_jobs.pop(jid, None)


def _run_sales_sync_job(job_id: str, service_secret: str, license_key: str):
    """バックグラウンドで楽天RMSから受注を取得し、各商品の販売数を更新する"""
    with _sync_jobs_lock:
        _sync_jobs[job_id] = {"status": "running", "result": None, "error": None, "started_at": time.time()}

    from app.services.rakuten_rms import fetch_sales_by_sku
    db = SessionLocal()
    try:
        # 60日分取得（元の設計値）。直近30日=recent、31〜60日前=prevで成長率も算出でき、
        # sales_90(=60日合計)が発注計算の日販ベースになる。
        # 並列取得＋100件/回でメモリは同時リクエスト分で頭打ちのため60日でも落ちない。
        sku_sales = asyncio.run(fetch_sales_by_sku(service_secret, license_key, days=60))

        products = db.query(RakutenProduct).filter(
            RakutenProduct.is_active == True,
            RakutenProduct.is_component == False,
        ).all()

        updated = 0
        for p in products:
            # 楽天SKU管理番号 または 商品管理番号(sku)で照合
            sales = sku_sales.get(p.rakuten_sku_id or "") or sku_sales.get(p.sku or "") or {}
            if sales:
                p.sales_30_recent  = sales.get("recent", 0)
                p.sales_30_prev    = sales.get("prev",   0)
                p.sales_90         = sales.get("total_90", 0)
                p.stockout_days_90 = sales.get("stockout_days", 0)
                p.sales_updated_at = _now_jst()
                updated += 1
        db.commit()

        with _sync_jobs_lock:
            _sync_jobs[job_id]["status"] = "done"
            _sync_jobs[job_id]["result"] = {
                "synced_skus": len(sku_sales),
                "updated_products": updated,
                "last_sync": _now_jst().isoformat(),
            }
    except Exception as e:
        db.rollback()
        with _sync_jobs_lock:
            _sync_jobs[job_id]["status"] = "error"
            _sync_jobs[job_id]["error"] = str(e)
    finally:
        db.close()


@router.post("/rms/sync/start")
def start_sales_sync(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """売上同期をバックグラウンドで開始し、job_idを返す（処理は数分かかる）"""
    _prune_sync_jobs()
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。楽天設定から登録してください。")
    job_id = str(uuid.uuid4())
    background_tasks.add_task(
        _run_sales_sync_job, job_id,
        settings.rms_service_secret, settings.rms_license_key,
    )
    return {"job_id": job_id}


@router.get("/rms/sync/status/{job_id}")
def get_sales_sync_status(job_id: str):
    """売上同期ジョブの状態と結果を返す"""
    with _sync_jobs_lock:
        job = _sync_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "ジョブが見つかりません")
    return {
        "status": job["status"],
        "result": job["result"],
        "error": job["error"],
        "elapsed": round(time.time() - job["started_at"], 1),
    }


class SalesApplyRequest(BaseModel):
    # {sku or rakuten_sku_id: {"recent":int,"prev":int,"total_90":int,"stockout_days":int}}
    sales: dict


@router.post("/rms/sales/apply")
def apply_sales(req: SalesApplyRequest, db: Session = Depends(get_db)):
    """GitHub Actions側で集計済みのSKU別販売数を受け取り、商品の販売数を更新する。
    重い受注取得はGitHub側(メモリ7GB)で行い、Renderは軽いDB書き込みだけを担当する
    （60日分の受注取得をRenderで走らせると512MB超過でOOMするため）。"""
    sales = req.sales or {}
    products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
        RakutenProduct.is_component == False,
    ).all()
    updated = 0
    for p in products:
        s = sales.get(p.rakuten_sku_id or "") or sales.get(p.sku or "") or {}
        if s:
            p.sales_30_recent  = s.get("recent", 0)
            p.sales_30_prev    = s.get("prev", 0)
            p.sales_90         = s.get("total_90", 0)
            p.stockout_days_90 = s.get("stockout_days", 0)
            p.sales_updated_at = _now_jst()
            updated += 1
    db.commit()
    return {
        "synced_skus": len(sales),
        "updated_products": updated,
        "last_sync": _now_jst().isoformat(),
    }


# ============ 日別販売数 ============

class DailySalesApplyRequest(BaseModel):
    # {sku: {date_str: qty, ...}, ...}
    daily: dict


@router.post("/rms/daily-sales/apply")
def apply_daily_sales(req: DailySalesApplyRequest, db: Session = Depends(get_db)):
    """GitHub Actions から日別×SKU の販売数を受け取り rakuten_daily_sales に upsert する。"""
    from app.models.rakuten_daily_sales import RakutenDailySales
    from datetime import date as _date
    upserted = 0
    for sku, day_map in (req.daily or {}).items():
        for day_str, qty in day_map.items():
            try:
                sale_date = _date.fromisoformat(day_str)
            except Exception:
                continue
            existing = db.query(RakutenDailySales).filter(
                RakutenDailySales.sale_date == sale_date,
                RakutenDailySales.sku == sku,
            ).first()
            if existing:
                existing.qty = qty
            else:
                db.add(RakutenDailySales(sale_date=sale_date, sku=sku, qty=qty))
            upserted += 1
    db.commit()
    return {"upserted": upserted}


@router.get("/daily-sales")
def get_daily_sales(days: int = 7, db: Session = Depends(get_db)):
    """日別×商品の販売数を返す。商品名はrakuten_productsから結合。"""
    from app.models.rakuten_daily_sales import RakutenDailySales
    from datetime import date as _date, timedelta as _td
    cutoff = _date.today() - _td(days=days)
    rows = db.query(RakutenDailySales).filter(
        RakutenDailySales.sale_date >= cutoff,
    ).order_by(RakutenDailySales.sale_date.desc()).all()

    products = db.query(RakutenProduct).filter(RakutenProduct.is_active == True).all()
    sku_name = {}
    for p in products:
        sku_name[p.sku] = p.name
        if p.rakuten_sku_id:
            sku_name[p.rakuten_sku_id] = p.name

    # {sku: {date: qty}}
    result: dict[str, dict] = {}
    for r in rows:
        d = r.sale_date.isoformat()
        if r.sku not in result:
            result[r.sku] = {}
        result[r.sku][d] = r.qty

    sku_list = []
    for sku, days_data in result.items():
        sku_list.append({
            "sku": sku,
            "name": sku_name.get(sku, sku),
            "daily": days_data,
            "total": sum(days_data.values()),
        })
    sku_list.sort(key=lambda x: x["total"], reverse=True)
    return {"data": sku_list, "days": days}


@router.post("/rms/daily-sales/mark-stockouts")
def mark_stockouts(db: Session = Depends(get_db)):
    """在庫0のSKUについて今日のrakuten_daily_salesにis_stockout=Trueを設定する。"""
    from app.models.rakuten_daily_sales import RakutenDailySales
    today = date.today()
    zero_stock_skus = [
        p.sku for p in db.query(RakutenProduct).filter(
            RakutenProduct.is_active == True,
            RakutenProduct.stock <= 0,
        ).all()
        if p.sku
    ]
    marked = 0
    for sku in zero_stock_skus:
        existing = db.query(RakutenDailySales).filter(
            RakutenDailySales.sale_date == today,
            RakutenDailySales.sku == sku,
        ).first()
        if existing:
            existing.is_stockout = True
        else:
            db.add(RakutenDailySales(sale_date=today, sku=sku, qty=0, is_stockout=True))
        marked += 1
    db.commit()
    return {"marked_stockouts": marked, "date": today.isoformat()}


# ============ スーパーセール(SS)販売数の集計・保存 ============

def _run_ss_sync_job(job_id: str, service_secret: str, license_key: str, period_key: str,
                     ss_start, ss_end):
    """SS期間の販売数を集計してrakuten_ss_salesに保存する（バックグラウンド）"""
    from app.services.rakuten_rms import fetch_ss_sales
    from app.models.rakuten_ss_sales import RakutenSsSales
    with _sync_jobs_lock:
        _sync_jobs[job_id] = {"status": "running", "result": None, "error": None, "started_at": time.time()}
    db = SessionLocal()
    try:
        sku_qty = asyncio.run(fetch_ss_sales(service_secret, license_key, ss_start, ss_end))

        # 表示対象商品（is_component=False）のSKUに紐付けて保存
        products = db.query(RakutenProduct).filter(
            RakutenProduct.is_active == True,
            RakutenProduct.is_component == False,
        ).all()

        saved = 0
        for p in products:
            qty = sku_qty.get(p.rakuten_sku_id or "")
            if qty is None:
                qty = sku_qty.get(p.sku or "")
            if qty is None:
                continue
            row = db.query(RakutenSsSales).filter(
                RakutenSsSales.sku == p.sku,
                RakutenSsSales.ss_period == period_key,
            ).first()
            if row:
                row.qty = qty
            else:
                db.add(RakutenSsSales(sku=p.sku, ss_period=period_key, qty=qty))
            saved += 1
        db.commit()

        with _sync_jobs_lock:
            _sync_jobs[job_id]["status"] = "done"
            _sync_jobs[job_id]["result"] = {
                "period": period_key,
                "matched_skus": len(sku_qty),
                "saved_products": saved,
            }
    except Exception as e:
        db.rollback()
        with _sync_jobs_lock:
            _sync_jobs[job_id]["status"] = "error"
            _sync_jobs[job_id]["error"] = str(e)
    finally:
        db.close()


@router.post("/rms/ss-sync/start")
def start_ss_sync(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """直近に終了したSS期間の販売数を集計・保存するジョブを開始する。
    SS期間は3/6/9/12月の4日20:00〜11日2:00固定。楽天APIは63日前までしか遡れないため、
    SS終了後63日以内に実行する必要がある。"""
    from app.services.rakuten_rms import ss_period_for
    _prune_sync_jobs()
    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "RMS APIキーが設定されていません。楽天設定から登録してください。")

    from datetime import timezone, timedelta
    jst = timezone(timedelta(hours=9))
    now = datetime.now(jst)
    period = ss_period_for(now)
    if not period:
        raise HTTPException(400, "対象となるSS期間が見つかりません。")
    period_key, ss_start, ss_end = period

    # SS終了から63日を超えていたらAPIで遡れない
    if (now - ss_end).days > 62:
        raise HTTPException(
            400,
            f"直近SS({period_key})は終了から63日を超えているため、楽天APIで遡れません。",
        )

    job_id = str(uuid.uuid4())
    background_tasks.add_task(
        _run_ss_sync_job, job_id,
        settings.rms_service_secret, settings.rms_license_key,
        period_key, ss_start, ss_end,
    )
    return {"job_id": job_id, "period": period_key}


@router.get("/ss-sales")
def get_ss_sales(period: Optional[str] = None, db: Session = Depends(get_db)):
    """保存済みのSS販売数を返す。period省略時は最新期間。
    戻り値: {"period": "2026-06", "sales": {sku: qty, ...}}"""
    from app.models.rakuten_ss_sales import RakutenSsSales
    if not period:
        latest = db.query(RakutenSsSales.ss_period).order_by(
            RakutenSsSales.ss_period.desc()
        ).first()
        period = latest[0] if latest else None
    if not period:
        return {"period": None, "sales": {}}
    rows = db.query(RakutenSsSales).filter(RakutenSsSales.ss_period == period).all()
    return {"period": period, "sales": {r.sku: r.qty for r in rows}}


# ===== 段階的push検証（段階2） =====


def _resolve_push_group(component_sku: str, db: Session) -> dict:
    """単品SKUを起点に、関連セットを探索し、単品在庫からセット在庫を再計算する。
    戻り値: {
      "component": {"sku", "stock", "manage_number", "name"},
      "sets": [{"sku", "db_stock", "calculated_stock", "manage_number", "name", "components"}, ...],
    }
    """
    import re as _re
    all_products = db.query(RakutenProduct).filter(
        RakutenProduct.is_active == True,
    ).all()
    sku_to_product = {p.sku: p for p in all_products}

    comp_product = sku_to_product.get(component_sku)
    if not comp_product:
        return None

    comp_stock = comp_product.stock or 0

    def parse_comps(p):
        try:
            return json.loads(p.set_components or "[]")
        except Exception:
            return []

    sets = []
    share_counts = build_component_share_counts(all_products)
    for p in all_products:
        comps = parse_comps(p)
        if not comps:
            continue
        comp_skus_in_set = [c.get("sku") for c in comps]
        if component_sku not in comp_skus_in_set:
            continue
        if p.is_component:
            continue
        s = (p.sku or "").strip()
        if not s or not _re.match(r'^[a-zA-Z0-9_\-]+$', s):
            continue

        merged = {}
        for c in comps:
            c_sku = c.get("sku")
            c_qty = c.get("qty") or 1
            merged[c_sku] = merged.get(c_sku, 0) + c_qty
        set_qty = None
        for c_sku, total_qty in merged.items():
            avail = calc_set_avail(sku_to_product.get(c_sku).stock or 0, total_qty, share_counts.get(c_sku, 0)) if sku_to_product.get(c_sku) else 0
            set_qty = avail if set_qty is None else min(set_qty, avail)

        manage_number = (p.rakuten_item_url or s.split("_")[0]).strip()
        sets.append({
            "sku": s,
            "db_stock": p.stock or 0,
            "calculated_stock": set_qty if set_qty is not None else 0,
            "manage_number": manage_number,
            "name": p.name,
            "components": comps,
        })

    comp_has_rms = bool((comp_product.rakuten_item_url or "").strip())
    comp_mn = (comp_product.rakuten_item_url or component_sku.split("_")[0]).strip()
    return {
        "component": {
            "sku": component_sku,
            "stock": comp_stock,
            "manage_number": comp_mn,
            "name": comp_product.name,
            "is_rms_sku": comp_has_rms,
        },
        "sets": sets,
    }


@router.get("/rms/debug-push-preview")
async def debug_push_preview(
    component_sku: str = "",
    db: Session = Depends(get_db),
):
    """単品SKUを指定 → 関連セットを自動探索 → 単品在庫からセット在庫を再計算して楽天値と比較。
    component_sku: 単品SKU（例: 244）。関連セットは自動で含まれる。
    """
    import re as _re
    from app.services.rakuten_rms import fetch_inventory_from_rms

    if not component_sku.strip():
        raise HTTPException(400, "component_sku を指定してください（例: 244）")

    sku_val = component_sku.strip()
    check_p = db.query(RakutenProduct).filter(RakutenProduct.sku == sku_val, RakutenProduct.is_active == True).first()
    if check_p and check_p.set_components and check_p.set_components != "[]":
        raise HTTPException(400, f"'{sku_val}' はセットSKUです。構成品の単品SKUを指定してください")

    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")

    group = _resolve_push_group(sku_val, db)
    if not group:
        raise HTTPException(404, f"SKU '{component_sku}' が見つかりません")

    comp = group["component"]
    sets = group["sets"]
    comp_is_rms = comp["is_rms_sku"]

    fetch_items = []
    if comp_is_rms:
        fetch_items.append({"manage_number": comp["manage_number"], "variant_id": comp["sku"]})
    for s in sets:
        fetch_items.append({"manage_number": s["manage_number"], "variant_id": s["sku"]})

    rms_stock = await fetch_inventory_from_rms(
        settings.rms_service_secret, settings.rms_license_key, fetch_items
    ) if fetch_items else {}

    rms_comp = rms_stock.get(comp["sku"]) if comp_is_rms else None
    comp_result = {
        "sku": comp["sku"],
        "manage_number": comp["manage_number"],
        "name": comp["name"],
        "role": "component (RMS販売SKU)" if comp_is_rms else "component (内部構成品・push対象外)",
        "is_rms_sku": comp_is_rms,
        "db_stock": comp["stock"],
        "push_value": comp["stock"] if comp_is_rms else None,
        "rms_stock": rms_comp,
        "diff": comp["stock"] - rms_comp if comp_is_rms and rms_comp is not None else None,
    }

    push_count = (1 if comp_is_rms else 0) + len(sets)
    set_results = []
    for s in sets:
        rms_qty = rms_stock.get(s["sku"])
        set_results.append({
            "sku": s["sku"],
            "manage_number": s["manage_number"],
            "name": s["name"],
            "role": "set",
            "db_stock": s["db_stock"],
            "push_value": s["calculated_stock"],
            "rms_stock": rms_qty,
            "diff": s["calculated_stock"] - rms_qty if rms_qty is not None else None,
            "components": s["components"],
        })

    return {
        "component": comp_result,
        "sets": set_results,
        "summary": {
            "total_push_targets": push_count,
            "component_pushable": comp_is_rms,
        },
    }


class DebugPushRequest(BaseModel):
    component_sku: str


@router.post("/rms/debug-push-execute")
async def debug_push_execute(
    req: DebugPushRequest,
    db: Session = Depends(get_db),
):
    """単品SKUを指定 → 関連セットを再計算 → 単品+セットをまとめてRMSにpush。
    RMS_PUSH_ENABLEDフラグを無視。push前後の楽天値と時刻を記録。
    """
    import re as _re, base64, httpx, asyncio
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td

    component_sku = req.component_sku.strip()
    if not component_sku:
        raise HTTPException(400, "component_sku を指定してください")

    check_p = db.query(RakutenProduct).filter(RakutenProduct.sku == component_sku, RakutenProduct.is_active == True).first()
    if check_p and check_p.set_components and check_p.set_components != "[]":
        raise HTTPException(400, f"'{component_sku}' はセットSKUです。構成品の単品SKUを指定してください")

    settings = _get_or_create_settings(db)
    if not settings.rms_service_secret or not settings.rms_license_key:
        raise HTTPException(400, "APIキーが設定されていません")

    group = _resolve_push_group(component_sku, db)
    if not group:
        raise HTTPException(404, f"SKU '{component_sku}' が見つかりません")

    comp = group["component"]
    sets = group["sets"]
    comp_is_rms = comp["is_rms_sku"]

    push_items = []
    if comp_is_rms:
        push_items.append({
            "manage_number": comp["manage_number"],
            "variant_id": comp["sku"],
            "quantity": comp["stock"],
            "role": "component",
        })
    for s in sets:
        push_items.append({
            "manage_number": s["manage_number"],
            "variant_id": s["sku"],
            "quantity": s["calculated_stock"],
            "role": "set",
        })

    if not push_items:
        raise HTTPException(400, "push対象がありません（内部構成品で関連セットもなし）")
    if len(push_items) > 30:
        raise HTTPException(400, f"push対象が{len(push_items)}件と多すぎます（上限30）")

    from app.services.rakuten_rms import fetch_inventory_from_rms

    fetch_items = [{"manage_number": i["manage_number"], "variant_id": i["variant_id"]} for i in push_items]
    rms_before = await fetch_inventory_from_rms(
        settings.rms_service_secret, settings.rms_license_key, fetch_items
    )

    jst = _tz(_td(hours=9))
    time_before = _dt.now(jst).strftime("%Y-%m-%d %H:%M:%S")

    token = base64.b64encode(
        f"{settings.rms_service_secret}:{settings.rms_license_key}".encode()
    ).decode()
    headers = {"Authorization": f"ESA {token}", "Content-Type": "application/json"}

    results = []
    async with httpx.AsyncClient(timeout=30) as client:
        for idx, item in enumerate(push_items):
            if idx > 0:
                await asyncio.sleep(0.5)
            mn = item["manage_number"]
            vid = item["variant_id"]
            qty = item["quantity"]
            url = f"https://api.rms.rakuten.co.jp/es/2.0/inventories/manage-numbers/{mn}/variants/{vid}"
            body = json.dumps({"mode": "ABSOLUTE", "quantity": qty}, ensure_ascii=False).encode("utf-8")
            attempts = 0
            last_status = None
            last_detail = None
            ok = False
            for attempt in range(4):
                attempts = attempt + 1
                try:
                    res = await client.put(url, headers=headers, content=body)
                    last_status = res.status_code
                    if res.status_code == 204:
                        ok = True
                        last_detail = None
                        break
                    elif res.status_code == 429:
                        last_detail = "429 Rate Limit"
                        await asyncio.sleep(2 ** attempt)
                    else:
                        last_detail = res.text[:200]
                        break
                except Exception as e:
                    last_detail = str(e)
                    break
            results.append({
                "sku": vid,
                "manage_number": mn,
                "role": item["role"],
                "pushed_qty": qty,
                "rms_before": rms_before.get(vid),
                "http_status": last_status,
                "ok": ok,
                "attempts": attempts,
                "detail": last_detail,
            })

    time_after = _dt.now(jst).strftime("%Y-%m-%d %H:%M:%S")

    rms_after = await fetch_inventory_from_rms(
        settings.rms_service_secret, settings.rms_license_key, fetch_items
    )
    time_verified = _dt.now(jst).strftime("%Y-%m-%d %H:%M:%S")

    for r in results:
        r["rms_after"] = rms_after.get(r["sku"])

    ok_count = sum(1 for r in results if r["ok"])

    try:
        import json as _json_ie
        from app.models.inventory_event import InventoryEvent
        pushed_list = [{"sku": r["sku"], "qty": r["pushed_qty"], "http": r["http_status"], "ok": r["ok"]} for r in results]
        sb = {r["sku"]: r["rms_before"] for r in results if r["rms_before"] is not None}
        sa = {r["sku"]: r["rms_after"] for r in results if r["rms_after"] is not None}
        errs = [{"sku": r["sku"], "detail": r["detail"]} for r in results if r["detail"]]
        ev_time = _dt.now(jst).replace(tzinfo=None)
        db.add(InventoryEvent(
            event_time=ev_time,
            event_type="debug_push",
            pushed=_json_ie.dumps(pushed_list, ensure_ascii=False),
            push_ok=ok_count,
            push_fail=len(results) - ok_count,
            errors=_json_ie.dumps(errs, ensure_ascii=False) if errs else None,
            stock_before=_json_ie.dumps(sb, ensure_ascii=False) if sb else None,
            stock_after=_json_ie.dumps(sa, ensure_ascii=False) if sa else None,
        ))
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    return {
        "ok": ok_count,
        "fail": len(results) - ok_count,
        "time_before_push": time_before,
        "time_after_push": time_after,
        "time_verified": time_verified,
        "results": results,
    }
