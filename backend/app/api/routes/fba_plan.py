from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
import io
import uuid
import time
import threading

from app.core.database import get_db, SessionLocal
from app.models.product import Product
from app.models.settings import OrderSettings
from app.models.order_history import OrderHistory
from app.services.calc import CalcSettings, weighted_daily, growth_mult, calc_sale_extra_days

router = APIRouter(prefix="/fba-plan", tags=["fba-plan"])

_jobs: dict = {}
_jobs_lock = threading.Lock()


def _prune_jobs():
    now = time.time()
    with _jobs_lock:
        for jid in [j for j, v in _jobs.items() if now - v.get("started_at", now) > 1200]:
            _jobs.pop(jid, None)
        if len(_jobs) > 30:
            for jid, _ in sorted(_jobs.items(), key=lambda kv: kv[1].get("started_at", 0))[:-30]:
                _jobs.pop(jid, None)


def _build_settings(row: Optional[OrderSettings]):
    if not row:
        return {
            "lt_order_to_warehouse": 7,
            "lt_shipping_request": 7,
            "lt_sea_to_fba": 18,
            "lt_air_to_fba": 10,
            "free_storage_days": 90,
            "air_threshold_days": 18,
            "hold_daily_threshold": 0.1,
        }
    return {
        "lt_order_to_warehouse": getattr(row, "lt_order_to_warehouse", 7) or 7,
        "lt_shipping_request": getattr(row, "lt_shipping_request", 7) or 7,
        "lt_sea_to_fba": getattr(row, "lt_sea_to_fba", 18) or 18,
        "lt_air_to_fba": getattr(row, "lt_air_to_fba", 10) or 10,
        "free_storage_days": getattr(row, "free_storage_days", 90) or 90,
        "air_threshold_days": getattr(row, "air_threshold_days", 18) or 18,
        "hold_daily_threshold": getattr(row, "hold_daily_threshold", 0.1) or 0.1,
    }


def _build_calc_settings(row: Optional[OrderSettings]) -> CalcSettings:
    if not row:
        return CalcSettings()
    return CalcSettings(
        lead_days=getattr(row, 'lead_days', 75) or 75,
        weight_d7=row.weight_d7,
        weight_d15=row.weight_d15,
        weight_d30=row.weight_d30,
        weight_d60=row.weight_d60,
        weight_d90=getattr(row, 'weight_d90', 0.30) or 0.30,
        growth_ratio_threshold=row.growth_ratio_threshold,
        growth_multiplier=min(row.growth_multiplier, 1.0),
        decline_ratio_threshold=row.decline_ratio_threshold,
        decline_multiplier=max(row.decline_multiplier, 0.5),
        min_order_qty=row.min_order_qty,
        sale_enabled=row.sale_enabled,
        sale_start=row.sale_start,
        sale_end=row.sale_end,
        sale_multiplier=getattr(row, 'sale_multiplier', 3.0) or 3.0,
    )


def _run_plan_job(job_id: str):
    with _jobs_lock:
        _jobs[job_id] = {"status": "running", "result": None, "error": None, "started_at": time.time()}

    db = SessionLocal()
    try:
        products = db.query(Product).filter(Product.is_active == True).all()
        if not products:
            with _jobs_lock:
                _jobs[job_id]["status"] = "done"
                _jobs[job_id]["result"] = {"items": [], "settings": {}}
            return

        settings_row = db.query(OrderSettings).first()
        plan_settings = _build_settings(settings_row)
        cs = _build_calc_settings(settings_row)

        from app.core.config import settings as app_settings
        if app_settings.SP_API_REFRESH_TOKEN:
            from app.services.amazon_api import fetch_inventory, fetch_all_sales
            from concurrent.futures import ThreadPoolExecutor
            asin_list = [p.asin for p in products if p.asin]
            with ThreadPoolExecutor(max_workers=2) as ex:
                f_inv = ex.submit(fetch_inventory)
                f_sales = ex.submit(fetch_all_sales, asin_list, getattr(settings_row, 'order_qty_cap', 0) or 0)
            inventory = f_inv.result()
            sales_7, sales_15, sales_30, sales_60, sales_90 = f_sales.result()
        else:
            inventory = {}
            sales_7 = sales_15 = sales_30 = sales_60 = sales_90 = {}

        from sqlalchemy import func as sqlfunc
        ordered_qty_by_sku = dict(
            db.query(OrderHistory.sku, sqlfunc.sum(OrderHistory.qty))
            .filter(OrderHistory.is_deleted == False)
            .group_by(OrderHistory.sku)
            .all()
        )

        free_days = plan_settings["free_storage_days"]
        lt_sea_total = plan_settings["lt_order_to_warehouse"] + plan_settings["lt_shipping_request"] + plan_settings["lt_sea_to_fba"]
        lt_air_total = plan_settings["lt_order_to_warehouse"] + plan_settings["lt_shipping_request"] + plan_settings["lt_air_to_fba"]
        air_threshold = plan_settings["air_threshold_days"]
        hold_threshold = plan_settings["hold_daily_threshold"]
        target_stock_days = free_days - plan_settings["lt_order_to_warehouse"]
        sale_extra = calc_sale_extra_days(cs)

        result = []
        for p in products:
            inv = inventory.get(p.fnsku, {})
            available = inv.get("available", 0)
            inbound = inv.get("inbound", 0)
            processing = inv.get("processing", 0)
            ordered = ordered_qty_by_sku.get(p.sku, 0)
            extra = p.extra_stock or 0

            s7 = sales_7.get(p.asin, 0)
            s15 = sales_15.get(p.asin, 0)
            s30 = sales_30.get(p.asin, 0)
            s60 = sales_60.get(p.asin, 0)
            s90 = sales_90.get(p.asin, 0)

            daily = weighted_daily(s7, s15, s30, s60, s90, cs)
            g = growth_mult(s7, s15, s90, cs)

            fba_stock = available
            pipeline_stock = available + inbound + processing + ordered + extra
            fba_days = int(fba_stock / daily) if daily > 0 else 9999
            pipeline_days = int(pipeline_stock / daily) if daily > 0 else 9999

            # 配送方法判定
            if daily < hold_threshold:
                ship_method = "hold"
            elif pipeline_days <= air_threshold:
                ship_method = "air"
            else:
                ship_method = "sea"

            # 推奨納品数 = 目標日数 × 日販 × 成長補正 + セール上乗せ − パイプライン在庫
            target = round(daily * g * (target_stock_days + sale_extra))
            recommended = max(0, target - pipeline_stock)
            set_size = max(1, p.set_size or 1)
            recommended_sets = -(-recommended // set_size) if recommended > 0 else 0

            result.append({
                "product_id": p.id,
                "sku": p.sku or "",
                "fnsku": p.fnsku or "",
                "asin": p.asin or "",
                "name": p.name or "",
                "photo_url": p.photo_url or "",
                "color": p.color or "",
                "size": p.size or "",
                "set_size": set_size,
                "price": p.price or 0,
                "category": p.category or "標準",
                "buy_url": p.buy_url or "",
                # 販売データ
                "sales_7": s7,
                "sales_15": s15,
                "sales_30": s30,
                "sales_60": s60,
                "sales_90": s90,
                "daily": round(daily, 2),
                "growth": round(g, 2),
                # 在庫データ
                "fba_available": available,
                "fba_inbound": inbound,
                "fba_processing": processing,
                "ordered": ordered,
                "extra_stock": extra,
                "pipeline_stock": pipeline_stock,
                # 残日数
                "fba_days": fba_days,
                "pipeline_days": pipeline_days,
                # 判定結果
                "ship_method": ship_method,
                "target_stock": target,
                "recommended_sets": recommended_sets,
                "recommended_pieces": recommended_sets * set_size,
                "plan_qty": recommended_sets,
            })

        with _jobs_lock:
            _jobs[job_id]["status"] = "done"
            _jobs[job_id]["result"] = {
                "items": result,
                "settings": plan_settings,
                "sale_extra_days": round(sale_extra, 1),
                "target_stock_days": target_stock_days,
                "lt_sea_total": lt_sea_total,
                "lt_air_total": lt_air_total,
            }

    except Exception as e:
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = str(e)
    finally:
        db.close()


@router.post("/start")
def start_plan(background_tasks: BackgroundTasks, force: bool = False):
    _prune_jobs()
    if force:
        from app.services.amazon_api import _cache
        _cache.clear()
    job_id = str(uuid.uuid4())
    background_tasks.add_task(_run_plan_job, job_id)
    return {"job_id": job_id}


@router.get("/status/{job_id}")
def get_plan_status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return {"status": "not_found"}
    return {
        "status": job["status"],
        "result": job["result"],
        "error": job["error"],
        "elapsed": round(time.time() - job["started_at"], 1),
    }


class PlanExportItem(BaseModel):
    sku: str
    fnsku: str
    plan_qty: int
    set_size: int = 1


class PlanExportRequest(BaseModel):
    items: list[PlanExportItem]


@router.post("/export-excel")
def export_plan_excel(req: PlanExportRequest, db: Session = Depends(get_db)):
    if not req.items:
        return {"error": "エクスポートする商品がありません"}

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "納品プラン"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = "Default prep owner"
    ws["C1"] = "Seller"
    ws.merge_cells("A2:B2")
    ws["A2"] = "Default labeling owner"
    ws["C2"] = "Seller"

    headers = ["Merchant SKU", "Quantity", "Prep owner", "Labeling owner",
               "Units per box", "Number of boxes", "Box length (cm)",
               "Box width (cm)", "Box height (cm)", "Box weight (kg)", "", "FNSKU", "出荷数", "セット数"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_qty = 0
    row = 5
    for item in req.items:
        if item.plan_qty <= 0:
            continue
        pieces = item.plan_qty * item.set_size
        ws.cell(row=row, column=1, value=item.sku).border = border
        ws.cell(row=row, column=2, value=pieces).border = border
        ws.cell(row=row, column=12, value=item.fnsku).border = border
        ws.cell(row=row, column=13, value=pieces).border = border
        ws.cell(row=row, column=14, value=item.set_size).border = border
        total_qty += pieces
        row += 1

    ws.cell(row=row, column=1, value="合計")
    ws.cell(row=row, column=2, value=total_qty)
    ws.cell(row=row, column=2).font = Font(bold=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["N"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fba_plan_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
